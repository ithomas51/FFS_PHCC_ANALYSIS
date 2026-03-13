#!/usr/bin/env python3
"""
centrix_vs_cms_analysis.py  --  Centrix Care OR vs CMS-Only Analysis
=====================================================================

Compare Centrix Care OR proposed rates against CMS 2026 Q1 OR (Medicare)
and OHA FFS Medicaid (Medicaid reference) only -- no PHCC data.

Code universe = Centrix HCPC codes.
For each code, NU (purchase) and RR (rental) rates are looked up
independently.  Blanks where no data exists.

Output: output/centrix_vs_cms_analysis.xlsx  (2 tabs: Summary, Detail)

See: METHODOLOGY_CENTRIX.md
"""

import re, math, sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# =====================================================================
# 0.  PATHS
# =====================================================================
ROOT    = Path(__file__).resolve().parent.parent
DATA    = ROOT / "data"
CMS_DIR = DATA / "cms"
CENTRIX = DATA / "CENTRIX"
OUTPUT  = ROOT / "output"

FILES = {
    "centrix": CENTRIX / "Centrix_Care_OR.csv",
    "cms_or":  CMS_DIR / "CMS_2026_Q1_OR.csv",
    "oha":     CMS_DIR / "OHA_FFS_09_2025_RAW.csv",
    "hcpcs":   CMS_DIR / "2026_CMS_HCPCS.csv",
}

TOLERANCE_PCT = 1.0  # +/-1% treated as no change

# =====================================================================
# 1.  HELPERS
# =====================================================================
VALID_RE = re.compile(r'^[A-Z][0-9]{4}$')


def _norm(raw) -> str:
    s = str(raw).strip().upper()
    return s if VALID_RE.match(s) else ""


def _norm_mod(raw) -> str:
    s = str(raw).strip().upper() if pd.notna(raw) else ""
    return s if s else ""


def _sf(val) -> float:
    """Safe float, stripping $ and commas."""
    try:
        s = str(val).strip().replace("$", "").replace(",", "")
        return float(s) if s else np.nan
    except (ValueError, TypeError):
        return np.nan


def _classify_note(raw_val: str):
    """Parse raw rate text -> (numeric_rate, note_text)."""
    s = str(raw_val).strip()
    if not s:
        return np.nan, ""
    s_clean = s.replace("$", "").replace(",", "")
    try:
        return float(s_clean), ""
    except ValueError:
        return np.nan, s


# =====================================================================
# 2.  DATA LOADERS
# =====================================================================

def load_centrix(path: Path) -> tuple:
    """Load Centrix proposed rates.
    Returns:
        lk   : {hcpcs: {"NU": num, "RR": num, "NU_raw": str, ...}}
        meta : {hcpcs: {"cat": str, "type": str}}
    """
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    meta: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPC", ""))
        if not hcpcs:
            continue
        mod = _norm_mod(r.get("MOD1", ""))
        raw = str(r.get("RATE", "")).strip()
        num, note = _classify_note(raw)

        slot = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {
                "NU": np.nan, "RR": np.nan,
                "NU_raw": "", "RR_raw": "",
                "NU_note": "", "RR_note": "",
            }
        lk[hcpcs][slot]           = num
        lk[hcpcs][f"{slot}_raw"]  = raw
        lk[hcpcs][f"{slot}_note"] = note

        if hcpcs not in meta:
            cat  = str(r.get("CAT", "")).strip()
            typ  = str(r.get("TYPE", "")).strip()
            meta[hcpcs] = {"cat": cat, "type": typ}

    return lk, meta


def load_cms(path: Path) -> dict:
    """Return {hcpcs: {"NU_nr": rate, "RR_nr": rate, "BLANK_nr": rate, ...}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod   = _norm_mod(r.get("Mod", ""))
        if not hcpcs:
            continue
        nr = _sf(r.get("OR (NR)", ""))
        rv = _sf(r.get("OR (R)", ""))

        tag = mod if mod else "BLANK"
        if hcpcs not in lk:
            lk[hcpcs] = {}
        lk[hcpcs][f"{tag}_nr"] = nr
        lk[hcpcs][f"{tag}_r"]  = rv
    return lk


def _cms_rate(lk: dict, hcpcs: str, slot: str) -> float:
    """CMS Non-Rural for code+slot.  Cascade: exact -> blank fallback."""
    rec = lk.get(hcpcs, {})
    val = rec.get(f"{slot}_nr", np.nan)
    if isinstance(val, float) and not math.isnan(val) and val > 0:
        return val
    val_blank = rec.get("BLANK_nr", np.nan)
    if isinstance(val_blank, float) and not math.isnan(val_blank):
        return val_blank
    return np.nan


def load_oha(path: Path) -> dict:
    """Return {hcpcs: {"NU": rate, "RR": rate}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("Procedure Code", ""))
        mod   = _norm_mod(r.get("Mod1", ""))
        if not hcpcs:
            continue
        price = _sf(r.get("Price", ""))
        slot  = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {"NU": np.nan, "RR": np.nan}
        lk[hcpcs][slot] = price
    return lk


def load_hcpcs_desc(path: Path) -> dict:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    return {
        _norm(r.get("HCPC", "")): str(r.get("SHORT DESCRIPTION", "")).strip()
        for _, r in df.iterrows() if _norm(r.get("HCPC", ""))
    }


# =====================================================================
# 3.  BUILD ANALYSIS TABLE
# =====================================================================

def _delta(a: float, b: float):
    if math.isnan(a) or math.isnan(b):
        return np.nan, np.nan
    d = a - b
    p = (d / b * 100) if b != 0 else np.nan
    return d, p


def _flag_cms(proposed: float, cms_nr: float, proposed_note: str) -> str:
    """Decision flag: Centrix proposed vs CMS benchmark."""
    if math.isnan(proposed) and proposed_note:
        return "NON-NUMERIC PROPOSED"
    if math.isnan(proposed):
        return ""
    if math.isnan(cms_nr):
        return "NO CMS RATE"
    d = proposed - cms_nr
    pct = (d / cms_nr * 100) if cms_nr != 0 else 0.0
    if abs(pct) <= TOLERANCE_PCT:
        return "AT CMS"
    if proposed > cms_nr:
        return "ABOVE CMS"
    return "BELOW CMS"


def build_table(
    centrix_lk: dict,
    centrix_meta: dict,
    cms_or: dict,
    oha_lk: dict,
    hcpcs_desc: dict,
) -> pd.DataFrame:
    """One row per Centrix HCPC code with NU + RR comparison to CMS."""

    rows = []
    for hcpcs in sorted(centrix_lk.keys()):
        # -- Centrix proposed --
        cx = centrix_lk[hcpcs]
        cx_nu      = cx.get("NU", np.nan)
        cx_rr      = cx.get("RR", np.nan)
        cx_nu_raw  = cx.get("NU_raw", "")
        cx_rr_raw  = cx.get("RR_raw", "")
        cx_nu_note = cx.get("NU_note", "")
        cx_rr_note = cx.get("RR_note", "")

        cm = centrix_meta.get(hcpcs, {})
        cx_cat  = cm.get("cat", "")
        cx_type = cm.get("type", "")

        in_cms = hcpcs in cms_or

        # -- CMS OR --
        cms_nu = _cms_rate(cms_or, hcpcs, "NU")
        cms_rr = _cms_rate(cms_or, hcpcs, "RR")

        # -- OHA --
        oha = oha_lk.get(hcpcs, {})
        oha_nu = oha.get("NU", np.nan)
        oha_rr = oha.get("RR", np.nan)

        # -- Deltas: Centrix vs CMS --
        d_nu, p_nu = _delta(cx_nu, cms_nu)
        d_rr, p_rr = _delta(cx_rr, cms_rr)

        # -- Flags --
        f_nu = _flag_cms(cx_nu, cms_nu, cx_nu_note)
        f_rr = _flag_cms(cx_rr, cms_rr, cx_rr_note)

        rows.append({
            "HCPC":              hcpcs,
            "Description":       hcpcs_desc.get(hcpcs, ""),
            "Centrix CAT":       cx_cat,
            "Centrix TYPE":      cx_type,
            "Centrix NU":        cx_nu,
            "Centrix NU Raw":    cx_nu_raw if cx_nu_note else "",
            "Centrix RR":        cx_rr,
            "Centrix RR Raw":    cx_rr_raw if cx_rr_note else "",
            "CMS OR NU":         cms_nu,
            "CMS OR RR":         cms_rr,
            "OHA NU":            oha_nu,
            "OHA RR":            oha_rr,
            "In CMS":            "Y" if in_cms else "N",
            "Delta NU":          d_nu,
            "Delta RR":          d_rr,
            "Delta% NU":         p_nu,
            "Delta% RR":         p_rr,
            "Flag NU":           f_nu,
            "Flag RR":           f_rr,
        })

    return pd.DataFrame(rows)


# =====================================================================
# 4.  XLSX FORMATTING
# =====================================================================
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="E2D0F0", end_color="E2D0F0", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(bold=True, size=11)
CURRENCY    = '"$"#,##0.00'
PCT_FMT     = '0.0"%"'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

FLAG_COLORS = {
    "BELOW CMS":            RED_FILL,
    "ABOVE CMS":            BLUE_FILL,
    "AT CMS":               GREEN_FILL,
    "NO CMS RATE":          GRAY_FILL,
    "NON-NUMERIC PROPOSED": PURPLE_FILL,
}

# 0-based column indices for currency and pct formatting
# Columns: 0-HCPC, 1-Desc, 2-CAT, 3-TYPE,
#   4-Centrix NU, 5-Centrix NU Raw, 6-Centrix RR, 7-Centrix RR Raw,
#   8-CMS NU, 9-CMS RR, 10-OHA NU, 11-OHA RR,
#   12-In CMS, 13-Delta NU, 14-Delta RR, 15-Delta% NU, 16-Delta% RR,
#   17-Flag NU, 18-Flag RR
CURRENCY_COLS = {4, 6, 8, 9, 10, 11, 13, 14}
PCT_COLS      = {15, 16}


def _flag_fill(text: str):
    if not text:
        return None
    for kw, fill in FLAG_COLORS.items():
        if kw in text:
            return fill
    return None


def _auto_width(ws, max_w=30):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col[:80]:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(mx + 2, 8), max_w)


def write_tab(ws, df: pd.DataFrame):
    if df.empty:
        return
    cols = list(df.columns)

    # Header
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col_name in enumerate(cols, 1):
            val = row[col_name]
            if isinstance(val, float) and math.isnan(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER

            idx = ci - 1
            if idx in CURRENCY_COLS and val is not None:
                cell.number_format = CURRENCY
            elif idx in PCT_COLS and val is not None:
                cell.number_format = PCT_FMT

        # Flag colouring
        for flag_col_name in ("Flag NU", "Flag RR"):
            if flag_col_name in cols:
                flag_idx = cols.index(flag_col_name) + 1
                flag_val = str(row.get(flag_col_name, "") or "")
                fill = _flag_fill(flag_val)
                if fill:
                    ws.cell(row=ri, column=flag_idx).fill = fill

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def write_summary(ws, stats: dict):
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1,
                    value="Centrix Care OR vs CMS -- Rate Analysis Summary")
    title.font = Font(bold=True, size=14)

    r = 3
    sections = [
        ("Code Universe", [
            ("Total Centrix HCPC codes", stats["centrix_total"]),
            ("In CMS 2026 Q1 OR", stats["in_cms"]),
            ("In OHA Medicaid", stats["in_oha"]),
            ("Not in CMS", stats["not_in_cms"]),
        ]),
        ("Centrix Rate Breakdown", [
            ("Numeric NU rates", stats["cx_nu_numeric"]),
            ("Numeric RR rates", stats["cx_rr_numeric"]),
            ("Non-numeric NU (MSRP etc.)", stats["cx_nu_text"]),
            ("Non-numeric RR (MSRP etc.)", stats["cx_rr_text"]),
        ]),
        ("NU Flag Distribution", stats["flag_nu_items"]),
        ("RR Flag Distribution", stats["flag_rr_items"]),
    ]

    for section_title, items in sections:
        ws.cell(row=r, column=1, value=section_title).font = BOLD_FONT
        r += 1
        for label, val in items:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        r += 1

    _auto_width(ws, max_w=50)


# =====================================================================
# 5.  MAIN
# =====================================================================

def main():
    missing = [k for k, p in FILES.items() if not p.exists()]
    if missing:
        print("ERROR -- missing files:")
        for k in missing:
            print(f"  {k}: {FILES[k]}")
        sys.exit(1)

    OUTPUT.mkdir(exist_ok=True)

    print("Loading data ...")
    hcpcs_desc = load_hcpcs_desc(FILES["hcpcs"])
    cms_or     = load_cms(FILES["cms_or"])
    oha_lk     = load_oha(FILES["oha"])
    centrix_lk, centrix_meta = load_centrix(FILES["centrix"])

    centrix_codes = set(centrix_lk.keys())
    cms_codes     = set(cms_or.keys())
    oha_codes     = set(oha_lk.keys())

    print(f"  Centrix: {len(centrix_codes)} unique codes")
    print(f"  CMS OR:  {len(cms_codes)} codes")
    print(f"  OHA:     {len(oha_codes)} codes")
    print(f"  Centrix in CMS: {len(centrix_codes & cms_codes)}")
    print(f"  Centrix in OHA: {len(centrix_codes & oha_codes)}")
    print(f"  Centrix NOT in CMS: {len(centrix_codes - cms_codes)}")

    # Build table
    print("\nBuilding comparison table ...")
    df = build_table(
        centrix_lk=centrix_lk,
        centrix_meta=centrix_meta,
        cms_or=cms_or,
        oha_lk=oha_lk,
        hcpcs_desc=hcpcs_desc,
    )
    print(f"  -> {len(df)} rows")

    # Centrix rate breakdown
    cx_nu_num = sum(1 for v in centrix_lk.values()
                    if not math.isnan(v.get("NU", np.nan)))
    cx_rr_num = sum(1 for v in centrix_lk.values()
                    if not math.isnan(v.get("RR", np.nan)))
    cx_nu_txt = sum(1 for v in centrix_lk.values()
                    if math.isnan(v.get("NU", np.nan)) and v.get("NU_note", ""))
    cx_rr_txt = sum(1 for v in centrix_lk.values()
                    if math.isnan(v.get("RR", np.nan)) and v.get("RR_note", ""))

    # Flag distributions
    flag_nu = df["Flag NU"].fillna("").value_counts()
    flag_rr = df["Flag RR"].fillna("").value_counts()
    flag_nu_items = [(f, int(c)) for f, c in flag_nu.items() if f]
    flag_rr_items = [(f, int(c)) for f, c in flag_rr.items() if f]

    summary = {
        "centrix_total": len(centrix_codes),
        "in_cms":        len(centrix_codes & cms_codes),
        "in_oha":        len(centrix_codes & oha_codes),
        "not_in_cms":    len(centrix_codes - cms_codes),
        "cx_nu_numeric": cx_nu_num,
        "cx_rr_numeric": cx_rr_num,
        "cx_nu_text":    cx_nu_txt,
        "cx_rr_text":    cx_rr_txt,
        "flag_nu_items": flag_nu_items,
        "flag_rr_items": flag_rr_items,
    }

    # Write XLSX
    out_path = OUTPUT / "centrix_vs_cms_analysis.xlsx"
    print(f"\nWriting {out_path} ...")
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary(ws_summary, summary)

    ws_detail = wb.create_sheet(title="Detail")
    write_tab(ws_detail, df)

    wb.save(out_path)
    print(f"Done -> {out_path}")
    print(f"  Tabs: Summary, Detail")


if __name__ == "__main__":
    main()
