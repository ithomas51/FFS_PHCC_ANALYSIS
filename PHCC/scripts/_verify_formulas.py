"""Verify formulas XLSX: check that formula columns contain XL formulas, not static values."""
from openpyxl import load_workbook

wb = load_workbook("output/integra_rate_analysis_v2_formulas.xlsx")
print(f"Tabs: {wb.sheetnames}\n")

# Check payer tab (Commercial)
ws = wb["Commercial"]
print("=== Commercial tab ===")
print(f"Row 1 (header): {[ws.cell(1, c).value for c in range(1, 19)]}")
print(f"Row 2 (first data):")
for c in range(1, 19):
    v = ws.cell(2, c).value
    h = ws.cell(1, c).value
    print(f"  {h}: {repr(v)}")

# Spot-check formula columns
print(f"\nRow 3 formulas check:")
for col_name in ["Δ Proposed–PHCC", "Δ%", "Proposed–CMS NR", "PHCC–CMS NR"]:
    for c in range(1, 19):
        if ws.cell(1, c).value == col_name:
            v2 = ws.cell(2, c).value
            v3 = ws.cell(3, c).value
            print(f"  {col_name}: row2={repr(v2)}, row3={repr(v3)}")
            break

# Check that detail table starts at row 1
print(f"\nA1 = {repr(ws.cell(1, 1).value)}")
print(f"A2 = {repr(ws.cell(2, 1).value)}")

# Find where summary starts (after data)
last_data = ws.max_row
print(f"Max row: {last_data}")
# Look for "Integra Commercial" summary header
for r in range(1, min(last_data + 1, 3300)):
    v = ws.cell(r, 1).value
    if v and "Rate Analysis Summary" in str(v):
        print(f"Summary starts at row {r}: {v}")
        break

# Check CV OR Contracted
print("\n=== CV OR Contracted ===")
ws2 = wb["CV OR Contracted"]
print(f"Headers: {[ws2.cell(1, c).value for c in range(1, 22)]}")
print(f"Row 2 sample:")
for c in range(1, 22):
    h = ws2.cell(1, c).value
    v = ws2.cell(2, c).value
    if h and ("Δ" in str(h) or "Integra" in str(h)):
        print(f"  {h}: {repr(v)}")

wb.close()
