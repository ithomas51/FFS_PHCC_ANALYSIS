#!/usr/bin/env python3
"""
centrix_cms_analysis.py  --  Centrix Care OR vs CMS / OHA Benchmark Analysis
=============================================================================

Compare Centrix Care OR proposed rates directly against public benchmarks:
  - CMS 2026 Q1 OR (Medicare benchmark)
  - OHA FFS Medicaid (Medicaid reference)

No PHCC contract data is used.  Code universe = Centrix HCPC codes only.
For each code, NU (purchase) and RR (rental) rates are compared
independently.  Blanks where no data exists.

Output: output/centrix_cms_analysis.xlsx
  Tabs: Summary, Centrix Fees, CMS NR Fees, OHA Fees,
        CMS Comparison, OHA Comparison, Delta Flags
"""

import re, math, sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# =====================================================================
# 0.  PATHS
# =====================================================================
ROOT    = Path(__file__).resolve().parent.parent
DATA    = ROOT / "data"
CMS_DIR = DATA / "cms"
CENTRIX = DATA / "CENTRIX"
OUTPUT  = ROOT / "output"

FILES = {
    "centrix": CENTRIX / "Centrix_Care_OR.csv",
    "cms_or":  CMS_DIR / "CMS_2026_Q1_OR.csv",
    "oha":     CMS_DIR / "OHA_FFS_09_2025_RAW.csv",
    "hcpcs":   CMS_DIR / "2026_CMS_HCPCS.csv",
}

TOLERANCE_PCT = 1.0  # +/-1% treated as no change

# =====================================================================
# 1.  HELPERS
# =====================================================================
VALID_RE = re.compile(r'^[A-Z][0-9]{4}$')


def _norm(raw) -> str:
    s = str(raw).strip().upper()
    return s if VALID_RE.match(s) else ""


def _norm_mod(raw) -> str:
    s = str(raw).strip().upper() if pd.notna(raw) else ""
    return s if s else ""


def _sf(val) -> float:
    """Safe float, stripping $ and commas."""
    try:
        s = str(val).strip().replace("$", "").replace(",", "")
        return float(s) if s else np.nan
    except (ValueError, TypeError):
        return np.nan


def _classify_note(raw_val: str):
    """Parse raw rate text -> (numeric_rate, note_text)."""
    s = str(raw_val).strip()
    if not s:
        return np.nan, ""
    s_clean = s.replace("$", "").replace(",", "")
    try:
        return float(s_clean), ""
    except ValueError:
        return np.nan, s


# =====================================================================
# 2.  DATA LOADERS
# =====================================================================

def load_centrix(path: Path) -> tuple:
    """Load Centrix proposed rates.
    Returns:
        lk   : {hcpcs: {"NU": num, "RR": num, "NU_raw": str, ...}}
        meta : {hcpcs: {"cat": str, "type": str}}
    """
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    meta: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPC", ""))
        if not hcpcs:
            continue
        mod = _norm_mod(r.get("MOD1", ""))
        raw = str(r.get("RATE", "")).strip()
        num, note = _classify_note(raw)

        slot = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {
                "NU": np.nan, "RR": np.nan,
                "NU_raw": "", "RR_raw": "",
                "NU_note": "", "RR_note": "",
            }
        lk[hcpcs][slot]           = num
        lk[hcpcs][f"{slot}_raw"]  = raw
        lk[hcpcs][f"{slot}_note"] = note

        if hcpcs not in meta:
            cat  = str(r.get("CAT", "")).strip()
            typ  = str(r.get("TYPE", "")).strip()
            meta[hcpcs] = {"cat": cat, "type": typ}

    return lk, meta


def load_cms(path: Path) -> dict:
    """Return {hcpcs: {"NU_nr": rate, "RR_nr": rate, "BLANK_nr": rate, ...}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod   = _norm_mod(r.get("Mod", ""))
        if not hcpcs:
            continue
        nr = _sf(r.get("OR (NR)", ""))
        rv = _sf(r.get("OR (R)", ""))

        tag = mod if mod else "BLANK"
        if hcpcs not in lk:
            lk[hcpcs] = {}
        lk[hcpcs][f"{tag}_nr"] = nr
        lk[hcpcs][f"{tag}_r"]  = rv
    return lk


def _cms_rate(lk: dict, hcpcs: str, slot: str) -> float:
    """CMS Non-Rural for code+slot.  Cascade: exact -> blank fallback."""
    rec = lk.get(hcpcs, {})
    val = rec.get(f"{slot}_nr", np.nan)
    if isinstance(val, float) and not math.isnan(val) and val > 0:
        return val
    val_blank = rec.get("BLANK_nr", np.nan)
    if isinstance(val_blank, float) and not math.isnan(val_blank):
        return val_blank
    val_raw = rec.get(f"{slot}_nr", np.nan)
    if isinstance(val_raw, float) and not math.isnan(val_raw):
        return val_raw
    return np.nan


def _cms_rural(lk: dict, hcpcs: str, slot: str) -> float:
    """CMS Rural for code+slot.  Cascade: exact -> blank fallback."""
    rec = lk.get(hcpcs, {})
    val = rec.get(f"{slot}_r", np.nan)
    if isinstance(val, float) and not math.isnan(val) and val > 0:
        return val
    val_blank = rec.get("BLANK_r", np.nan)
    if isinstance(val_blank, float) and not math.isnan(val_blank):
        return val_blank
    return np.nan


def load_oha(path: Path) -> dict:
    """Return {hcpcs: {"NU": rate, "RR": rate}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("Procedure Code", ""))
        mod   = _norm_mod(r.get("Mod1", ""))
        if not hcpcs:
            continue
        price = _sf(r.get("Price", ""))
        slot  = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {"NU": np.nan, "RR": np.nan}
        lk[hcpcs][slot] = price
    return lk


def load_hcpcs_desc(path: Path) -> dict:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    return {
        _norm(r.get("HCPC", "")): str(r.get("SHORT DESCRIPTION", "")).strip()
        for _, r in df.iterrows() if _norm(r.get("HCPC", ""))
    }


# =====================================================================
# 3.  BUILD SPLIT TABLES  (one per sheet)
# =====================================================================

def _delta(a: float, b: float):
    if math.isnan(a) or math.isnan(b):
        return np.nan, np.nan
    d = a - b
    p = (d / b * 100) if b != 0 else np.nan
    return d, p


def _flag(proposed: float, benchmark: float, proposed_note: str) -> str:
    if math.isnan(proposed) and proposed_note:
        return "NON-NUMERIC"
    if math.isnan(proposed):
        return ""
    if math.isnan(benchmark):
        return "NO BENCHMARK"
    d = proposed - benchmark
    pct = (d / benchmark * 100) if benchmark != 0 else 0.0
    if abs(pct) <= TOLERANCE_PCT:
        return "AT BENCHMARK"
    if proposed > benchmark:
        return "ABOVE"
    return "BELOW"


def build_all_tables(
    centrix_lk, centrix_meta, cms_or, oha_lk, hcpcs_desc,
) -> dict:
    """Return dict of sheet_name -> DataFrame."""

    centrix_rows = []
    cms_rows = []
    oha_rows = []
    cms_cmp_rows = []
    oha_cmp_rows = []
    flag_rows = []

    for hcpcs in sorted(centrix_lk.keys()):
        cx   = centrix_lk[hcpcs]
        cm   = centrix_meta.get(hcpcs, {})
        desc = hcpcs_desc.get(hcpcs, "")

        cx_nu      = cx.get("NU", np.nan)
        cx_rr      = cx.get("RR", np.nan)
        cx_nu_note = cx.get("NU_note", "")
        cx_rr_note = cx.get("RR_note", "")

        # -- Centrix Fees --
        centrix_rows.append({
            "HCPC":        hcpcs,
            "Description": desc,
            "CAT":         cm.get("cat", ""),
            "TYPE":        cm.get("type", ""),
            "NU Rate":     cx_nu,
            "RR Rate":     cx_rr,
            "NU Note":     cx_nu_note,
            "RR Note":     cx_rr_note,
        })

        # -- CMS NR Fees --
        cms_nu_nr = _cms_rate(cms_or, hcpcs, "NU")
        cms_rr_nr = _cms_rate(cms_or, hcpcs, "RR")
        cms_nu_r  = _cms_rural(cms_or, hcpcs, "NU")
        cms_rr_r  = _cms_rural(cms_or, hcpcs, "RR")
        cms_rows.append({
            "HCPC":        hcpcs,
            "Description": desc,
            "NR NU":       cms_nu_nr,
            "NR RR":       cms_rr_nr,
            "Rural NU":    cms_nu_r,
            "Rural RR":    cms_rr_r,
        })

        # -- OHA Fees --
        oha = oha_lk.get(hcpcs, {})
        oha_nu = oha.get("NU", np.nan)
        oha_rr = oha.get("RR", np.nan)
        oha_rows.append({
            "HCPC":        hcpcs,
            "Description": desc,
            "OHA NU":      oha_nu,
            "OHA RR":      oha_rr,
        })

        # -- CMS Comparison --
        d_cms_nu, p_cms_nu = _delta(cx_nu, cms_nu_nr)
        d_cms_rr, p_cms_rr = _delta(cx_rr, cms_rr_nr)
        f_cms_nu = _flag(cx_nu, cms_nu_nr, cx_nu_note)
        f_cms_rr = _flag(cx_rr, cms_rr_nr, cx_rr_note)
        cms_cmp_rows.append({
            "HCPC":        hcpcs,
            "Description": desc,
            "Centrix NU":  cx_nu,
            "CMS NR NU":   cms_nu_nr,
            "Delta $":     d_cms_nu,
            "Delta %":     p_cms_nu,
            "Flag NU":     f_cms_nu,
            "Centrix RR":  cx_rr,
            "CMS NR RR":   cms_rr_nr,
            "Delta $ RR":  d_cms_rr,
            "Delta % RR":  p_cms_rr,
            "Flag RR":     f_cms_rr,
        })

        # -- OHA Comparison --
        d_oha_nu, p_oha_nu = _delta(cx_nu, oha_nu)
        d_oha_rr, p_oha_rr = _delta(cx_rr, oha_rr)
        f_oha_nu = _flag(cx_nu, oha_nu, cx_nu_note)
        f_oha_rr = _flag(cx_rr, oha_rr, cx_rr_note)
        oha_cmp_rows.append({
            "HCPC":        hcpcs,
            "Description": desc,
            "Centrix NU":  cx_nu,
            "OHA NU":      oha_nu,
            "Delta $":     d_oha_nu,
            "Delta %":     p_oha_nu,
            "Flag NU":     f_oha_nu,
            "Centrix RR":  cx_rr,
            "OHA RR":      oha_rr,
            "Delta $ RR":  d_oha_rr,
            "Delta % RR":  p_oha_rr,
            "Flag RR":     f_oha_rr,
        })

        # -- Delta Flags (combined) --
        in_cms = hcpcs in cms_or
        in_oha_src = hcpcs in oha_lk
        if in_cms and in_oha_src:
            bench = "CMS+OHA"
        elif in_cms:
            bench = "CMS"
        elif in_oha_src:
            bench = "OHA"
        else:
            bench = "NONE"

        flag_rows.append({
            "HCPC":          hcpcs,
            "Description":   desc,
            "Benchmark":     bench,
            "CMS Flag NU":   f_cms_nu,
            "CMS Flag RR":   f_cms_rr,
            "OHA Flag NU":   f_oha_nu,
            "OHA Flag RR":   f_oha_rr,
        })

    return {
        "Centrix Fees":    pd.DataFrame(centrix_rows),
        "CMS NR Fees":     pd.DataFrame(cms_rows),
        "OHA Fees":        pd.DataFrame(oha_rows),
        "CMS Comparison":  pd.DataFrame(cms_cmp_rows),
        "OHA Comparison":  pd.DataFrame(oha_cmp_rows),
        "Delta Flags":     pd.DataFrame(flag_rows),
    }


# =====================================================================
# 4.  XLSX FORMATTING  (column-name driven, not positional)
# =====================================================================
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="E2D0F0", end_color="E2D0F0", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(bold=True, size=11)
CURRENCY    = '"$"#,##0.00'
PCT_FMT     = '0.0"%"'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

FLAG_COLORS = {
    "BELOW":         RED_FILL,
    "ABOVE":         BLUE_FILL,
    "AT BENCHMARK":  GREEN_FILL,
    "NO BENCHMARK":  GRAY_FILL,
    "NON-NUMERIC":   PURPLE_FILL,
}

# Column-name substrings that determine formatting
_CURRENCY_KEYS = ("Rate", "NU", "RR", "Delta $")
_PCT_KEYS      = ("Delta %",)
_FLAG_KEYS     = ("Flag",)


def _is_currency_col(name: str) -> bool:
    if "%" in name or "Flag" in name or "Note" in name:
        return False
    return any(k in name for k in _CURRENCY_KEYS)


def _is_pct_col(name: str) -> bool:
    return any(k in name for k in _PCT_KEYS)


def _is_flag_col(name: str) -> bool:
    return any(k in name for k in _FLAG_KEYS)


def _flag_fill(text: str):
    if not text:
        return None
    for kw, fill in FLAG_COLORS.items():
        if kw in text:
            return fill
    return None


def _auto_width(ws, max_w=30):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col[:80]:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(mx + 2, 8), max_w)


def write_tab(ws, df: pd.DataFrame):
    if df.empty:
        return
    cols = list(df.columns)
    cur_set  = {i for i, c in enumerate(cols) if _is_currency_col(c)}
    pct_set  = {i for i, c in enumerate(cols) if _is_pct_col(c)}
    flag_set = {i for i, c in enumerate(cols) if _is_flag_col(c)}

    # Header
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col_name in enumerate(cols, 1):
            val = row[col_name]
            if isinstance(val, float) and math.isnan(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER

            idx = ci - 1
            if idx in cur_set and val is not None:
                cell.number_format = CURRENCY
            elif idx in pct_set and val is not None:
                cell.number_format = PCT_FMT

            if idx in flag_set:
                fill = _flag_fill(str(val or ""))
                if fill:
                    cell.fill = fill

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def write_summary(ws, stats: dict):
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1,
                    value="Centrix Care OR -- CMS / OHA Benchmark Analysis Summary")
    title.font = Font(bold=True, size=14)

    r = 3
    sections = [
        ("Code Universe", [
            ("Total Centrix HCPC codes", stats["centrix"]),
            ("In CMS OR", stats["in_cms"]),
            ("In OHA", stats["in_oha"]),
            ("In Both CMS+OHA", stats["in_both"]),
            ("No Benchmark", stats["no_bench"]),
        ]),
        ("Centrix Rate Breakdown", [
            ("Numeric NU rates", stats["cx_nu_numeric"]),
            ("Numeric RR rates", stats["cx_rr_numeric"]),
            ("Non-numeric NU (MSRP etc.)", stats["cx_nu_text"]),
            ("Non-numeric RR (MSRP etc.)", stats["cx_rr_text"]),
        ]),
    ]

    for section_title, items in sections:
        ws.cell(row=r, column=1, value=section_title).font = BOLD_FONT
        r += 1
        for label, val in items:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        r += 1

    # Flag distribution
    for label, dist in stats.get("flags", {}).items():
        ws.cell(row=r, column=1, value=f"-- {label} Flag Distribution --").font = BOLD_FONT
        r += 1
        for flag_text, cnt in dist:
            ws.cell(row=r, column=1, value=f"  {flag_text}")
            ws.cell(row=r, column=2, value=cnt)
            r += 1
        r += 1

    _auto_width(ws, max_w=50)


# =====================================================================
# 5.  MAIN
# =====================================================================

def main():
    missing = [k for k, p in FILES.items() if not p.exists()]
    if missing:
        print("ERROR -- missing files:")
        for k in missing:
            print(f"  {k}: {FILES[k]}")
        sys.exit(1)

    OUTPUT.mkdir(exist_ok=True)

    print("=" * 60)
    print("Centrix Care OR -- CMS / OHA Benchmark Analysis")
    print("=" * 60)

    print("\nLoading data ...")
    hcpcs_desc = load_hcpcs_desc(FILES["hcpcs"])
    cms_or     = load_cms(FILES["cms_or"])
    oha_lk     = load_oha(FILES["oha"])
    centrix_lk, centrix_meta = load_centrix(FILES["centrix"])
    print(f"  Centrix: {len(centrix_lk)} unique codes")
    print(f"  CMS OR:  {len(cms_or)} codes")
    print(f"  OHA:     {len(oha_lk)} codes")
    print(f"  HCPCS descriptions: {len(hcpcs_desc)}")

    centrix_codes = set(centrix_lk.keys())
    cms_codes     = set(cms_or.keys())
    oha_codes     = set(oha_lk.keys())
    in_cms  = len(centrix_codes & cms_codes)
    in_oha  = len(centrix_codes & oha_codes)
    in_both = len(centrix_codes & cms_codes & oha_codes)
    no_bench = len(centrix_codes - cms_codes - oha_codes)

    print(f"\nCentrix codes: {len(centrix_codes)}")
    print(f"  In CMS:        {in_cms}")
    print(f"  In OHA:        {in_oha}")
    print(f"  In Both:       {in_both}")
    print(f"  No Benchmark:  {no_bench}")

    cx_nu_num = sum(1 for v in centrix_lk.values()
                    if not math.isnan(v.get("NU", np.nan)))
    cx_rr_num = sum(1 for v in centrix_lk.values()
                    if not math.isnan(v.get("RR", np.nan)))
    cx_nu_txt = sum(1 for v in centrix_lk.values()
                    if math.isnan(v.get("NU", np.nan)) and v.get("NU_note", ""))
    cx_rr_txt = sum(1 for v in centrix_lk.values()
                    if math.isnan(v.get("RR", np.nan)) and v.get("RR_note", ""))

    print("\nBuilding tables ...")
    sheets = build_all_tables(
        centrix_lk, centrix_meta, cms_or, oha_lk, hcpcs_desc,
    )
    for name, sdf in sheets.items():
        print(f"  {name}: {len(sdf)} rows")

    # Flag distributions (from CMS Comparison sheet)
    cms_cmp = sheets["CMS Comparison"]
    flag_nu = cms_cmp["Flag NU"].fillna("").value_counts()
    flag_rr = cms_cmp["Flag RR"].fillna("").value_counts()
    flag_dist = {
        "CMS NU": [(f, int(c)) for f, c in flag_nu.items() if f],
        "CMS RR": [(f, int(c)) for f, c in flag_rr.items() if f],
    }
    oha_cmp = sheets["OHA Comparison"]
    oha_fnu = oha_cmp["Flag NU"].fillna("").value_counts()
    oha_frr = oha_cmp["Flag RR"].fillna("").value_counts()
    flag_dist["OHA NU"] = [(f, int(c)) for f, c in oha_fnu.items() if f]
    flag_dist["OHA RR"] = [(f, int(c)) for f, c in oha_frr.items() if f]

    for slot, items in flag_dist.items():
        print(f"  {slot} flags: {', '.join(f'{f}={c}' for f, c in items)}")

    summary = {
        "centrix":       len(centrix_codes),
        "in_cms":        in_cms,
        "in_oha":        in_oha,
        "in_both":       in_both,
        "no_bench":      no_bench,
        "cx_nu_numeric": cx_nu_num,
        "cx_rr_numeric": cx_rr_num,
        "cx_nu_text":    cx_nu_txt,
        "cx_rr_text":    cx_rr_txt,
        "flags":         flag_dist,
    }

    # Write XLSX
    out_path = OUTPUT / "centrix_cms_analysis.xlsx"
    print(f"\nWriting {out_path.name} ...")
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary(ws_summary, summary)

    for sheet_name, sdf in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        write_tab(ws, sdf)

    wb.save(out_path)
    n_codes = len(sheets["Centrix Fees"])
    print(f"\nDone -> {out_path}")
    print(f"  Tabs: {', '.join(wb.sheetnames)}")
    print(f"  {n_codes} codes analysed across {len(sheets)} sheets")


if __name__ == "__main__":
    main()
