#!/usr/bin/env python3
"""
unified_code_analysis.py  —  Code-Centric Fee Schedule Comparison (v3)
======================================================================

Build a UNION of all unique HCPCS codes from Integra PHP FFS files and PHCC
cleaned schedule files.  For each code, look up NU (purchase) and RR (rental)
rates independently from every source.  If a code is not found in a source,
its rates are left blank — no imputed calculations.

Sources:
  - Integra PHP FFS: 4 payer files (Commercial, ASO, Medicare, Medicaid)
  - PHCC Cleaned: 3 schedules (OR Contracted, OR Participating, WA Participating)
  - CMS 2026 Q1: OR + WA fee schedules (Medicare benchmark)
  - OHA Medicaid: Oregon fee-for-service (Medicaid benchmark)
  - CMS HCPCS 2026: Code descriptions

Output: output/unified_code_analysis.xlsx  (5 tabs)

Prerequisite: python scripts/clean_phcc_files.py
See: METHODOLOGY_v3.md for full design documentation.
"""

import re, math, sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════
# 0.  PATHS & CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════
ROOT    = Path(__file__).resolve().parent.parent
CLEANED = ROOT / "data" / "cleaned"
INTEGRA = ROOT / "data" / "INTEGRA_PHP_FFS"
CMS_DIR = ROOT / "data" / "cms"
OUTPUT  = ROOT / "output"

FILES = {
    "or_contracted":      CLEANED / "PHCC_OR_CONTRACTED_CLEAN.csv",
    "or_participating":   CLEANED / "PHCC_OR_PARTICIPATING_CLEAN.csv",
    "wa_participating":   CLEANED / "PHCC_WA_PARTICIPATING_CLEAN.csv",
    "integra_commercial": INTEGRA / "Integra_PHP_CARVEOUTS_COMMERCIAL.csv",
    "integra_aso":        INTEGRA / "Integra_PHP_CARVEOUTS_ASO.csv",
    "integra_medicare":   INTEGRA / "Integra_PHP_CARVEOUTS_MEDICARE.csv",
    "integra_medicaid":   INTEGRA / "INTEGRA_PHP_CARVEOUTS_MEDICAID.csv",
    "cms_or":             CMS_DIR / "CMS_2026_Q1_OR.csv",
    "cms_wa":             CMS_DIR / "CMS_2026_Q1_WA.csv",
    "oha":                CMS_DIR / "OHA_FFS_09_2025_RAW.csv",
    "hcpcs":              CMS_DIR / "2026_CMS_HCPCS.csv",
}

# Per-payer: which Integra file, rate column, and OR_CONTRACTED rate prefix
PAYER_CFG = {
    "Commercial": {
        "integra_key": "integra_commercial",
        "rate_col":    "Commercial",
        "or_c_prefix": "Commercial",
    },
    "ASO": {
        "integra_key": "integra_aso",
        "rate_col":    "ASO/Commercial",
        "or_c_prefix": "Commercial",
    },
    "Medicare": {
        "integra_key": "integra_medicare",
        "rate_col":    "Medicare",
        "or_c_prefix": "Managed",
    },
    "Medicaid": {
        "integra_key": "integra_medicaid",
        "rate_col":    "Medicaid",
        "or_c_prefix": "Managed",
    },
}

TOLERANCE_PCT = 1.0  # ±1 % treated as "no change"

# ═══════════════════════════════════════════════════════════════════════
# 1.  HELPERS
# ═══════════════════════════════════════════════════════════════════════
VALID_RE = re.compile(r'^[A-Z][0-9]{4}$')
MEDICARE_PCT_RE = re.compile(
    r'Medicare\s+Allowable\s+less\s+(\d+)\s*%', re.IGNORECASE)


def _norm(raw) -> str:
    """Normalise an HCPCS code string."""
    s = str(raw).strip().upper()
    return s if VALID_RE.match(s) else ""


def _sf(val) -> float:
    """Safe float conversion, stripping $ and ,."""
    try:
        s = str(val).strip().replace("$", "").replace(",", "")
        return float(s) if s else np.nan
    except (ValueError, TypeError):
        return np.nan


def _norm_mod(raw) -> str:
    s = str(raw).strip().upper() if pd.notna(raw) else ""
    return s if s else ""


def _classify_note(raw_val: str) -> tuple:
    """Parse raw rate text → (numeric_rate, note_text)."""
    s = str(raw_val).strip()
    if not s:
        return np.nan, ""
    s_clean = s.replace("$", "").replace(",", "")
    try:
        return float(s_clean), ""
    except ValueError:
        return np.nan, s


def _resolve_pct_of_medicare(note_detail: str, cms_nr: float) -> float:
    m = MEDICARE_PCT_RE.search(str(note_detail))
    if m and not math.isnan(cms_nr):
        pct = int(m.group(1))
        return round(cms_nr * (1 - pct / 100), 2)
    return np.nan


# ═══════════════════════════════════════════════════════════════════════
# 2.  DATA LOADERS
# ═══════════════════════════════════════════════════════════════════════

def load_integra(path: Path, rate_col: str) -> dict:
    """Return {hcpcs: {"NU": rate, "RR": rate, "NU_raw": ..., ...}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod = _norm_mod(r.get("Mod 1", ""))
        if not hcpcs:
            continue
        raw = str(r.get(rate_col, "")).strip()
        num, note = _classify_note(raw)

        slot = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {
                "NU": np.nan, "RR": np.nan,
                "NU_raw": "", "RR_raw": "",
                "NU_note": "", "RR_note": "",
            }
        lk[hcpcs][slot]           = num
        lk[hcpcs][f"{slot}_raw"]  = raw
        lk[hcpcs][f"{slot}_note"] = note
    return lk


def load_phcc(path: Path, schedule_key: str, or_c_prefix: str = "") -> dict:
    """Return {hcpcs: {"NU": rate, "RR": rate, "NU_note_type": ..., ...}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)

    lk: dict = {}
    for _, r in df.iterrows():
        if str(r.get("hcpcs_is_valid", "")).strip() != "True":
            continue
        hcpcs = str(r.get("hcpcs_normalised", "")).strip()
        mod = str(r.get("modifier_normalised", "")).strip()
        if not hcpcs:
            continue

        # Determine rate column based on modifier and schedule
        if schedule_key == "or_contracted":
            prefix = or_c_prefix
            if mod == "RR":
                rate_base = f"{prefix} Rental Rate"
            else:
                rate_base = f"{prefix} Purchase Rate"
        else:
            if mod == "RR":
                rate_base = "Rental Rate"
            else:
                rate_base = "Purchase Rate"

        num       = _sf(r.get(f"{rate_base}_numeric", ""))
        raw       = str(r.get(f"{rate_base}_raw", "")).strip()
        note_type = str(r.get(f"{rate_base}_note_type", "")).strip()
        note_det  = str(r.get(f"{rate_base}_note_detail", "")).strip()

        slot = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {
                "NU": np.nan, "RR": np.nan,
                "NU_raw": "", "RR_raw": "",
                "NU_note_type": "", "RR_note_type": "",
                "NU_note_detail": "", "RR_note_detail": "",
            }
        lk[hcpcs][slot]                = num
        lk[hcpcs][f"{slot}_raw"]       = raw
        lk[hcpcs][f"{slot}_note_type"] = note_type
        lk[hcpcs][f"{slot}_note_detail"] = note_det
    return lk


def load_cms(path: Path, nr_col: str, r_col: str) -> dict:
    """Return {hcpcs: {"NU_nr": rate, "NU_r": rate, "RR_nr": ..., "BLANK_nr": ...}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod = _norm_mod(r.get("Mod", ""))
        if not hcpcs:
            continue
        nr = _sf(r.get(nr_col, ""))
        rv = _sf(r.get(r_col, ""))

        if hcpcs not in lk:
            lk[hcpcs] = {}

        tag = mod if mod else "BLANK"
        lk[hcpcs][f"{tag}_nr"] = nr
        lk[hcpcs][f"{tag}_r"]  = rv
    return lk


def _cms_rate(lk: dict, hcpcs: str, slot: str) -> float:
    """Get CMS Non-Rural rate for code+slot.  Cascade: exact → blank."""
    rec = lk.get(hcpcs, {})
    val = rec.get(f"{slot}_nr", np.nan)
    if isinstance(val, float) and (not math.isnan(val)) and val > 0:
        return val
    val_blank = rec.get("BLANK_nr", np.nan)
    if isinstance(val_blank, float) and not math.isnan(val_blank):
        return val_blank
    # Return the slot value even if 0
    val_raw = rec.get(f"{slot}_nr", np.nan)
    if isinstance(val_raw, float) and not math.isnan(val_raw):
        return val_raw
    return np.nan


def load_oha(path: Path) -> dict:
    """Return {hcpcs: {"NU": rate, "RR": rate}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("Procedure Code", ""))
        mod = _norm_mod(r.get("Mod1", ""))
        if not hcpcs:
            continue
        price = _sf(r.get("Price", ""))
        slot = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {"NU": np.nan, "RR": np.nan}
        lk[hcpcs][slot] = price
    return lk


def load_hcpcs_desc(path: Path) -> dict:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    return {
        _norm(r.get("HCPC", "")): str(r.get("SHORT DESCRIPTION", "")).strip()
        for _, r in df.iterrows() if _norm(r.get("HCPC", ""))
    }


# ═══════════════════════════════════════════════════════════════════════
# 3.  BUILD PER-PAYER TABLE
# ═══════════════════════════════════════════════════════════════════════

def _delta(a: float, b: float):
    """Return (delta, pct) or (NaN, NaN)."""
    if math.isnan(a) or math.isnan(b):
        return np.nan, np.nan
    d = a - b
    p = (d / b * 100) if b != 0 else np.nan
    return d, p


def _flag(proposed: float, current: float, cms_nr: float,
          in_phcc: bool, is_phcc_only: bool) -> str:
    """Compute decision flag for one modifier slot."""
    if is_phcc_only:
        return "PHCC ONLY" if not math.isnan(current) else ""
    if math.isnan(proposed):
        return ""
    if not in_phcc:
        return "NEW CODE"
    if math.isnan(current):
        return "PHCC NON-NUMERIC"

    d = proposed - current
    pct = (d / current * 100) if current != 0 else 0.0
    if abs(pct) <= TOLERANCE_PCT:
        return "NO CHANGE"
    if proposed > current:
        return "RATE INCREASE"
    # proposed < current
    if not math.isnan(cms_nr):
        return "BELOW CURRENT" if proposed >= cms_nr else "BELOW CMS FLOOR"
    return "BELOW CURRENT"


def _systemic(current: float, cms_nr: float, flag: str) -> str:
    """Append PHCC BELOW CMS if current < CMS."""
    if (not math.isnan(current) and not math.isnan(cms_nr)
            and current < cms_nr):
        return f"{flag} | PHCC BELOW CMS" if flag else "PHCC BELOW CMS"
    return flag


def build_payer_table(
    payer: str,
    universe: set,
    integra_lk: dict,
    phcc_or_c: dict,
    phcc_or_p: dict,
    phcc_wa_p: dict,
    cms_or: dict,
    cms_wa: dict,
    oha_lk: dict,
    hcpcs_desc: dict,
) -> pd.DataFrame:
    """One row per unique HCPCS, with NU and RR columns."""

    rows = []
    for hcpcs in sorted(universe):
        # ── Source classification ──
        in_integra = hcpcs in integra_lk
        in_phcc    = (hcpcs in phcc_or_c
                      or hcpcs in phcc_or_p
                      or hcpcs in phcc_wa_p)
        is_phcc_only = (not in_integra) and in_phcc

        if in_integra and in_phcc:
            source = "BOTH"
        elif in_integra:
            source = "INTEGRA_ONLY"
        else:
            source = "PHCC_ONLY"

        # ── Integra rates ──
        ig = integra_lk.get(hcpcs, {})
        int_nu = ig.get("NU", np.nan)
        int_rr = ig.get("RR", np.nan)
        int_note = ig.get("NU_note", "") or ig.get("RR_note", "")

        # ── PHCC rates per schedule ──
        orc = phcc_or_c.get(hcpcs, {})
        orp = phcc_or_p.get(hcpcs, {})
        wap = phcc_wa_p.get(hcpcs, {})

        or_c_nu = orc.get("NU", np.nan)
        or_c_rr = orc.get("RR", np.nan)
        or_p_nu = orp.get("NU", np.nan)
        or_p_rr = orp.get("RR", np.nan)
        wa_p_nu = wap.get("NU", np.nan)
        wa_p_rr = wap.get("RR", np.nan)

        # ── Resolve PERCENT_OF_MEDICARE_ALLOWABLE ──
        for slot, cms_dict in [("NU", cms_or), ("RR", cms_or)]:
            for tag, sd in [("or_c", orc), ("or_p", orp)]:
                if sd.get(f"{slot}_note_type") == "PERCENT_OF_MEDICARE_ALLOWABLE":
                    cms_val = _cms_rate(cms_dict, hcpcs, slot)
                    resolved = _resolve_pct_of_medicare(
                        sd.get(f"{slot}_note_detail", ""), cms_val)
                    if not math.isnan(resolved):
                        if tag == "or_c":
                            if slot == "NU":
                                or_c_nu = resolved
                            else:
                                or_c_rr = resolved
                        else:
                            if slot == "NU":
                                or_p_nu = resolved
                            else:
                                or_p_rr = resolved

        for slot in ("NU", "RR"):
            if wap.get(f"{slot}_note_type") == "PERCENT_OF_MEDICARE_ALLOWABLE":
                cms_val = _cms_rate(cms_wa, hcpcs, slot)
                resolved = _resolve_pct_of_medicare(
                    wap.get(f"{slot}_note_detail", ""), cms_val)
                if not math.isnan(resolved):
                    if slot == "NU":
                        wa_p_nu = resolved
                    else:
                        wa_p_rr = resolved

        # ── CMS rates (NU and RR independently) ──
        cms_or_nu = _cms_rate(cms_or, hcpcs, "NU")
        cms_or_rr = _cms_rate(cms_or, hcpcs, "RR")
        cms_wa_nu = _cms_rate(cms_wa, hcpcs, "NU")
        cms_wa_rr = _cms_rate(cms_wa, hcpcs, "RR")

        # ── OHA rates ──
        oha = oha_lk.get(hcpcs, {})
        oha_nu = oha.get("NU", np.nan)
        oha_rr = oha.get("RR", np.nan)

        # ── Primary PHCC for delta (OR_C → OR_P → WA_P) ──
        phcc_pri_nu, src_nu = np.nan, ""
        for rate, src in [(or_c_nu, "OR_Contracted"),
                          (or_p_nu, "OR_Participating"),
                          (wa_p_nu, "WA_Participating")]:
            if not math.isnan(rate):
                phcc_pri_nu, src_nu = rate, src
                break

        phcc_pri_rr, src_rr = np.nan, ""
        for rate, src in [(or_c_rr, "OR_Contracted"),
                          (or_p_rr, "OR_Participating"),
                          (wa_p_rr, "WA_Participating")]:
            if not math.isnan(rate):
                phcc_pri_rr, src_rr = rate, src
                break

        # CMS benchmark matched to the selected PHCC source
        cms_pri_nu = (cms_or_nu if src_nu.startswith("OR") or not src_nu
                      else cms_wa_nu)
        cms_pri_rr = (cms_or_rr if src_rr.startswith("OR") or not src_rr
                      else cms_wa_rr)

        # ── Deltas ──
        d_nu, p_nu = _delta(int_nu, phcc_pri_nu)
        d_rr, p_rr = _delta(int_rr, phcc_pri_rr)

        # ── Flags ──
        f_nu = _flag(int_nu, phcc_pri_nu, cms_pri_nu, in_phcc, is_phcc_only)
        f_rr = _flag(int_rr, phcc_pri_rr, cms_pri_rr, in_phcc, is_phcc_only)
        f_nu = _systemic(phcc_pri_nu, cms_pri_nu, f_nu)
        f_rr = _systemic(phcc_pri_rr, cms_pri_rr, f_rr)

        rows.append({
            "HCPCS":           hcpcs,
            "Description":     hcpcs_desc.get(hcpcs, ""),
            "Source":          source,
            "Integra NU":      int_nu,
            "Integra RR":      int_rr,
            "Integra Note":    int_note,
            "OR Contract NU":  or_c_nu,
            "OR Contract RR":  or_c_rr,
            "OR Partic NU":    or_p_nu,
            "OR Partic RR":    or_p_rr,
            "WA Partic NU":    wa_p_nu,
            "WA Partic RR":    wa_p_rr,
            "CMS OR NU":       cms_or_nu,
            "CMS OR RR":       cms_or_rr,
            "CMS WA NU":       cms_wa_nu,
            "CMS WA RR":       cms_wa_rr,
            "OHA NU":          oha_nu,
            "OHA RR":          oha_rr,
            "Δ NU":            d_nu,
            "Δ RR":            d_rr,
            "Δ% NU":           p_nu,
            "Δ% RR":           p_rr,
            "Flag NU":         f_nu,
            "Flag RR":         f_rr,
            "PHCC Source NU":  src_nu,
            "PHCC Source RR":  src_rr,
        })

    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════
# 4.  XLSX FORMATTING
# ═══════════════════════════════════════════════════════════════════════
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
LTGRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(bold=True, size=11)
CURRENCY    = '"$"#,##0.00'
PCT_FMT     = '0.0"%"'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

FLAG_COLORS = {
    "BELOW CMS FLOOR": RED_FILL,
    "PHCC BELOW CMS":  ORANGE_FILL,
    "BELOW CURRENT":   YELLOW_FILL,
    "RATE INCREASE":   BLUE_FILL,
    "NO CHANGE":       GREEN_FILL,
    "NEW CODE":        GRAY_FILL,
    "PHCC ONLY":       LTGRAY_FILL,
    "PHCC NON-NUMERIC": GRAY_FILL,
    "REVIEW":          GRAY_FILL,
}

# Column indices (0-based) that should be formatted as currency
CURRENCY_COLS = {3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19}
PCT_COLS      = {20, 21}

# Column names in order
COL_ORDER = [
    "HCPCS", "Description", "Source",
    "Integra NU", "Integra RR", "Integra Note",
    "OR Contract NU", "OR Contract RR",
    "OR Partic NU", "OR Partic RR",
    "WA Partic NU", "WA Partic RR",
    "CMS OR NU", "CMS OR RR",
    "CMS WA NU", "CMS WA RR",
    "OHA NU", "OHA RR",
    "Δ NU", "Δ RR",
    "Δ% NU", "Δ% RR",
    "Flag NU", "Flag RR",
    "PHCC Source NU", "PHCC Source RR",
]


def _flag_fill(text: str):
    """Choose fill based on flag keywords."""
    if not text:
        return None
    for kw, fill in FLAG_COLORS.items():
        if kw in text:
            return fill
    return None


def _auto_width(ws, max_w=30):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col[:80]:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(mx + 2, 8), max_w)


def write_payer_tab(ws, df: pd.DataFrame, payer: str):
    """Write a payer dataframe to a worksheet with formatting."""
    if df.empty:
        return

    # Reorder columns
    cols = [c for c in COL_ORDER if c in df.columns]
    df = df[cols]

    # Header row
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col_name in enumerate(cols, 1):
            val = row[col_name]
            # Convert NaN → None for cleaner output
            if isinstance(val, float) and math.isnan(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER

            # Number formatting
            idx = ci - 1  # 0-based
            if idx in CURRENCY_COLS and val is not None:
                cell.number_format = CURRENCY
            elif idx in PCT_COLS and val is not None:
                cell.number_format = PCT_FMT

        # Flag colouring
        for flag_col_name in ("Flag NU", "Flag RR"):
            if flag_col_name in cols:
                flag_idx = cols.index(flag_col_name) + 1
                flag_val = str(row.get(flag_col_name, "") or "")
                fill = _flag_fill(flag_val)
                if fill:
                    ws.cell(row=ri, column=flag_idx).fill = fill

    # Freeze panes + auto-filter
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def write_summary(ws, stats: dict):
    """Write the summary/methodology tab."""
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1,
                    value="Unified Code-Centric Fee Schedule Analysis — Summary")
    title.font = Font(bold=True, size=14)

    r = 3
    sections = [
        ("Code Universe", [
            ("Total unique HCPCS codes", stats["universe"]),
            ("In Integra", stats["integra"]),
            ("In PHCC", stats["phcc"]),
            ("In Both", stats["both"]),
            ("Integra Only", stats["integra_only"]),
            ("PHCC Only", stats["phcc_only"]),
        ]),
    ]

    for section_title, items in sections:
        ws.cell(row=r, column=1, value=section_title).font = BOLD_FONT
        r += 1
        for label, val in items:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        r += 1

    # Per-payer stats
    for payer, pstats in stats.get("payers", {}).items():
        ws.cell(row=r, column=1, value=f"{payer} — Matching Statistics").font = BOLD_FONT
        r += 1
        for label, val in pstats:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        r += 1

    _auto_width(ws, max_w=50)


def compute_payer_stats(df: pd.DataFrame, payer: str) -> list:
    """Compute summary statistics for a payer table."""
    total = len(df)
    both = len(df[df["Source"] == "BOTH"])
    int_only = len(df[df["Source"] == "INTEGRA_ONLY"])
    phcc_only = len(df[df["Source"] == "PHCC_ONLY"])

    int_nu_pop = df["Integra NU"].notna().sum()
    int_rr_pop = df["Integra RR"].notna().sum()

    # PHCC match rate (codes where at least one PHCC schedule has a NU rate)
    phcc_nu_any = (
        df["OR Contract NU"].notna()
        | df["OR Partic NU"].notna()
        | df["WA Partic NU"].notna()
    ).sum()
    phcc_rr_any = (
        df["OR Contract RR"].notna()
        | df["OR Partic RR"].notna()
        | df["WA Partic RR"].notna()
    ).sum()

    # Flag distribution for NU
    flag_nu = df["Flag NU"].fillna("").value_counts()
    flag_rr = df["Flag RR"].fillna("").value_counts()

    items = [
        ("Total codes", total),
        ("BOTH (in Integra + PHCC)", both),
        ("INTEGRA_ONLY", int_only),
        ("PHCC_ONLY", phcc_only),
        ("", ""),
        ("Integra NU rates populated", int(int_nu_pop)),
        ("Integra RR rates populated", int(int_rr_pop)),
        ("PHCC NU (any schedule) populated", int(phcc_nu_any)),
        ("PHCC RR (any schedule) populated", int(phcc_rr_any)),
        ("", ""),
        ("— NU Flag Distribution —", ""),
    ]
    for flag, cnt in flag_nu.items():
        if flag:
            items.append((f"  {flag}", int(cnt)))
    items.append(("", ""))
    items.append(("— RR Flag Distribution —", ""))
    for flag, cnt in flag_rr.items():
        if flag:
            items.append((f"  {flag}", int(cnt)))

    return items


# ═══════════════════════════════════════════════════════════════════════
# 5.  MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    # ── Verify files ──
    missing = [k for k, p in FILES.items() if not p.exists()]
    if missing:
        print("ERROR — missing files:")
        for k in missing:
            print(f"  {k}: {FILES[k]}")
        print("\nRun 'python scripts/clean_phcc_files.py' first.")
        sys.exit(1)

    OUTPUT.mkdir(exist_ok=True)

    # ── Load reference data ──
    print("Loading data …")
    hcpcs_desc = load_hcpcs_desc(FILES["hcpcs"])
    cms_or = load_cms(FILES["cms_or"], "OR (NR)", "OR (R)")
    cms_wa = load_cms(FILES["cms_wa"], "WA (NR)", "WA (R)")
    oha_lk = load_oha(FILES["oha"])

    # ── Load Integra (per payer) ──
    integra_lks = {}
    for payer, cfg in PAYER_CFG.items():
        integra_lks[payer] = load_integra(
            FILES[cfg["integra_key"]], cfg["rate_col"])
        print(f"  Integra {payer}: {len(integra_lks[payer])} codes")

    # ── Build code universe from ALL Integra payers ──
    integra_codes = set()
    for lk in integra_lks.values():
        integra_codes.update(lk.keys())

    # ── Load PHCC (per payer due to OR_CONTRACTED payer-dependent cols) ──
    # For each payer, OR_CONTRACTED uses a different prefix
    phcc_lks = {}
    phcc_codes = set()
    for payer, cfg in PAYER_CFG.items():
        or_c = load_phcc(FILES["or_contracted"], "or_contracted",
                         cfg["or_c_prefix"])
        or_p = load_phcc(FILES["or_participating"], "or_participating")
        wa_p = load_phcc(FILES["wa_participating"], "wa_participating")
        phcc_lks[payer] = {"or_c": or_c, "or_p": or_p, "wa_p": wa_p}
        phcc_codes.update(or_c.keys())
        phcc_codes.update(or_p.keys())
        phcc_codes.update(wa_p.keys())
        print(f"  PHCC for {payer}: OR_C={len(or_c)}, "
              f"OR_P={len(or_p)}, WA_P={len(wa_p)}")

    # ── Build universe ──
    universe = integra_codes | phcc_codes
    overlap = integra_codes & phcc_codes
    print(f"\nCode universe: {len(universe)} "
          f"(Integra={len(integra_codes)}, PHCC={len(phcc_codes)}, "
          f"Both={len(overlap)}, "
          f"Integra-only={len(integra_codes - phcc_codes)}, "
          f"PHCC-only={len(phcc_codes - integra_codes)})")

    # ── Build per-payer tables ──
    payer_tables = {}
    payer_stats = {}
    for payer in PAYER_CFG:
        print(f"\nBuilding {payer} table …")
        df = build_payer_table(
            payer=payer,
            universe=universe,
            integra_lk=integra_lks[payer],
            phcc_or_c=phcc_lks[payer]["or_c"],
            phcc_or_p=phcc_lks[payer]["or_p"],
            phcc_wa_p=phcc_lks[payer]["wa_p"],
            cms_or=cms_or,
            cms_wa=cms_wa,
            oha_lk=oha_lk,
            hcpcs_desc=hcpcs_desc,
        )
        payer_tables[payer] = df
        payer_stats[payer] = compute_payer_stats(df, payer)
        print(f"  → {len(df)} rows")

    # ── Summary stats ──
    summary = {
        "universe": len(universe),
        "integra": len(integra_codes),
        "phcc": len(phcc_codes),
        "both": len(overlap),
        "integra_only": len(integra_codes - phcc_codes),
        "phcc_only": len(phcc_codes - integra_codes),
        "payers": payer_stats,
    }

    # ── Write XLSX ──
    out_path = OUTPUT / "unified_code_analysis.xlsx"
    print(f"\nWriting {out_path} …")
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary(ws_summary, summary)

    for payer, df in payer_tables.items():
        ws = wb.create_sheet(title=payer)
        write_payer_tab(ws, df, payer)

    wb.save(out_path)
    print(f"✓ Saved: {out_path}")
    print(f"  Tabs: Summary, {', '.join(payer_tables.keys())}")


if __name__ == "__main__":
    main()
