"""
INTEGRA PHP FFS — Executive Rate Analysis  v2
===============================================
Changes from v1:
  • T5 range matching: Integra L-codes now match PHCC CATEGORY_RANGE rows
    (e.g. L3000 matches "L3000-L4631")
  • "Medicare Allowable less X%" resolved to dollar amounts using CMS NR data
    (e.g. "Medicare Allowable less 20%" + CMS NR $100 → $80.00)
  • New decision tree: Proposed vs Current FIRST, then CMS benchmark only
    when proposed is below current
  • Added "Prevailing State Rates" handling (treated as note, flagged)

Prerequisite: python scripts/clean_phcc_files.py
Run:          python scripts/integra_rate_analysis_v2.py

Output: PHCC/output/integra_rate_analysis.xlsx
"""
from __future__ import annotations
import re, os, sys, math
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ───────────────────────────────────────────────────────────────────────
# 0.  PATHS & CONFIG
# ───────────────────────────────────────────────────────────────────────
PHCC_ROOT = Path(__file__).resolve().parent.parent
CLEANED   = PHCC_ROOT / "data" / "cleaned"
INTEGRA   = PHCC_ROOT / "data" / "INTEGRA_PHP_FFS"
CMS_DIR   = PHCC_ROOT / "data" / "cms"
OUTPUT    = PHCC_ROOT / "output"
OUTPUT.mkdir(exist_ok=True)

FILES = {
    "or_contracted":    CLEANED / "PHCC_OR_CONTRACTED_CLEAN.csv",
    "or_participating": CLEANED / "PHCC_OR_PARTICIPATING_CLEAN.csv",
    "wa_participating": CLEANED / "PHCC_WA_PARTICIPATING_CLEAN.csv",
    "integra_commercial": INTEGRA / "Integra_PHP_CARVEOUTS_COMMERCIAL.csv",
    "integra_aso":        INTEGRA / "Integra_PHP_CARVEOUTS_ASO.csv",
    "integra_medicare":   INTEGRA / "Integra_PHP_CARVEOUTS_MEDICARE.csv",
    "integra_medicaid":   INTEGRA / "INTEGRA_PHP_CARVEOUTS_MEDICAID.csv",
    "cms_or": CMS_DIR / "CMS_2026_Q1_OR.csv",
    "cms_wa": CMS_DIR / "CMS_2026_Q1_WA.csv",
    "oha":    CMS_DIR / "OHA_FFS_09_2025_RAW.csv",
    "hcpcs":  CMS_DIR / "2026_CMS_HCPCS.csv",
    # Raw contract files (for Contract View tabs)
    "or_contracted_raw":    PHCC_ROOT / "data" / "Contract" / "PHCC_OR_CONTRACTED.csv",
    "or_participating_raw": PHCC_ROOT / "data" / "Contract" / "PHCC_OR_PARTICIPATING.csv",
    "wa_participating_raw": PHCC_ROOT / "data" / "Contract" / "PHCC_WA_PARTICIPATING.csv",
}

# Payer configs: (integra_key, payer_label, integra_rate_col,
#                  phcc_keys, state, benchmark_type)
PAYER_CONFIGS = {
    "Commercial": [
        ("integra_commercial", "Commercial", "Commercial",
         ["or_contracted", "or_participating"], "OR", "CMS"),
        ("integra_commercial", "Commercial", "Commercial",
         ["wa_participating"], "WA", "CMS"),
    ],
    "ASO": [
        ("integra_aso", "ASO", "ASO/Commercial",
         ["or_contracted", "or_participating"], "OR", "CMS"),
        ("integra_aso", "ASO", "ASO/Commercial",
         ["wa_participating"], "WA", "CMS"),
    ],
    "Medicare": [
        ("integra_medicare", "Medicare", "Medicare",
         ["or_contracted", "or_participating"], "OR", "CMS"),
        ("integra_medicare", "Medicare", "Medicare",
         ["wa_participating"], "WA", "CMS"),
    ],
    "Medicaid": [
        ("integra_medicaid", "Medicaid", "Medicaid",
         ["or_contracted", "or_participating"], "OR", "OHA"),
        ("integra_medicaid", "Medicaid", "Medicaid",
         ["wa_participating"], "WA", None),
    ],
}

# ───────────────────────────────────────────────────────────────────────
# 1.  PURE HELPERS
# ───────────────────────────────────────────────────────────────────────
VALID_HCPCS_RE = re.compile(r'^[A-Z][0-9]{4}$')
MEDICARE_ALLOWABLE_RE = re.compile(
    r'Medicare\s+Allowable\s+less\s+(\d+)\s*%', re.IGNORECASE)

def _norm(raw) -> str:
    if pd.isna(raw) or str(raw).strip() == "":
        return ""
    return str(raw).strip().upper()

def _safe_float(val) -> float:
    if pd.isna(val):
        return np.nan
    s = str(val).strip().replace("$", "").replace(",", "")
    if s == "" or s.lower() == "nan":
        return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan

def _norm_mod(raw) -> str:
    if pd.isna(raw) or str(raw).strip() == "":
        return ""
    return str(raw).strip().upper().rstrip("*")

def _classify_note(val) -> tuple[float, str]:
    """Return (numeric_or_nan, note_text)."""
    if pd.isna(val):
        return np.nan, ""
    s = str(val).strip()
    if s == "":
        return np.nan, ""
    cleaned = s.replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned), ""
    except ValueError:
        return np.nan, s

def _resolve_pct_of_medicare(note_detail: str, cms_nr: float) -> float:
    """Compute rate from 'Medicare Allowable less X%' using CMS NR."""
    m = MEDICARE_ALLOWABLE_RE.search(note_detail)
    if m and not math.isnan(cms_nr):
        pct = int(m.group(1))
        return round(cms_nr * (1 - pct / 100), 2)
    return np.nan

def _in_range(code: str, start: str, end: str) -> bool:
    """Check if a single HCPCS code falls within [start, end] range."""
    if not code or not start or not end:
        return False
    if len(code) < 2 or len(start) < 2 or len(end) < 2:
        return False
    if code[0] != start[0] or code[0] != end[0]:
        return False
    try:
        c = int(code[1:])
        s = int(start[1:])
        e = int(end[1:])
        return s <= c <= e
    except ValueError:
        return False

def _to_num(s: str):
    """Convert string to float if numeric, otherwise return original string."""
    if not s:
        return ""
    try:
        return float(s.replace("$", "").replace(",", ""))
    except (ValueError, TypeError):
        return s

CROSS_MOD = {"NU": ["RR", ""], "RR": ["NU", ""], "AU": ["NU", ""],
             "KF": ["NU", ""], "": ["NU", "RR"]}

# ───────────────────────────────────────────────────────────────────────
# 2.  LOADERS
# ───────────────────────────────────────────────────────────────────────

def load_cleaned_phcc(path: Path, label: str) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False)

def _pick_rate(row, schedule_label: str, payer: str, modifier: str,
               cms_nr: float = np.nan) -> tuple[float, str]:
    """Return (numeric_rate, raw_value).  Resolves Medicare Allowable % rates."""
    if schedule_label == "PHCC_OR_CONTRACTED":
        prefix = "Managed" if payer in ("Medicare", "Medicaid") else "Commercial"
        rate_col = f"{prefix} Rental Rate" if modifier == "RR" else f"{prefix} Purchase Rate"
    else:
        rate_col = "Rental Rate" if modifier == "RR" else "Purchase Rate"

    raw = str(row.get(f"{rate_col}_raw", ""))
    num = _safe_float(row.get(f"{rate_col}_numeric", ""))
    note_type = str(row.get(f"{rate_col}_note_type", ""))
    note_detail = str(row.get(f"{rate_col}_note_detail", ""))

    # Resolve "Medicare Allowable less X%" → actual dollar amount
    if math.isnan(num) and note_type == "PERCENT_OF_MEDICARE_ALLOWABLE":
        resolved = _resolve_pct_of_medicare(note_detail, cms_nr)
        if not math.isnan(resolved):
            raw = f"{note_detail} → ${resolved:.2f}"
            num = resolved

    return num, raw

def load_integra(path: Path, rate_col: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df.columns = [c.strip() for c in df.columns]
    rows = []
    for idx, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod1 = _norm_mod(r.get("Mod 1", ""))
        raw = str(r.get(rate_col, "")).strip()
        num, note = _classify_note(raw)
        rows.append({
            "hcpcs": hcpcs, "mod": mod1,
            "proposed_rate": num, "proposed_raw": raw, "proposed_note": note,
        })
    return pd.DataFrame(rows)

def load_cms(path: Path, nr_col: str, r_col: str) -> dict[str, dict]:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    lk = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod = _norm_mod(r.get("Mod", ""))
        if not hcpcs:
            continue
        lk[f"{hcpcs}|{mod}"] = {
            "nr": _safe_float(r.get(nr_col, "")),
            "r":  _safe_float(r.get(r_col, "")),
        }
    return lk

def load_oha(path: Path) -> dict[str, float]:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    lk = {}
    for _, r in df.iterrows():
        code = _norm(r.get("Procedure Code", ""))
        mod = _norm_mod(r.get("Mod1", ""))
        if code:
            lk[f"{code}|{mod}"] = _safe_float(r.get("Price", ""))
    return lk

def load_hcpcs_desc(path: Path) -> dict[str, str]:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    return {_norm(r.get("HCPC", "")): str(r.get("SHORT DESCRIPTION", "")).strip()
            for _, r in df.iterrows() if _norm(r.get("HCPC", ""))}


def load_raw_contract(path: Path) -> pd.DataFrame:
    """Load a raw (uncleaned) PHCC contract CSV for Contract View."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    # Normalise HCPCS + Mod columns (handle both 'Mod' and 'Modifier' headers)
    hcpcs_col = "HCPCS"
    mod_col = "Mod" if "Mod" in df.columns else "Modifier"
    df["_hcpcs"] = df[hcpcs_col].str.strip().str.upper()
    df["_mod"] = df[mod_col].str.strip().str.upper()
    return df


def _build_integra_lk(integra_dfs: dict) -> dict:
    """Build {hcpcs|mod: {payer: rate, payer_raw: raw}} from loaded Integra DFs."""
    payer_map = {
        "integra_commercial": "Commercial",
        "integra_aso":        "ASO",
        "integra_medicare":   "Medicare",
        "integra_medicaid":   "Medicaid",
    }
    lk = {}
    for key, payer_name in payer_map.items():
        df = integra_dfs[key]
        for _, row in df.iterrows():
            h = row["hcpcs"]
            m = row["mod"]
            if not h:
                continue
            k = f"{h}|{m}"
            lk.setdefault(k, {})[payer_name] = row["proposed_rate"]
            lk[k][f"{payer_name}_raw"] = row["proposed_raw"]
    return lk


# ───────────────────────────────────────────────────────────────────────
# 3.  PHCC LOOKUP + MATCHING  (v2 — range-aware)
# ───────────────────────────────────────────────────────────────────────

def build_phcc(df: pd.DataFrame):
    """Return (key_lk, code_lk, range_lk).

    range_lk is a list of (range_start, range_end, modifier, row)
    for CATEGORY_RANGE entries that cover large code spans.
    """
    key_lk, code_lk, range_lk = {}, {}, []
    for _, row in df.iterrows():
        h = str(row.get("hcpcs_normalised", "")).strip()
        m = str(row.get("modifier_normalised", "")).strip()
        key_lk.setdefault(f"{h}|{m}", []).append(row)
        code_lk.setdefault(h, []).append(row)

        # Catalog CATEGORY_RANGE entries for range matching
        issue = str(row.get("hcpcs_issue_type", "")).strip()
        if issue == "CATEGORY_RANGE":
            rs = str(row.get("range_start", "")).strip()
            re_ = str(row.get("range_end", "")).strip()
            if rs and re_:
                range_lk.append((rs, re_, m, row))

    return key_lk, code_lk, range_lk


def best_match(hcpcs: str, mod: str,
               key_lk: dict, code_lk: dict, range_lk: list):
    """Return (phcc_row_or_None, tier_label, cross_mod_used).

    Tiers:
      T1 = exact HCPCS + Mod
      T2 = proposed mod → blank PHCC
      T3 = cross-modifier fallback
      T4 = HCPCS-only (any modifier)
      T5 = CATEGORY_RANGE (code falls within a range row)
    """
    # T1: exact
    rows = key_lk.get(f"{hcpcs}|{mod}", [])
    if rows:
        return rows[0], "T1", ""
    # T2: proposed mod → blank PHCC
    if mod:
        rows = key_lk.get(f"{hcpcs}|", [])
        if rows:
            return rows[0], "T2", ""
    # T3: cross-modifier
    for alt in CROSS_MOD.get(mod, ["NU", "RR", ""]):
        if alt == mod or (alt == "" and mod):
            continue
        rows = key_lk.get(f"{hcpcs}|{alt}", [])
        if rows:
            return rows[0], "T3", alt
    # T4: any row for that HCPCS
    rows = code_lk.get(hcpcs, [])
    if rows:
        return rows[0], "T4", str(rows[0].get("modifier_normalised", ""))
    # T5: range match — code falls within CATEGORY_RANGE
    for rs, re_, r_mod, r_row in range_lk:
        if _in_range(hcpcs, rs, re_):
            return r_row, "T5_RANGE", r_mod
    return None, "NO_MATCH", ""


def cms_cascade(hcpcs: str, mod: str, cms_lk: dict) -> tuple[float, float, str]:
    """B1→B4 cascade. Returns (nr_rate, rural_rate, match_tier)."""
    for tier, try_mod in [("B1", mod), ("B2", "NU"), ("B3", "RR"), ("B4", "")]:
        rec = cms_lk.get(f"{hcpcs}|{try_mod}")
        if rec:
            return rec["nr"], rec["r"], tier
    return np.nan, np.nan, "NOT_FOUND"


# ───────────────────────────────────────────────────────────────────────
# 4.  BUILD COMPARISON TABLE PER PAYER  (v2 — new decision tree)
# ───────────────────────────────────────────────────────────────────────
TOLERANCE_PCT = 1.0  # ±1% treated as "no change"

def build_payer_table(
    payer: str,
    configs: list,
    integra_dfs: dict,
    phcc_dfs: dict, phcc_key_lks: dict, phcc_code_lks: dict,
    phcc_range_lks: dict,
    cms_or: dict, cms_wa: dict, oha_lk: dict,
    hcpcs_desc: dict,
) -> pd.DataFrame:
    """Build the slim comparison rows for one payer type.

    Decision tree (v2):
      1. Match Proposed → PHCC Current  (T1-T5)
      2. Resolve "Medicare Allowable less X%" → CMS NR × (1 - X/100)
      3. Compare Proposed vs Current:
         • NO MATCH         → "NEW CODE"            (gray)
         • ±1%              → "NO CHANGE"           (no color)
         • Proposed > Cur   → "RATE INCREASE"        (blue)
         • Proposed < Cur   → Step 4
      4. Proposed < Current → check CMS benchmark:
         • Proposed ≥ CMS   → "BELOW CURRENT"        (yellow)
         • Proposed < CMS   → "BELOW CMS FLOOR"      (red)
         • No CMS data      → "BELOW CURRENT"        (yellow)
      Additional:
         • PHCC Current < CMS → append " | PHCC BELOW CMS" (systemic)
    """
    rows = []
    for (integra_key, payer_label, rate_col,
         phcc_keys, state, bench_type) in configs:

        integra_df = integra_dfs[integra_key]
        cms_lk = cms_or if state == "OR" else cms_wa

        for _, prop in integra_df.iterrows():
            hcpcs = prop["hcpcs"]
            mod = prop["mod"]
            proposed_num = prop["proposed_rate"]
            proposed_raw = prop["proposed_raw"]
            proposed_note = prop["proposed_note"]

            if not hcpcs:
                continue

            # CMS benchmark for this code
            cms_nr, cms_r, cms_tier = cms_cascade(hcpcs, mod, cms_lk)

            # OHA benchmark (Medicaid only)
            oha_rate = np.nan
            if bench_type == "OHA":
                for oha_tier, try_mod in [("B1", mod), ("B2", "NU"),
                                          ("B3", "RR"), ("B4", "")]:
                    oha_val = oha_lk.get(f"{hcpcs}|{try_mod}", np.nan)
                    if not math.isnan(oha_val):
                        oha_rate = oha_val
                        break

            # Match against each PHCC schedule
            for sched_key in phcc_keys:
                label = {
                    "or_contracted": "PHCC_OR_CONTRACTED",
                    "or_participating": "PHCC_OR_PARTICIPATING",
                    "wa_participating": "PHCC_WA_PARTICIPATING",
                }[sched_key]

                phcc_row, tier, cross_mod = best_match(
                    hcpcs, mod,
                    phcc_key_lks[sched_key],
                    phcc_code_lks[sched_key],
                    phcc_range_lks[sched_key])

                if phcc_row is not None:
                    eff_mod = cross_mod if cross_mod else mod
                    # Pass cms_nr so _pick_rate can resolve Medicare Allowable %
                    cur_num, cur_raw = _pick_rate(
                        phcc_row, label, payer, eff_mod, cms_nr)
                else:
                    cur_num, cur_raw = np.nan, ""

                # ── Deltas ──
                delta_prop_phcc = np.nan
                pct_prop_phcc = np.nan
                have_both = (not math.isnan(proposed_num)
                             and not math.isnan(cur_num))
                if have_both:
                    delta_prop_phcc = proposed_num - cur_num
                    if cur_num != 0:
                        pct_prop_phcc = delta_prop_phcc / cur_num * 100

                delta_cur_cms_nr = np.nan
                if not math.isnan(cur_num) and not math.isnan(cms_nr):
                    delta_cur_cms_nr = cur_num - cms_nr

                delta_prop_cms_nr = np.nan
                if not math.isnan(proposed_num) and not math.isnan(cms_nr):
                    delta_prop_cms_nr = proposed_num - cms_nr

                # ── Decision tree flags (v2) ──
                flag = ""
                sort_priority = 4  # default: neutral / OK

                if tier == "NO_MATCH":
                    flag = "NEW CODE"
                    sort_priority = 5
                elif not have_both:
                    # Matched but one side is non-numeric
                    if math.isnan(proposed_num) and proposed_note:
                        flag = f"REVIEW: {proposed_note[:40]}"
                    elif math.isnan(cur_num):
                        flag = "PHCC RATE NON-NUMERIC"
                    else:
                        flag = "REVIEW"
                    sort_priority = 6
                else:
                    # Step 3: compare proposed vs current
                    if cur_num != 0 and abs(pct_prop_phcc) <= TOLERANCE_PCT:
                        flag = "NO CHANGE"
                        sort_priority = 7  # lowest concern
                    elif proposed_num > cur_num:
                        flag = "RATE INCREASE"
                        sort_priority = 3
                    else:
                        # Step 4: proposed < current → check CMS
                        if not math.isnan(cms_nr):
                            if proposed_num >= cms_nr:
                                flag = "BELOW CURRENT"
                                sort_priority = 2
                            else:
                                flag = "BELOW CMS FLOOR"
                                sort_priority = 0  # worst
                        else:
                            flag = "BELOW CURRENT"
                            sort_priority = 2

                # Append systemic issue if PHCC current < CMS
                if (not math.isnan(cur_num)
                        and not math.isnan(cms_nr)
                        and cur_num < cms_nr):
                    if flag:
                        flag += " | PHCC BELOW CMS"
                    else:
                        flag = "PHCC BELOW CMS"
                    # Bump priority if it's the only flag
                    if sort_priority > 1:
                        sort_priority = 1

                rows.append({
                    "State": state,
                    "Schedule": label.replace("PHCC_", ""),
                    "HCPCS": hcpcs,
                    "Mod": mod,
                    "Description": hcpcs_desc.get(hcpcs, ""),
                    "Proposed Rate": proposed_num,
                    "PHCC Current": cur_num,
                    "Δ Proposed–PHCC": delta_prop_phcc,
                    "Δ%": pct_prop_phcc,
                    "CMS NR": cms_nr,
                    "CMS Rural": cms_r,
                    "Proposed–CMS NR": delta_prop_cms_nr,
                    "PHCC–CMS NR": delta_cur_cms_nr,
                    "OHA Rate": oha_rate if bench_type == "OHA" else np.nan,
                    "Flag": flag,
                    "Match": tier,
                    "PHCC Raw": cur_raw,
                    "Note": proposed_note if proposed_note else "",
                    "_sort": sort_priority,
                })

    df = pd.DataFrame(rows)
    if len(df) > 0:
        df = df.sort_values(["_sort", "HCPCS", "State"], ascending=True)
        df = df.drop(columns=["_sort"])
    return df


# ───────────────────────────────────────────────────────────────────────
# 4b. CONTRACT VIEW BUILDER
# ───────────────────────────────────────────────────────────────────────

def _integra_rate(integra_lk: dict, hcpcs: str, mod: str, payer: str) -> float:
    """Look up Integra proposed rate for hcpcs|mod, with NU/RR/''/blank fallback."""
    for try_mod in [mod, "NU", "RR", ""]:
        rec = integra_lk.get(f"{hcpcs}|{try_mod}")
        if rec and not math.isnan(rec.get(payer, np.nan)):
            return rec[payer]
    return np.nan


def build_contract_view(
    raw_df: pd.DataFrame,
    schedule_key: str,
    integra_lk: dict,
    phcc_key_lk: dict,
    phcc_code_lk: dict,
    phcc_range_lk: list,
    cms_lk: dict,
    oha_lk: dict,
    hcpcs_desc: dict,
) -> pd.DataFrame:
    """Build a Contract View DataFrame for one raw contract file.

    Left side  = raw contract columns (original text).
    Right side = Integra proposed rates + resolved current + CMS + deltas + flag.
    """
    is_contracted = (schedule_key == "or_contracted")
    schedule_label = {
        "or_contracted":    "PHCC_OR_CONTRACTED",
        "or_participating": "PHCC_OR_PARTICIPATING",
        "wa_participating": "PHCC_WA_PARTICIPATING",
    }[schedule_key]

    rows = []
    for _, r in raw_df.iterrows():
        hcpcs = r["_hcpcs"]
        if not VALID_HCPCS_RE.match(hcpcs):
            continue
        mod_raw = r["_mod"]

        # Raw contract text columns
        desc_raw  = str(r.get("Description", "")).strip()
        unit_raw  = str(r.get("Billing Unit", "")).strip()
        comments  = str(r.get("Comments", "")).strip()

        if is_contracted:
            mgd_rent  = str(r.get("Managed Rental Rate", "")).strip()
            mgd_purch = str(r.get("Managed Purchase Rate", "")).strip()
            com_rent  = str(r.get("Commercial Rental Rate", "")).strip()
            com_purch = str(r.get("Commercial Purchase Rate", "")).strip()
        else:
            rent_raw  = str(r.get("Rental Rate", "")).strip()
            purch_raw = str(r.get("Purchase Rate", "")).strip()

        # Split NU/RR modifiers into separate rows
        mods_to_process = []
        clean_mod = mod_raw.replace("**", "").replace("*", "")
        if "/" in clean_mod:
            mods_to_process = [m.strip() for m in clean_mod.split("/")
                               if m.strip() in ("NU", "RR")]
        elif clean_mod in ("NU", "RR", ""):
            mods_to_process = [clean_mod if clean_mod else "NU"]
        else:
            # Composite modifiers like RR,QG,QF — take first component
            mods_to_process = [clean_mod.split(",")[0].strip()]
            if mods_to_process[0] not in ("NU", "RR"):
                mods_to_process = ["NU"]

        for mod in mods_to_process:
            # ── Match PHCC cleaned row ──
            phcc_row, tier, cross_mod = best_match(
                hcpcs, mod, phcc_key_lk, phcc_code_lk, phcc_range_lk)

            # ── CMS benchmark ──
            cms_nr, cms_r, _ = cms_cascade(hcpcs, mod, cms_lk)

            # ── OHA ──
            oha_rate = np.nan
            for _, try_mod in [("B1", mod), ("B2", "NU"), ("B3", "RR"), ("B4", "")]:
                oha_val = oha_lk.get(f"{hcpcs}|{try_mod}", np.nan)
                if not math.isnan(oha_val):
                    oha_rate = oha_val
                    break

            # ── Current numeric from cleaned data ──
            if is_contracted:
                eff_mod = cross_mod if cross_mod else mod
                mgd_num, mgd_num_raw = (
                    _pick_rate(phcc_row, schedule_label, "Medicare", eff_mod, cms_nr)
                    if phcc_row is not None else (np.nan, ""))
                com_num, com_num_raw = (
                    _pick_rate(phcc_row, schedule_label, "Commercial", eff_mod, cms_nr)
                    if phcc_row is not None else (np.nan, ""))
            else:
                eff_mod = cross_mod if cross_mod else mod
                cur_num, cur_raw = (
                    _pick_rate(phcc_row, schedule_label, "Commercial", eff_mod, cms_nr)
                    if phcc_row is not None else (np.nan, ""))

            # ── Integra proposed rates ──
            int_comm = _integra_rate(integra_lk, hcpcs, mod, "Commercial")
            int_aso  = _integra_rate(integra_lk, hcpcs, mod, "ASO")
            int_med  = _integra_rate(integra_lk, hcpcs, mod, "Medicare")
            int_mcd  = _integra_rate(integra_lk, hcpcs, mod, "Medicaid")

            # ── Primary delta: Integra Commercial vs current rate ──
            if is_contracted:
                primary_proposed = int_comm
                primary_current  = com_num
            else:
                primary_proposed = int_comm
                primary_current  = cur_num

            delta, pct = np.nan, np.nan
            have_both = (not math.isnan(primary_proposed)
                         and not math.isnan(primary_current))
            if have_both:
                delta = primary_proposed - primary_current
                if primary_current != 0:
                    pct = delta / primary_current * 100

            # ── Flag (same logic as payer tabs) ──
            flag = ""
            if tier == "NO_MATCH":
                flag = "NO PHCC MATCH"
            elif not have_both:
                if math.isnan(primary_proposed):
                    flag = "NO INTEGRA RATE"
                elif math.isnan(primary_current):
                    flag = "NON-NUMERIC CURRENT"
            else:
                if primary_current != 0 and abs(pct) <= TOLERANCE_PCT:
                    flag = "NO CHANGE"
                elif primary_proposed > primary_current:
                    flag = "RATE INCREASE"
                else:
                    if not math.isnan(cms_nr):
                        if primary_proposed >= cms_nr:
                            flag = "BELOW CURRENT"
                        else:
                            flag = "BELOW CMS FLOOR"
                    else:
                        flag = "BELOW CURRENT"

            # Systemic: current < CMS
            if (not math.isnan(primary_current)
                    and not math.isnan(cms_nr)
                    and primary_current < cms_nr):
                flag = f"{flag} | PHCC BELOW CMS" if flag else "PHCC BELOW CMS"

            # ── Build row ──
            if is_contracted:
                row = {
                    "HCPCS":                hcpcs,
                    "Mod":                  mod,
                    "Description":          desc_raw,
                    "Billing Unit":         unit_raw,
                    "Managed Rental":       _to_num(mgd_rent) if mod == "RR" else "",
                    "Managed Purchase":     _to_num(mgd_purch) if mod == "NU" else "",
                    "Commercial Rental":    _to_num(com_rent) if mod == "RR" else "",
                    "Commercial Purchase":  _to_num(com_purch) if mod == "NU" else "",
                    "Comments":             comments,
                    # --- Comparison ---
                    "Managed $":            mgd_num,
                    "Commercial $":         com_num,
                    "Integra Comm":         int_comm,
                    "Integra ASO":          int_aso,
                    "Integra Medicare":     int_med,
                    "Integra Medicaid":     int_mcd,
                    "CMS NR":               cms_nr,
                    "OHA Rate":             oha_rate,
                    "Δ Integra–Current":    delta,
                    "Δ%":                   pct,
                    "Flag":                 flag,
                }
            else:
                row = {
                    "HCPCS":                hcpcs,
                    "Mod":                  mod,
                    "Description":          desc_raw,
                    "Billing Unit":         unit_raw,
                    "Rental Rate":          _to_num(rent_raw) if mod == "RR" else "",
                    "Purchase Rate":        _to_num(purch_raw) if mod == "NU" else "",
                    "Comments":             comments,
                    # --- Comparison ---
                    "Current $":            cur_num,
                    "Integra Comm":         int_comm,
                    "Integra ASO":          int_aso,
                    "Integra Medicare":     int_med,
                    "Integra Medicaid":     int_mcd,
                    "CMS NR":               cms_nr,
                    "OHA Rate":             oha_rate,
                    "Δ Integra–Current":    delta,
                    "Δ%":                   pct,
                    "Flag":                 flag,
                }
            rows.append(row)

    return pd.DataFrame(rows)


def _write_contract_tab(ws: Worksheet, df: pd.DataFrame,
                        contract_col_count: int):
    """Write Contract View tab with blue headers (contract) / green (comparison)."""
    if df.empty:
        return
    cols = list(df.columns)

    CONTRACT_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    COMPARE_FILL  = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

    # Headers
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = CONTRACT_FILL if ci <= contract_col_count else COMPARE_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Identify currency / pct columns by name
    currency_names = {"Managed $", "Commercial $", "Current $",
                      "Integra Comm", "Integra ASO", "Integra Medicare",
                      "Integra Medicaid", "CMS NR", "OHA Rate",
                      "Δ Integra–Current"}
    pct_names = {"Δ%"}

    # Data rows
    for ri, (_, drow) in enumerate(df.iterrows(), 2):
        for ci, col_name in enumerate(cols, 1):
            val = drow[col_name]
            if isinstance(val, float) and math.isnan(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            if col_name in currency_names and val is not None:
                cell.number_format = CURRENCY
            elif col_name in pct_names and val is not None:
                cell.number_format = PCT_FMT

        # Flag coloring
        if "Flag" in cols:
            flag_ci = cols.index("Flag") + 1
            flag_val = str(drow.get("Flag", "") or "")
            fill = _flag_fill(flag_val)
            if fill:
                ws.cell(row=ri, column=flag_ci).fill = fill

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


# ───────────────────────────────────────────────────────────────────────
# 5.  XLSX FORMATTING
# ───────────────────────────────────────────────────────────────────────
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(bold=True, size=11)
CURRENCY    = '"$"#,##0.00'
PCT_FMT     = '0.0"%"'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

# Map flag keywords → fill color
FLAG_COLORS = {
    "BELOW CMS FLOOR": RED_FILL,
    "PHCC BELOW CMS":  ORANGE_FILL,
    "BELOW CURRENT":   YELLOW_FILL,
    "RATE INCREASE":   BLUE_FILL,
    "NO CHANGE":       GREEN_FILL,
    "NEW CODE":        GRAY_FILL,
    "REVIEW":          GRAY_FILL,
}


def _auto_width(ws, max_w=32):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col[:60]:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, max_w)


def _flag_fill(flag_text: str):
    """Return the fill color for a flag string (first matching keyword wins)."""
    if not flag_text:
        return None
    # Priority order matters: check most severe first
    for keyword in ("BELOW CMS FLOOR", "PHCC BELOW CMS",
                    "BELOW CURRENT", "RATE INCREASE",
                    "NO CHANGE", "NEW CODE", "REVIEW"):
        if keyword in flag_text:
            return FLAG_COLORS[keyword]
    return None


def _write_table(ws: Worksheet, df: pd.DataFrame, start_row: int) -> int:
    """Write df as a formatted table starting at start_row. Returns next row."""
    if len(df) == 0:
        ws.cell(row=start_row, column=1, value="No data").font = Font(italic=True)
        return start_row + 2

    cols = list(df.columns)
    # Drop OHA Rate column if all NaN
    if "OHA Rate" in cols and df["OHA Rate"].isna().all():
        df = df.drop(columns=["OHA Rate"])
        cols = list(df.columns)

    # Headers
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data
    for ri, (_, drow) in enumerate(df.iterrows(), start_row + 1):
        for ci, col_name in enumerate(cols, 1):
            val = drow[col_name]
            cell = ws.cell(row=ri, column=ci)
            if isinstance(val, float) and math.isnan(val):
                cell.value = None
            elif isinstance(val, (np.floating,)):
                cell.value = float(val) if not math.isnan(float(val)) else None
            elif isinstance(val, (np.bool_, bool)):
                cell.value = bool(val)
            elif isinstance(val, (np.integer,)):
                cell.value = int(val)
            else:
                cell.value = val
            cell.border = THIN_BORDER

    max_row = start_row + len(df)

    # Currency / pct formatting
    currency_cols = {"Proposed Rate", "PHCC Current", "Δ Proposed–PHCC",
                     "CMS NR", "CMS Rural", "Proposed–CMS NR",
                     "PHCC–CMS NR", "OHA Rate"}
    pct_cols = {"Δ%"}
    for ci, h in enumerate(cols, 1):
        if h in currency_cols:
            for r in range(start_row + 1, max_row + 1):
                ws.cell(row=r, column=ci).number_format = CURRENCY
        elif h in pct_cols:
            for r in range(start_row + 1, max_row + 1):
                ws.cell(row=r, column=ci).number_format = PCT_FMT

    # Row-level flag coloring
    flag_ci = cols.index("Flag") + 1 if "Flag" in cols else None
    if flag_ci:
        for r in range(start_row + 1, max_row + 1):
            fval = str(ws.cell(row=r, column=flag_ci).value or "")
            fill = _flag_fill(fval)
            if fill:
                ws.cell(row=r, column=flag_ci).fill = fill

    # Negative delta coloring
    for ci, h in enumerate(cols, 1):
        if h in ("Δ Proposed–PHCC", "Proposed–CMS NR", "PHCC–CMS NR"):
            for r in range(start_row + 1, max_row + 1):
                v = ws.cell(row=r, column=ci).value
                if v is not None and isinstance(v, (int, float)) and v < 0:
                    ws.cell(row=r, column=ci).fill = RED_FILL
                elif v is not None and isinstance(v, (int, float)) and v > 0:
                    ws.cell(row=r, column=ci).fill = GREEN_FILL

    # Freeze + filter
    ws.freeze_panes = ws.cell(row=start_row + 1, column=4)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(cols))}{max_row}"

    return max_row + 2


def _write_payer_summary(ws: Worksheet, df: pd.DataFrame, payer: str, row: int) -> int:
    """Write summary stats for one payer tab. Returns next row."""
    total = len(df)
    matched = len(df[df["Match"] != "NO_MATCH"]) if "Match" in df.columns else 0
    range_matched = len(df[df["Match"] == "T5_RANGE"]) if "Match" in df.columns else 0
    no_match = total - matched

    # Count flags
    flags = df["Flag"].astype(str) if "Flag" in df.columns else pd.Series(dtype=str)
    below_cms_floor = flags.str.contains("BELOW CMS FLOOR", na=False).sum()
    below_current = flags.str.contains("BELOW CURRENT", na=False, regex=False).sum()
    rate_increase = flags.str.contains("RATE INCREASE", na=False).sum()
    no_change = flags.str.contains("NO CHANGE", na=False).sum()
    phcc_below_cms = flags.str.contains("PHCC BELOW CMS", na=False).sum()
    medicare_resolved = (df["PHCC Raw"].astype(str)
                         .str.contains("→", na=False).sum()
                         if "PHCC Raw" in df.columns else 0)

    # Average deltas (matched only)
    m = df[df["Match"] != "NO_MATCH"] if "Match" in df.columns else df
    avg_delta = (m["Δ Proposed–PHCC"].dropna().mean()
                 if "Δ Proposed–PHCC" in m.columns else np.nan)

    ws.cell(row=row, column=1,
            value=f"Integra {payer} — Rate Analysis Summary").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1,
            value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = Font(
                italic=True, size=10, color="666666")
    row += 2

    # --- Flag legend ---
    ws.cell(row=row, column=1, value="FLAG LEGEND").font = Font(
        bold=True, size=12, color="C00000")
    row += 1
    legend = [
        ("BELOW CMS FLOOR", RED_FILL,
         "Proposed is below PHCC current AND below CMS Medicare floor — negotiate UP."),
        ("BELOW CURRENT", YELLOW_FILL,
         "Proposed is below PHCC current but at/above CMS floor."),
        ("RATE INCREASE", BLUE_FILL,
         "Proposed exceeds PHCC current rate (rate increase)."),
        ("NO CHANGE", GREEN_FILL,
         "Proposed ≈ PHCC current (within ±1%)."),
        ("PHCC BELOW CMS", ORANGE_FILL,
         "PHCC's current rate is already below CMS — systemic gap."),
        ("NEW CODE", GRAY_FILL,
         "No PHCC match — Integra proposing a rate for a code PHCC doesn't cover."),
    ]
    for label, fill, desc in legend:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=desc).font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="OVERVIEW").font = Font(
        bold=True, size=12, color="4472C4")
    row += 1

    items = [
        ("Total proposed codes", total),
        ("Matched to PHCC current", matched),
        ("  ↳ via range match (T5)", range_matched),
        ("  ↳ Medicare Allowable % resolved", medicare_resolved),
        ("No PHCC match (new codes)", no_match),
        ("", ""),
        ("BELOW CMS FLOOR (proposed < current & < CMS)", int(below_cms_floor)),
        ("BELOW CURRENT (proposed < current, ≥ CMS)", int(below_current)),
        ("RATE INCREASE (proposed > current)", int(rate_increase)),
        ("NO CHANGE (± 1%)", int(no_change)),
        ("PHCC BELOW CMS (systemic gap)", int(phcc_below_cms)),
        ("", ""),
        ("Avg Δ Proposed–PHCC (matched)",
         f"${avg_delta:.2f}" if not math.isnan(avg_delta) else "N/A"),
    ]
    for label, val in items:
        if label == "":
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=row, column=2, value=val)
        if isinstance(val, str) and "$" in val and "-" in val:
            cell.fill = RED_FILL
        row += 1

    # By-state breakdown
    row += 1
    ws.cell(row=row, column=1, value="BY STATE").font = Font(
        bold=True, size=11, color="4472C4")
    row += 1
    sh = ["State", "Total", "Matched", "Range(T5)",
          "Below CMS Floor", "Below Current", "Rate↑",
          "No Change", "PHCC<CMS", "Avg Δ"]
    for c, h in enumerate(sh, 1):
        ws.cell(row=row, column=c, value=h).font = BOLD_FONT
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

    for state in sorted(df["State"].unique()):
        sg = df[df["State"] == state]
        sm = sg[sg["Match"] != "NO_MATCH"]
        sf = sg["Flag"].astype(str)
        a_delta = sm["Δ Proposed–PHCC"].dropna().mean()
        ws.cell(row=row, column=1, value=state)
        ws.cell(row=row, column=2, value=len(sg))
        ws.cell(row=row, column=3, value=len(sm))
        ws.cell(row=row, column=4,
                value=int((sg["Match"] == "T5_RANGE").sum()))
        ws.cell(row=row, column=5,
                value=int(sf.str.contains("BELOW CMS FLOOR", na=False).sum()))
        ws.cell(row=row, column=6,
                value=int(sf.str.contains("BELOW CURRENT", na=False, regex=False).sum()))
        ws.cell(row=row, column=7,
                value=int(sf.str.contains("RATE INCREASE", na=False).sum()))
        ws.cell(row=row, column=8,
                value=int(sf.str.contains("NO CHANGE", na=False).sum()))
        ws.cell(row=row, column=9,
                value=int(sf.str.contains("PHCC BELOW CMS", na=False).sum()))
        ws.cell(row=row, column=10,
                value=round(a_delta, 2) if not math.isnan(a_delta) else 0)
        ws.cell(row=row, column=10).number_format = CURRENCY
        for c in range(1, 11):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1
    return row


# ───────────────────────────────────────────────────────────────────────
# 6.  EXECUTIVE SUMMARY TAB
# ───────────────────────────────────────────────────────────────────────

def _write_exec_summary(ws: Worksheet, payer_tables: dict[str, pd.DataFrame]):
    """Cross-payer executive overview."""
    ws.cell(row=1, column=1,
            value="Integra PHP FFS — Executive Rate Analysis  v2").font = Font(
                bold=True, size=16)
    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = Font(
                italic=True, size=10, color="666666")

    row = 4
    ws.cell(row=row, column=1, value="METHODOLOGY").font = Font(
        bold=True, size=12, color="C00000")
    row += 1
    method = [
        "• Integra PHP proposed rates compared to PHCC current contracted rates.",
        "• CMS 2026 Q1 DMEPOS Fee Schedule (Non-Rural + Rural) as Medicare benchmark.",
        "• OHA FFS Sept 2025 as Medicaid benchmark (OR only).",
        "• Match tiers: T1=exact, T2=mod→blank, T3=cross-mod, T4=HCPCS-only, T5=range.",
        '• "Medicare Allowable less X%" resolved to CMS NR × (1 − X/100).',
        "",
        "DECISION TREE:",
        "  1. Match Integra code → PHCC current contract (T1–T5)",
        "  2. Compare Proposed vs PHCC Current:",
        "     • ±1%            → NO CHANGE     (green)",
        "     • Proposed > Cur  → RATE INCREASE  (blue)",
        "     • Proposed < Cur  → check CMS benchmark →",
        "        ○ ≥ CMS NR    → BELOW CURRENT  (yellow)",
        "        ○ < CMS NR    → BELOW CMS FLOOR (red)",
        "  3. If PHCC Current < CMS → PHCC BELOW CMS (orange, systemic gap)",
        "",
        "FLAG COLOR KEY:",
        "  🔴 BELOW CMS FLOOR — Proposed below current AND below Medicare floor.",
        "  🟠 PHCC BELOW CMS — PHCC's current rate already below Medicare floor.",
        "  🟡 BELOW CURRENT — Proposed below current, but at/above CMS floor.",
        "  🔵 RATE INCREASE — Proposed exceeds PHCC current.",
        "  🟢 NO CHANGE — Proposed ≈ current (within ±1%).",
        "  ⚪ NEW CODE — No PHCC match; Integra proposing for uncovered code.",
    ]
    for line in method:
        ws.cell(row=row, column=1, value=line).font = Font(size=10)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="CROSS-PAYER SUMMARY").font = Font(
        bold=True, size=12, color="4472C4")
    row += 1

    hdr = ["Payer", "Total", "Matched", "Range(T5)", "New Codes",
           "Below CMS Floor", "Below Current", "Rate Increase",
           "No Change", "PHCC<CMS", "Avg Δ"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    for payer_name in ["Commercial", "ASO", "Medicare", "Medicaid"]:
        df = payer_tables.get(payer_name, pd.DataFrame())
        if len(df) == 0:
            continue
        total = len(df)
        matched = len(df[df["Match"] != "NO_MATCH"])
        range_m = int((df["Match"] == "T5_RANGE").sum())
        no_match = total - matched
        flags = df["Flag"].astype(str)
        bcf = int(flags.str.contains("BELOW CMS FLOOR", na=False).sum())
        bc = int(flags.str.contains("BELOW CURRENT", na=False, regex=False).sum())
        ri_ = int(flags.str.contains("RATE INCREASE", na=False).sum())
        nc = int(flags.str.contains("NO CHANGE", na=False).sum())
        pcb = int(flags.str.contains("PHCC BELOW CMS", na=False).sum())
        m = df[df["Match"] != "NO_MATCH"]
        avg_d = m["Δ Proposed–PHCC"].dropna().mean()

        vals = [payer_name, total, matched, range_m, no_match,
                bcf, bc, ri_, nc, pcb,
                f"${avg_d:.2f}" if not math.isnan(avg_d) else "N/A"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
        # Color key cells
        if bcf > 0:
            ws.cell(row=row, column=6).fill = RED_FILL
        if pcb > 0:
            ws.cell(row=row, column=10).fill = ORANGE_FILL
        row += 1

    # Highest-impact codes across all payers
    row += 2
    ws.cell(row=row, column=1,
            value="TOP ATTENTION CODES (Below CMS Floor — All Payers)"
            ).font = Font(bold=True, size=12)
    row += 1

    all_below = []
    for payer_name, df in payer_tables.items():
        if len(df) == 0:
            continue
        below = df[df["Flag"].astype(str).str.contains(
            "BELOW CMS FLOOR", na=False)].copy()
        below["Payer"] = payer_name
        all_below.append(below)

    if all_below:
        combined = pd.concat(all_below, ignore_index=True)
        combined = combined.sort_values("Proposed–CMS NR", ascending=True)
        top_cols = ["Payer", "State", "HCPCS", "Mod", "Description",
                    "Proposed Rate", "PHCC Current", "CMS NR",
                    "Proposed–CMS NR", "Flag"]
        top_cols = [c for c in top_cols if c in combined.columns]
        top = combined[top_cols].head(25)

        for c, h in enumerate(top_cols, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
        row += 1
        for _, tr in top.iterrows():
            for c, col in enumerate(top_cols, 1):
                val = tr[col]
                cell = ws.cell(row=row, column=c)
                if isinstance(val, float) and math.isnan(val):
                    cell.value = None
                elif isinstance(val, (np.floating,)):
                    cell.value = float(val) if not math.isnan(float(val)) else None
                else:
                    cell.value = val
                cell.border = THIN_BORDER
            row += 1
        for c, h in enumerate(top_cols, 1):
            if h in ("Proposed Rate", "PHCC Current", "CMS NR", "Proposed–CMS NR"):
                for r in range(row - len(top), row):
                    ws.cell(row=r, column=c).number_format = CURRENCY
    else:
        ws.cell(row=row, column=1,
                value="No codes below CMS benchmark.").font = Font(italic=True)

    _auto_width(ws)


# ───────────────────────────────────────────────────────────────────────
# 7.  MAIN
# ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("INTEGRA PHP FFS — Executive Rate Analysis  v2")
    print(f"Run: {datetime.now():%Y-%m-%d %H:%M}")
    print("=" * 70)

    # Verify cleaned data
    for key in ("or_contracted", "or_participating", "wa_participating"):
        if not FILES[key].exists():
            print(f"ERROR: {FILES[key]} not found. Run clean_phcc_files.py first.")
            sys.exit(1)

    # Load PHCC cleaned
    print("\n[1] Loading PHCC cleaned schedules…")
    phcc_dfs = {}
    phcc_key_lks, phcc_code_lks, phcc_range_lks = {}, {}, {}
    for key, label in [("or_contracted", "PHCC_OR_CONTRACTED"),
                        ("or_participating", "PHCC_OR_PARTICIPATING"),
                        ("wa_participating", "PHCC_WA_PARTICIPATING")]:
        df = load_cleaned_phcc(FILES[key], label)
        phcc_dfs[key] = df
        k, c, rng = build_phcc(df)
        phcc_key_lks[key] = k
        phcc_code_lks[key] = c
        phcc_range_lks[key] = rng
        print(f"    {label}: {len(df)} rows, {len(rng)} range entries")

    # Load Integra
    print("\n[2] Loading Integra proposed…")
    integra_dfs = {}
    for key, rate_col in [("integra_commercial", "Commercial"),
                           ("integra_aso", "ASO/Commercial"),
                           ("integra_medicare", "Medicare"),
                           ("integra_medicaid", "Medicaid")]:
        integra_dfs[key] = load_integra(FILES[key], rate_col)
        print(f"    {key}: {len(integra_dfs[key])} rows")

    # Load benchmarks
    print("\n[3] Loading benchmarks…")
    cms_or = load_cms(FILES["cms_or"], "OR (NR)", "OR (R)")
    cms_wa = load_cms(FILES["cms_wa"], "WA (NR)", "WA (R)")
    oha = load_oha(FILES["oha"])
    hcpcs_desc = load_hcpcs_desc(FILES["hcpcs"])
    print(f"    CMS OR: {len(cms_or)} keys, CMS WA: {len(cms_wa)} keys")
    print(f"    OHA: {len(oha)} keys, HCPCS descriptions: {len(hcpcs_desc)}")

    # Build per-payer tables
    print("\n[4] Building comparison tables…")
    payer_tables = {}
    for payer, configs in PAYER_CONFIGS.items():
        df = build_payer_table(
            payer, configs,
            integra_dfs, phcc_dfs, phcc_key_lks, phcc_code_lks,
            phcc_range_lks,
            cms_or, cms_wa, oha, hcpcs_desc)
        payer_tables[payer] = df
        matched = len(df[df["Match"] != "NO_MATCH"]) if len(df) > 0 else 0
        range_m = int((df["Match"] == "T5_RANGE").sum()) if len(df) > 0 else 0
        flags = df["Flag"].astype(str) if len(df) > 0 else pd.Series(dtype=str)
        bcf = int(flags.str.contains("BELOW CMS FLOOR", na=False).sum())
        bc = int(flags.str.contains("BELOW CURRENT", na=False, regex=False).sum())
        print(f"    {payer}: {len(df)} rows, {matched} matched "
              f"({range_m} range), "
              f"{bcf} below CMS floor, {bc} below current")

    # Build Integra lookup for Contract View
    print("\n[5] Building Integra lookup for Contract View…")
    integra_lk = _build_integra_lk(integra_dfs)
    print(f"    Integra lookup: {len(integra_lk)} HCPCS|Mod keys")

    # Load raw contract files + build Contract Views
    print("\n[6] Building Contract View tabs…")
    cv_configs = [
        ("or_contracted",    "CV OR Contracted",    9),
        ("or_participating", "CV OR Participating",  7),
        ("wa_participating", "CV WA Participating",  7),
    ]
    cv_tables = {}
    for sched_key, tab_name, n_contract_cols in cv_configs:
        raw_path = FILES[f"{sched_key}_raw"]
        if not raw_path.exists():
            print(f"    SKIP {tab_name}: {raw_path} not found")
            continue
        raw_df = load_raw_contract(raw_path)
        cms_lk = cms_or if "wa" not in sched_key else cms_wa
        cv_df = build_contract_view(
            raw_df, sched_key, integra_lk,
            phcc_key_lks[sched_key], phcc_code_lks[sched_key],
            phcc_range_lks[sched_key],
            cms_lk, oha, hcpcs_desc)
        cv_tables[tab_name] = (cv_df, n_contract_cols)
        print(f"    {tab_name}: {len(raw_df)} raw rows -> {len(cv_df)} view rows")

    # Write XLSX
    print("\n[7] Writing Excel workbook…")
    out_path = OUTPUT / "integra_rate_analysis_v2.xlsx"
    wb = Workbook()

    # Tab 1: Executive Summary
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    _write_exec_summary(ws_exec, payer_tables)

    # Tabs 2-5: Per-payer detail
    for payer_name in ["Commercial", "ASO", "Medicare", "Medicaid"]:
        df = payer_tables[payer_name]
        ws = wb.create_sheet(payer_name)
        next_row = _write_payer_summary(ws, df, payer_name, 1)
        ws.cell(row=next_row, column=1,
                value="DETAIL TABLE").font = Font(bold=True, size=12)
        _write_table(ws, df, next_row + 1)
        _auto_width(ws)
        print(f"    Tab: {payer_name} — {len(df)} rows")

    # Tabs 6-8: Contract View
    for tab_name, (cv_df, n_contract_cols) in cv_tables.items():
        ws = wb.create_sheet(tab_name)
        _write_contract_tab(ws, cv_df, n_contract_cols)
        print(f"    Tab: {tab_name} — {len(cv_df)} rows")

    wb.save(out_path)
    print(f"\n[XLSX] {out_path.name} saved with {len(wb.sheetnames)} tabs")
    print(f"    Tabs: {', '.join(wb.sheetnames)}")
    print(f"\n{'='*70}")
    print(f"DONE — {out_path}")
    print(f"{'='*70}")
    return out_path


if __name__ == "__main__":
    main()
