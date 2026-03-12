"""
INTEGRA PHP FFS — Executive Rate Analysis
==========================================
Produces a focused Excel workbook with one tab per Integra payer type,
showing proposed rates vs PHCC current vs CMS benchmarks with
color-coded flags for exec decision-making.

Prerequisite: python scripts/clean_phcc_files.py
Run:          python scripts/integra_rate_analysis.py

Output: PHCC/output/integra_rate_analysis.xlsx
"""
from __future__ import annotations
import re, os, sys, math
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ───────────────────────────────────────────────────────────────────────
# 0.  PATHS & CONFIG
# ───────────────────────────────────────────────────────────────────────
PHCC_ROOT = Path(__file__).resolve().parent.parent
CLEANED   = PHCC_ROOT / "data" / "cleaned"
INTEGRA   = PHCC_ROOT / "data" / "INTEGRA_PHP_FFS"
CMS_DIR   = PHCC_ROOT / "data" / "cms"
OUTPUT    = PHCC_ROOT / "output"
OUTPUT.mkdir(exist_ok=True)

FILES = {
    "or_contracted":    CLEANED / "PHCC_OR_CONTRACTED_CLEAN.csv",
    "or_participating": CLEANED / "PHCC_OR_PARTICIPATING_CLEAN.csv",
    "wa_participating": CLEANED / "PHCC_WA_PARTICIPATING_CLEAN.csv",
    "integra_commercial": INTEGRA / "Integra_PHP_CARVEOUTS_COMMERCIAL.csv",
    "integra_aso":        INTEGRA / "Integra_PHP_CARVEOUTS_ASO.csv",
    "integra_medicare":   INTEGRA / "Integra_PHP_CARVEOUTS_MEDICARE.csv",
    "integra_medicaid":   INTEGRA / "INTEGRA_PHP_CARVEOUTS_MEDICAID.csv",
    "cms_or": CMS_DIR / "CMS_2026_Q1_OR.csv",
    "cms_wa": CMS_DIR / "CMS_2026_Q1_WA.csv",
    "oha":    CMS_DIR / "OHA_FFS_09_2025_RAW.csv",
    "hcpcs":  CMS_DIR / "2026_CMS_HCPCS.csv",
}

# Payer configs: (integra_key, payer_label, integra_rate_col,
#                  phcc_keys, state, benchmark_type)
PAYER_CONFIGS = {
    "Commercial": [
        ("integra_commercial", "Commercial", "Commercial",
         ["or_contracted", "or_participating"], "OR", "CMS"),
        ("integra_commercial", "Commercial", "Commercial",
         ["wa_participating"], "WA", "CMS"),
    ],
    "ASO": [
        ("integra_aso", "ASO", "ASO/Commercial",
         ["or_contracted", "or_participating"], "OR", "CMS"),
        ("integra_aso", "ASO", "ASO/Commercial",
         ["wa_participating"], "WA", "CMS"),
    ],
    "Medicare": [
        ("integra_medicare", "Medicare", "Medicare",
         ["or_contracted", "or_participating"], "OR", "CMS"),
        ("integra_medicare", "Medicare", "Medicare",
         ["wa_participating"], "WA", "CMS"),
    ],
    "Medicaid": [
        ("integra_medicaid", "Medicaid", "Medicaid",
         ["or_contracted", "or_participating"], "OR", "OHA"),
        ("integra_medicaid", "Medicaid", "Medicaid",
         ["wa_participating"], "WA", None),
    ],
}

# ───────────────────────────────────────────────────────────────────────
# 1.  PURE HELPERS (reused from analyze_fee_schedules)
# ───────────────────────────────────────────────────────────────────────
VALID_HCPCS_RE = re.compile(r'^[A-Z][0-9]{4}$')

def _norm(raw) -> str:
    if pd.isna(raw) or str(raw).strip() == "":
        return ""
    return str(raw).strip().upper()

def _safe_float(val) -> float:
    if pd.isna(val):
        return np.nan
    s = str(val).strip().replace("$", "").replace(",", "")
    if s == "" or s.lower() == "nan":
        return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan

def _norm_mod(raw) -> str:
    if pd.isna(raw) or str(raw).strip() == "":
        return ""
    return str(raw).strip().upper().rstrip("*")

def _classify_note(val) -> tuple[float, str]:
    """Return (numeric_or_nan, note_text)."""
    if pd.isna(val):
        return np.nan, ""
    s = str(val).strip()
    if s == "":
        return np.nan, ""
    cleaned = s.replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned), ""
    except ValueError:
        return np.nan, s

CROSS_MOD = {"NU": ["RR", ""], "RR": ["NU", ""], "AU": ["NU", ""],
             "KF": ["NU", ""], "": ["NU", "RR"]}

# ───────────────────────────────────────────────────────────────────────
# 2.  LOADERS
# ───────────────────────────────────────────────────────────────────────

def load_cleaned_phcc(path: Path, label: str) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False)

def _pick_rate(row, schedule_label: str, payer: str, modifier: str) -> tuple[float, str]:
    """Return (numeric_rate, raw_value)."""
    if schedule_label == "PHCC_OR_CONTRACTED":
        prefix = "Managed" if payer in ("Medicare", "Medicaid") else "Commercial"
        rate_col = f"{prefix} Rental Rate" if modifier == "RR" else f"{prefix} Purchase Rate"
    else:
        rate_col = "Rental Rate" if modifier == "RR" else "Purchase Rate"

    raw = str(row.get(f"{rate_col}_raw", ""))
    num = _safe_float(row.get(f"{rate_col}_numeric", ""))
    return num, raw

def load_integra(path: Path, rate_col: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df.columns = [c.strip() for c in df.columns]
    rows = []
    for idx, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod1 = _norm_mod(r.get("Mod 1", ""))
        raw = str(r.get(rate_col, "")).strip()
        num, note = _classify_note(raw)
        rows.append({
            "hcpcs": hcpcs, "mod": mod1,
            "proposed_rate": num, "proposed_raw": raw, "proposed_note": note,
        })
    return pd.DataFrame(rows)

def load_cms(path: Path, nr_col: str, r_col: str) -> dict[str, dict]:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    lk = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod = _norm_mod(r.get("Mod", ""))
        if not hcpcs:
            continue
        lk[f"{hcpcs}|{mod}"] = {
            "nr": _safe_float(r.get(nr_col, "")),
            "r":  _safe_float(r.get(r_col, "")),
        }
    return lk

def load_oha(path: Path) -> dict[str, float]:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    lk = {}
    for _, r in df.iterrows():
        code = _norm(r.get("Procedure Code", ""))
        mod = _norm_mod(r.get("Mod1", ""))
        if code:
            lk[f"{code}|{mod}"] = _safe_float(r.get("Price", ""))
    return lk

def load_hcpcs_desc(path: Path) -> dict[str, str]:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    return {_norm(r.get("HCPC", "")): str(r.get("SHORT DESCRIPTION", "")).strip()
            for _, r in df.iterrows() if _norm(r.get("HCPC", ""))}


# ───────────────────────────────────────────────────────────────────────
# 3.  PHCC LOOKUP + MATCHING (simplified best-match)
# ───────────────────────────────────────────────────────────────────────

def build_phcc(df: pd.DataFrame):
    """Return (key_lk: {HCPCS|MOD → row}, code_lk: {HCPCS → [rows]})."""
    key_lk, code_lk = {}, {}
    for _, row in df.iterrows():
        h = str(row.get("hcpcs_normalised", "")).strip()
        m = str(row.get("modifier_normalised", "")).strip()
        key_lk.setdefault(f"{h}|{m}", []).append(row)
        code_lk.setdefault(h, []).append(row)
    return key_lk, code_lk


def best_match(hcpcs: str, mod: str, key_lk: dict, code_lk: dict):
    """Return (phcc_row_or_None, tier_label, cross_mod_used)."""
    # T1: exact
    rows = key_lk.get(f"{hcpcs}|{mod}", [])
    if rows:
        return rows[0], "T1", ""
    # T2: proposed mod → blank PHCC
    if mod:
        rows = key_lk.get(f"{hcpcs}|", [])
        if rows:
            return rows[0], "T2", ""
    # T3: cross-modifier
    for alt in CROSS_MOD.get(mod, ["NU", "RR", ""]):
        if alt == mod or (alt == "" and mod):
            continue
        rows = key_lk.get(f"{hcpcs}|{alt}", [])
        if rows:
            return rows[0], "T3", alt
    # T4: any row for that HCPCS
    rows = code_lk.get(hcpcs, [])
    if rows:
        return rows[0], "T4", str(rows[0].get("modifier_normalised", ""))
    return None, "NO_MATCH", ""


def cms_cascade(hcpcs: str, mod: str, cms_lk: dict) -> tuple[float, float, str]:
    """B1→B4 cascade. Returns (nr_rate, rural_rate, match_tier)."""
    for tier, try_mod in [("B1", mod), ("B2", "NU"), ("B3", "RR"), ("B4", "")]:
        rec = cms_lk.get(f"{hcpcs}|{try_mod}")
        if rec:
            return rec["nr"], rec["r"], tier
    return np.nan, np.nan, "NOT_FOUND"


# ───────────────────────────────────────────────────────────────────────
# 4.  BUILD COMPARISON TABLE PER PAYER
# ───────────────────────────────────────────────────────────────────────

def build_payer_table(
    payer: str,
    configs: list,
    integra_dfs: dict,
    phcc_dfs: dict, phcc_key_lks: dict, phcc_code_lks: dict,
    cms_or: dict, cms_wa: dict, oha_lk: dict,
    hcpcs_desc: dict,
) -> pd.DataFrame:
    """Build the slim comparison rows for one payer type."""

    rows = []
    for (integra_key, payer_label, rate_col,
         phcc_keys, state, bench_type) in configs:

        integra_df = integra_dfs[integra_key]
        cms_lk = cms_or if state == "OR" else cms_wa

        for _, prop in integra_df.iterrows():
            hcpcs = prop["hcpcs"]
            mod = prop["mod"]
            proposed_num = prop["proposed_rate"]
            proposed_raw = prop["proposed_raw"]
            proposed_note = prop["proposed_note"]

            if not hcpcs:
                continue

            # CMS benchmark
            cms_nr, cms_r, cms_tier = cms_cascade(hcpcs, mod, cms_lk)

            # OHA benchmark (Medicaid only)
            oha_rate = np.nan
            if bench_type == "OHA":
                for oha_tier, try_mod in [("B1", mod), ("B2", "NU"), ("B3", "RR"), ("B4", "")]:
                    oha_val = oha_lk.get(f"{hcpcs}|{try_mod}", np.nan)
                    if not math.isnan(oha_val):
                        oha_rate = oha_val
                        break

            # Match against each PHCC schedule
            for sched_key in phcc_keys:
                label = {
                    "or_contracted": "PHCC_OR_CONTRACTED",
                    "or_participating": "PHCC_OR_PARTICIPATING",
                    "wa_participating": "PHCC_WA_PARTICIPATING",
                }[sched_key]

                phcc_row, tier, cross_mod = best_match(
                    hcpcs, mod,
                    phcc_key_lks[sched_key], phcc_code_lks[sched_key])

                if phcc_row is not None:
                    eff_mod = cross_mod if cross_mod else mod
                    cur_num, cur_raw = _pick_rate(phcc_row, label, payer, eff_mod)
                else:
                    cur_num, cur_raw = np.nan, ""

                # Compute deltas
                delta_prop_phcc = np.nan
                pct_prop_phcc = np.nan
                if not math.isnan(proposed_num) and not math.isnan(cur_num):
                    delta_prop_phcc = proposed_num - cur_num
                    if cur_num != 0:
                        pct_prop_phcc = delta_prop_phcc / cur_num * 100

                delta_cur_cms_nr = np.nan
                if not math.isnan(cur_num) and not math.isnan(cms_nr):
                    delta_cur_cms_nr = cur_num - cms_nr

                delta_prop_cms_nr = np.nan
                if not math.isnan(proposed_num) and not math.isnan(cms_nr):
                    delta_prop_cms_nr = proposed_num - cms_nr

                # Flags
                flags = []
                if not math.isnan(proposed_num) and not math.isnan(cur_num):
                    if cur_num > proposed_num:
                        flags.append("PHCC ABOVE PROPOSED")
                    elif cur_num < proposed_num:
                        flags.append("PROPOSED ABOVE PHCC")
                if not math.isnan(proposed_num) and not math.isnan(cms_nr):
                    if proposed_num < cms_nr:
                        flags.append("PROPOSED BELOW CMS")
                if not math.isnan(cur_num) and not math.isnan(cms_nr):
                    if cur_num < cms_nr:
                        flags.append("PHCC BELOW CMS")

                flag_str = " | ".join(flags) if flags else ""

                # Determine primary flag color category for sorting
                if "PROPOSED BELOW CMS" in flags and "PHCC BELOW CMS" in flags:
                    sort_priority = 0  # Both below CMS — worst
                elif "PROPOSED BELOW CMS" in flags:
                    sort_priority = 1
                elif "PHCC BELOW CMS" in flags:
                    sort_priority = 2
                elif "PHCC ABOVE PROPOSED" in flags:
                    sort_priority = 3
                elif not flags and tier == "NO_MATCH":
                    sort_priority = 5
                else:
                    sort_priority = 4

                rows.append({
                    "State": state,
                    "Schedule": label.replace("PHCC_", ""),
                    "HCPCS": hcpcs,
                    "Mod": mod,
                    "Description": hcpcs_desc.get(hcpcs, ""),
                    "Proposed Rate": proposed_num,
                    "PHCC Current": cur_num,
                    "Δ Proposed vs PHCC": delta_prop_phcc,
                    "Δ%": pct_prop_phcc,
                    "CMS NR": cms_nr,
                    "CMS Rural": cms_r,
                    "Proposed vs CMS NR": delta_prop_cms_nr,
                    "PHCC vs CMS NR": delta_cur_cms_nr,
                    "OHA Rate": oha_rate if bench_type == "OHA" else np.nan,
                    "Flag": flag_str,
                    "Match": tier,
                    "Note": proposed_note if proposed_note else "",
                    "_sort": sort_priority,
                })

    df = pd.DataFrame(rows)
    if len(df) > 0:
        df = df.sort_values(["_sort", "HCPCS", "State"], ascending=True)
        df = df.drop(columns=["_sort"])
    return df


# ───────────────────────────────────────────────────────────────────────
# 5.  XLSX FORMATTING
# ───────────────────────────────────────────────────────────────────────
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(bold=True, size=11)
CURRENCY    = '"$"#,##0.00'
PCT_FMT     = '0.0"%"'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))


def _auto_width(ws, max_w=32):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col[:60]:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, max_w)


def _write_table(ws: Worksheet, df: pd.DataFrame, start_row: int) -> int:
    """Write df as a formatted table starting at start_row. Returns next row."""
    if len(df) == 0:
        ws.cell(row=start_row, column=1, value="No data").font = Font(italic=True)
        return start_row + 2

    cols = list(df.columns)
    # Drop OHA Rate column if all NaN
    if "OHA Rate" in cols and df["OHA Rate"].isna().all():
        df = df.drop(columns=["OHA Rate"])
        cols = list(df.columns)

    # Headers
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data
    for ri, (_, drow) in enumerate(df.iterrows(), start_row + 1):
        for ci, col_name in enumerate(cols, 1):
            val = drow[col_name]
            cell = ws.cell(row=ri, column=ci)
            if isinstance(val, float) and math.isnan(val):
                cell.value = None
            elif isinstance(val, (np.floating,)):
                cell.value = float(val) if not math.isnan(float(val)) else None
            elif isinstance(val, (np.bool_, bool)):
                cell.value = bool(val)
            elif isinstance(val, (np.integer,)):
                cell.value = int(val)
            else:
                cell.value = val
            cell.border = THIN_BORDER

    max_row = start_row + len(df)

    # Currency / pct formatting
    currency_cols = {"Proposed Rate", "PHCC Current", "Δ Proposed vs PHCC",
                     "CMS NR", "CMS Rural", "Proposed vs CMS NR",
                     "PHCC vs CMS NR", "OHA Rate"}
    pct_cols = {"Δ%"}
    for ci, h in enumerate(cols, 1):
        if h in currency_cols:
            for r in range(start_row + 1, max_row + 1):
                ws.cell(row=r, column=ci).number_format = CURRENCY
        elif h in pct_cols:
            for r in range(start_row + 1, max_row + 1):
                ws.cell(row=r, column=ci).number_format = PCT_FMT

    # Row-level flag coloring
    flag_ci = cols.index("Flag") + 1 if "Flag" in cols else None
    if flag_ci:
        for r in range(start_row + 1, max_row + 1):
            fval = str(ws.cell(row=r, column=flag_ci).value or "")
            if "PROPOSED BELOW CMS" in fval and "PHCC BELOW CMS" in fval:
                fill = RED_FILL
            elif "PROPOSED BELOW CMS" in fval:
                fill = RED_FILL
            elif "PHCC BELOW CMS" in fval:
                fill = YELLOW_FILL
            elif "PHCC ABOVE PROPOSED" in fval:
                fill = GREEN_FILL
            else:
                fill = None
            if fill:
                ws.cell(row=r, column=flag_ci).fill = fill

    # Negative delta coloring
    for ci, h in enumerate(cols, 1):
        if h in ("Δ Proposed vs PHCC", "Proposed vs CMS NR", "PHCC vs CMS NR"):
            for r in range(start_row + 1, max_row + 1):
                v = ws.cell(row=r, column=ci).value
                if v is not None and isinstance(v, (int, float)) and v < 0:
                    ws.cell(row=r, column=ci).fill = RED_FILL
                elif v is not None and isinstance(v, (int, float)) and v > 0:
                    ws.cell(row=r, column=ci).fill = GREEN_FILL

    # Freeze + filter
    ws.freeze_panes = ws.cell(row=start_row + 1, column=4)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(cols))}{max_row}"

    return max_row + 2


def _write_payer_summary(ws: Worksheet, df: pd.DataFrame, payer: str, row: int) -> int:
    """Write summary stats for one payer tab. Returns next row."""
    total = len(df)
    matched = len(df[df["Match"] != "NO_MATCH"]) if "Match" in df.columns else 0
    no_match = total - matched

    # Count flags
    flags = df["Flag"].astype(str) if "Flag" in df.columns else pd.Series(dtype=str)
    phcc_above = flags.str.contains("PHCC ABOVE PROPOSED", na=False).sum()
    prop_below_cms = flags.str.contains("PROPOSED BELOW CMS", na=False).sum()
    phcc_below_cms = flags.str.contains("PHCC BELOW CMS", na=False).sum()
    prop_above_phcc = flags.str.contains("PROPOSED ABOVE PHCC", na=False).sum()

    # Average deltas (matched only)
    m = df[df["Match"] != "NO_MATCH"] if "Match" in df.columns else df
    avg_delta = m["Δ Proposed vs PHCC"].dropna().mean() if "Δ Proposed vs PHCC" in m.columns else np.nan

    ws.cell(row=row, column=1,
            value=f"Integra {payer} — Rate Analysis Summary").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1,
            value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = Font(italic=True, size=10, color="666666")
    row += 2

    ws.cell(row=row, column=1, value="FLAG LEGEND").font = Font(bold=True, size=12, color="C00000")
    row += 1
    legend = [
        ("PHCC ABOVE PROPOSED", GREEN_FILL,
         "PHCC's current rate exceeds Integra's proposed — rate decrease."),
        ("PROPOSED BELOW CMS", RED_FILL,
         "Integra proposes below CMS Medicare floor — negotiate UP."),
        ("PHCC BELOW CMS", YELLOW_FILL,
         "PHCC's current rate is already below CMS — systemic gap."),
    ]
    for label, fill, desc in legend:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=desc).font = Font(size=10)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="OVERVIEW").font = Font(bold=True, size=12, color="4472C4")
    row += 1

    items = [
        ("Total proposed codes", total),
        ("Matched to PHCC current", matched),
        ("No PHCC match (new codes)", no_match),
        ("", ""),
        ("PHCC ABOVE PROPOSED (rate cut)", phcc_above),
        ("PROPOSED ABOVE PHCC (rate increase)", prop_above_phcc),
        ("PROPOSED BELOW CMS (below floor)", prop_below_cms),
        ("PHCC BELOW CMS (systemic gap)", phcc_below_cms),
        ("", ""),
        ("Avg Δ Proposed vs PHCC (matched)",
         f"${avg_delta:.2f}" if not math.isnan(avg_delta) else "N/A"),
    ]
    for label, val in items:
        if label == "":
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=row, column=2, value=val)
        if isinstance(val, str) and "$" in val and "-" in val:
            cell.fill = RED_FILL
        row += 1

    # By-state breakdown
    row += 1
    ws.cell(row=row, column=1, value="BY STATE").font = Font(bold=True, size=11, color="4472C4")
    row += 1
    sh = ["State", "Total", "Matched", "PHCC>Proposed", "Prop<CMS", "PHCC<CMS", "Avg Δ"]
    for c, h in enumerate(sh, 1):
        ws.cell(row=row, column=c, value=h).font = BOLD_FONT
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

    for state in sorted(df["State"].unique()):
        sg = df[df["State"] == state]
        sm = sg[sg["Match"] != "NO_MATCH"]
        sf = sg["Flag"].astype(str)
        a_delta = sm["Δ Proposed vs PHCC"].dropna().mean()
        ws.cell(row=row, column=1, value=state)
        ws.cell(row=row, column=2, value=len(sg))
        ws.cell(row=row, column=3, value=len(sm))
        ws.cell(row=row, column=4, value=int(sf.str.contains("PHCC ABOVE PROPOSED", na=False).sum()))
        ws.cell(row=row, column=5, value=int(sf.str.contains("PROPOSED BELOW CMS", na=False).sum()))
        ws.cell(row=row, column=6, value=int(sf.str.contains("PHCC BELOW CMS", na=False).sum()))
        ws.cell(row=row, column=7,
                value=round(a_delta, 2) if not math.isnan(a_delta) else 0)
        ws.cell(row=row, column=7).number_format = CURRENCY
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1
    return row


# ───────────────────────────────────────────────────────────────────────
# 6.  EXECUTIVE SUMMARY TAB
# ───────────────────────────────────────────────────────────────────────

def _write_exec_summary(ws: Worksheet, payer_tables: dict[str, pd.DataFrame]):
    """Cross-payer executive overview."""
    ws.cell(row=1, column=1,
            value="Integra PHP FFS — Executive Rate Analysis").font = Font(bold=True, size=16)
    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = Font(italic=True, size=10, color="666666")

    row = 4
    ws.cell(row=row, column=1, value="METHODOLOGY").font = Font(bold=True, size=12, color="C00000")
    row += 1
    method = [
        "• Integra PHP proposed rates compared to PHCC current contracted rates.",
        "• CMS 2026 Q1 DMEPOS Fee Schedule (Non-Rural + Rural) used as Medicare benchmark.",
        "• OHA FFS Sept 2025 used as Medicaid benchmark (OR only).",
        "• Match tiers: T1=exact, T2=mod→blank, T3=cross-modifier, T4=HCPCS-only.",
        "• Codes with no PHCC match = Integra proposing rates for codes PHCC doesn't currently cover.",
        "",
        "FLAG DEFINITIONS:",
        "  🟢 PHCC ABOVE PROPOSED — Our rate exceeds Integra's offer (rate decrease).",
        "  🔴 PROPOSED BELOW CMS — Integra asks below Medicare fee schedule floor.",
        "  🟡 PHCC BELOW CMS — Our current rate is already below Medicare floor.",
    ]
    for line in method:
        ws.cell(row=row, column=1, value=line).font = Font(size=10)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="CROSS-PAYER SUMMARY").font = Font(bold=True, size=12, color="4472C4")
    row += 1

    hdr = ["Payer", "Total Codes", "Matched", "No Match",
           "PHCC > Proposed", "Proposed < CMS", "PHCC < CMS",
           "Avg Δ (matched)", "Attention Codes"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    for payer in ["Commercial", "ASO", "Medicare", "Medicaid"]:
        df = payer_tables.get(payer, pd.DataFrame())
        if len(df) == 0:
            continue
        total = len(df)
        matched = len(df[df["Match"] != "NO_MATCH"])
        no_match = total - matched
        flags = df["Flag"].astype(str)
        pa = int(flags.str.contains("PHCC ABOVE PROPOSED", na=False).sum())
        pb = int(flags.str.contains("PROPOSED BELOW CMS", na=False).sum())
        pcb = int(flags.str.contains("PHCC BELOW CMS", na=False).sum())
        m = df[df["Match"] != "NO_MATCH"]
        avg_d = m["Δ Proposed vs PHCC"].dropna().mean()
        attention = pb + pcb  # codes needing negotiation review

        vals = [payer, total, matched, no_match, pa, pb, pcb,
                f"${avg_d:.2f}" if not math.isnan(avg_d) else "N/A",
                attention]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
        # Color the Proposed < CMS cell
        if pb > 0:
            ws.cell(row=row, column=6).fill = RED_FILL
        if pcb > 0:
            ws.cell(row=row, column=7).fill = YELLOW_FILL
        row += 1

    # Highest-impact codes across all payers
    row += 2
    ws.cell(row=row, column=1,
            value="TOP ATTENTION CODES (Proposed Below CMS — All Payers)").font = Font(bold=True, size=12)
    row += 1

    all_below = []
    for payer, df in payer_tables.items():
        if len(df) == 0:
            continue
        below = df[df["Flag"].astype(str).str.contains("PROPOSED BELOW CMS", na=False)].copy()
        below["Payer"] = payer
        all_below.append(below)

    if all_below:
        combined = pd.concat(all_below, ignore_index=True)
        combined = combined.sort_values("Proposed vs CMS NR", ascending=True)
        top_cols = ["Payer", "State", "HCPCS", "Mod", "Description",
                    "Proposed Rate", "CMS NR", "Proposed vs CMS NR", "PHCC Current"]
        top_cols = [c for c in top_cols if c in combined.columns]
        top = combined[top_cols].head(25)

        for c, h in enumerate(top_cols, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
        row += 1
        for _, tr in top.iterrows():
            for c, col in enumerate(top_cols, 1):
                val = tr[col]
                cell = ws.cell(row=row, column=c)
                if isinstance(val, float) and math.isnan(val):
                    cell.value = None
                elif isinstance(val, (np.floating,)):
                    cell.value = float(val) if not math.isnan(float(val)) else None
                else:
                    cell.value = val
                cell.border = THIN_BORDER
            row += 1
        # Currency fmt
        for c, h in enumerate(top_cols, 1):
            if h in ("Proposed Rate", "CMS NR", "Proposed vs CMS NR", "PHCC Current"):
                for r in range(row - len(top), row):
                    ws.cell(row=r, column=c).number_format = CURRENCY
    else:
        ws.cell(row=row, column=1, value="No codes below CMS benchmark.").font = Font(italic=True)

    _auto_width(ws)


# ───────────────────────────────────────────────────────────────────────
# 7.  MAIN
# ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("INTEGRA PHP FFS — Executive Rate Analysis")
    print(f"Run: {datetime.now():%Y-%m-%d %H:%M}")
    print("=" * 70)

    # Verify cleaned data
    for key in ("or_contracted", "or_participating", "wa_participating"):
        if not FILES[key].exists():
            print(f"ERROR: {FILES[key]} not found. Run clean_phcc_files.py first.")
            sys.exit(1)

    # Load PHCC cleaned
    print("\n[1] Loading PHCC cleaned schedules…")
    phcc_dfs, phcc_key_lks, phcc_code_lks = {}, {}, {}
    for key, label in [("or_contracted", "PHCC_OR_CONTRACTED"),
                        ("or_participating", "PHCC_OR_PARTICIPATING"),
                        ("wa_participating", "PHCC_WA_PARTICIPATING")]:
        df = load_cleaned_phcc(FILES[key], label)
        phcc_dfs[key] = df
        k, c = build_phcc(df)
        phcc_key_lks[key] = k
        phcc_code_lks[key] = c
        print(f"    {label}: {len(df)} rows")

    # Load Integra
    print("\n[2] Loading Integra proposed…")
    integra_dfs = {}
    for key, rate_col in [("integra_commercial", "Commercial"),
                           ("integra_aso", "ASO/Commercial"),
                           ("integra_medicare", "Medicare"),
                           ("integra_medicaid", "Medicaid")]:
        integra_dfs[key] = load_integra(FILES[key], rate_col)
        print(f"    {key}: {len(integra_dfs[key])} rows")

    # Load benchmarks
    print("\n[3] Loading benchmarks…")
    cms_or = load_cms(FILES["cms_or"], "OR (NR)", "OR (R)")
    cms_wa = load_cms(FILES["cms_wa"], "WA (NR)", "WA (R)")
    oha = load_oha(FILES["oha"])
    hcpcs_desc = load_hcpcs_desc(FILES["hcpcs"])
    print(f"    CMS OR: {len(cms_or)} keys, CMS WA: {len(cms_wa)} keys")
    print(f"    OHA: {len(oha)} keys, HCPCS descriptions: {len(hcpcs_desc)}")

    # Build per-payer tables
    print("\n[4] Building comparison tables…")
    payer_tables = {}
    for payer, configs in PAYER_CONFIGS.items():
        df = build_payer_table(
            payer, configs,
            integra_dfs, phcc_dfs, phcc_key_lks, phcc_code_lks,
            cms_or, cms_wa, oha, hcpcs_desc)
        payer_tables[payer] = df
        matched = len(df[df["Match"] != "NO_MATCH"]) if len(df) > 0 else 0
        flags = df["Flag"].astype(str) if len(df) > 0 else pd.Series(dtype=str)
        print(f"    {payer}: {len(df)} rows, {matched} matched, "
              f"{int(flags.str.contains('PROPOSED BELOW CMS', na=False).sum())} below CMS")

    # Write XLSX
    print("\n[5] Writing Excel workbook…")
    out_path = OUTPUT / "integra_rate_analysis.xlsx"
    wb = Workbook()

    # Tab 1: Executive Summary
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    _write_exec_summary(ws_exec, payer_tables)

    # Tabs 2-5: Per-payer detail
    for payer in ["Commercial", "ASO", "Medicare", "Medicaid"]:
        df = payer_tables[payer]
        ws = wb.create_sheet(payer)
        next_row = _write_payer_summary(ws, df, payer, 1)
        ws.cell(row=next_row, column=1,
                value="DETAIL TABLE").font = Font(bold=True, size=12)
        _write_table(ws, df, next_row + 1)
        _auto_width(ws)
        print(f"    Tab: {payer} — {len(df)} rows")

    wb.save(out_path)
    print(f"\n[XLSX] {out_path.name} saved with {len(wb.sheetnames)} tabs")
    print(f"    Tabs: {', '.join(wb.sheetnames)}")
    print(f"\n{'='*70}")
    print(f"DONE — {out_path}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
