#!/usr/bin/env python3
"""
centrix_rate_analysis.py  --  Centrix Care OR Fee Schedule Analysis
===================================================================

Compare Centrix Care OR proposed rates against:
  - PHCC OR Contracted current rates (Managed + Commercial)
  - CMS 2026 Q1 OR (Medicare benchmark)
  - OHA FFS Medicaid (Medicaid reference)

Code universe = UNION(Centrix HCPC codes, PHCC OR Contracted codes).
For each code, NU (purchase) and RR (rental) rates are looked up
independently.  Blanks where no data exists.

Output: output/centrix_rate_analysis.xlsx  (3 tabs: Summary, vs Managed, vs Commercial)

Prerequisite: python scripts/clean_phcc_files.py   (produces PHCC_OR_CONTRACTED_CLEAN.csv)
See: METHODOLOGY_CENTRIX.md
"""

import re, math, sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# =====================================================================
# 0.  PATHS
# =====================================================================
ROOT    = Path(__file__).resolve().parent.parent
DATA    = ROOT / "data"
CLEANED = DATA / "cleaned"
CMS_DIR = DATA / "cms"
CENTRIX = DATA / "CENTRIX"
OUTPUT  = ROOT / "output"

CONTRACT = DATA / "Contract"

FILES = {
    "centrix":          CENTRIX / "Centrix_Care_OR.csv",
    "or_contracted":    CLEANED / "PHCC_OR_CONTRACTED_CLEAN.csv",
    "or_contracted_raw": CONTRACT / "PHCC_OR_CONTRACTED.csv",
    "cms_or":           CMS_DIR / "CMS_2026_Q1_OR.csv",
    "oha":              CMS_DIR / "OHA_FFS_09_2025_RAW.csv",
    "hcpcs":            CMS_DIR / "2026_CMS_HCPCS.csv",
}

TOLERANCE_PCT = 1.0  # +/-1% treated as no change

# =====================================================================
# 1.  HELPERS
# =====================================================================
VALID_RE = re.compile(r'^[A-Z][0-9]{4}$')
MEDICARE_PCT_RE = re.compile(
    r'Medicare\s+Allowable\s+less\s+(\d+)\s*%', re.IGNORECASE)


def _norm(raw) -> str:
    s = str(raw).strip().upper()
    return s if VALID_RE.match(s) else ""


def _norm_mod(raw) -> str:
    s = str(raw).strip().upper() if pd.notna(raw) else ""
    return s if s else ""


def _sf(val) -> float:
    """Safe float, stripping $ and commas."""
    try:
        s = str(val).strip().replace("$", "").replace(",", "")
        return float(s) if s else np.nan
    except (ValueError, TypeError):
        return np.nan


def _classify_note(raw_val: str):
    """Parse raw rate text -> (numeric_rate, note_text)."""
    s = str(raw_val).strip()
    if not s:
        return np.nan, ""
    s_clean = s.replace("$", "").replace(",", "")
    try:
        return float(s_clean), ""
    except ValueError:
        return np.nan, s


def _resolve_pct_of_medicare(note_detail: str, cms_nr: float) -> float:
    m = MEDICARE_PCT_RE.search(str(note_detail))
    if m and not math.isnan(cms_nr):
        pct = int(m.group(1))
        return round(cms_nr * (1 - pct / 100), 2)
    return np.nan


# =====================================================================
# 2.  DATA LOADERS
# =====================================================================

def load_centrix(path: Path) -> tuple:
    """Load Centrix proposed rates.
    Returns:
        lk   : {hcpcs: {"NU": num, "RR": num, "NU_raw": str, ...}}
        meta : {hcpcs: {"cat": str, "type": str}}  (first seen values)
    """
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    meta: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPC", ""))
        if not hcpcs:
            continue
        mod = _norm_mod(r.get("MOD1", ""))
        raw = str(r.get("RATE", "")).strip()
        num, note = _classify_note(raw)

        slot = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {
                "NU": np.nan, "RR": np.nan,
                "NU_raw": "", "RR_raw": "",
                "NU_note": "", "RR_note": "",
            }
        lk[hcpcs][slot]           = num
        lk[hcpcs][f"{slot}_raw"]  = raw
        lk[hcpcs][f"{slot}_note"] = note

        if hcpcs not in meta:
            cat  = str(r.get("CAT", "")).strip()
            typ  = str(r.get("TYPE", "")).strip()
            meta[hcpcs] = {"cat": cat, "type": typ}

    return lk, meta


def load_phcc_or_contracted(path: Path, prefix: str) -> dict:
    """Load PHCC OR Contracted (cleaned).
    prefix = 'Managed' or 'Commercial'.
    Returns {hcpcs: {"NU": num, "RR": num, "NU_note_type": str, ...}}
    """
    df = pd.read_csv(path, dtype=str, keep_default_na=False)

    lk: dict = {}
    for _, r in df.iterrows():
        if str(r.get("hcpcs_is_valid", "")).strip() != "True":
            continue
        hcpcs = str(r.get("hcpcs_normalised", "")).strip()
        mod   = str(r.get("modifier_normalised", "")).strip()
        if not hcpcs:
            continue

        if mod == "RR":
            rate_base = f"{prefix} Rental Rate"
        else:
            rate_base = f"{prefix} Purchase Rate"

        num       = _sf(r.get(f"{rate_base}_numeric", ""))
        raw       = str(r.get(f"{rate_base}_raw", "")).strip()
        note_type = str(r.get(f"{rate_base}_note_type", "")).strip()
        note_det  = str(r.get(f"{rate_base}_note_detail", "")).strip()

        slot = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {
                "NU": np.nan, "RR": np.nan,
                "NU_raw": "", "RR_raw": "",
                "NU_note_type": "", "RR_note_type": "",
                "NU_note_detail": "", "RR_note_detail": "",
            }
        lk[hcpcs][slot]                  = num
        lk[hcpcs][f"{slot}_raw"]         = raw
        lk[hcpcs][f"{slot}_note_type"]   = note_type
        lk[hcpcs][f"{slot}_note_detail"] = note_det
    return lk


def load_cms(path: Path) -> dict:
    """Return {hcpcs: {"NU_nr": rate, "RR_nr": rate, "BLANK_nr": rate, ...}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("HCPCS", ""))
        mod   = _norm_mod(r.get("Mod", ""))
        if not hcpcs:
            continue
        nr = _sf(r.get("OR (NR)", ""))
        rv = _sf(r.get("OR (R)", ""))

        tag = mod if mod else "BLANK"
        if hcpcs not in lk:
            lk[hcpcs] = {}
        lk[hcpcs][f"{tag}_nr"] = nr
        lk[hcpcs][f"{tag}_r"]  = rv
    return lk


def _cms_rate(lk: dict, hcpcs: str, slot: str) -> float:
    """CMS Non-Rural for code+slot.  Cascade: exact -> blank fallback."""
    rec = lk.get(hcpcs, {})
    val = rec.get(f"{slot}_nr", np.nan)
    if isinstance(val, float) and not math.isnan(val) and val > 0:
        return val
    val_blank = rec.get("BLANK_nr", np.nan)
    if isinstance(val_blank, float) and not math.isnan(val_blank):
        return val_blank
    val_raw = rec.get(f"{slot}_nr", np.nan)
    if isinstance(val_raw, float) and not math.isnan(val_raw):
        return val_raw
    return np.nan


def load_oha(path: Path) -> dict:
    """Return {hcpcs: {"NU": rate, "RR": rate}}."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    lk: dict = {}
    for _, r in df.iterrows():
        hcpcs = _norm(r.get("Procedure Code", ""))
        mod   = _norm_mod(r.get("Mod1", ""))
        if not hcpcs:
            continue
        price = _sf(r.get("Price", ""))
        slot  = "RR" if mod == "RR" else "NU"

        if hcpcs not in lk:
            lk[hcpcs] = {"NU": np.nan, "RR": np.nan}
        lk[hcpcs][slot] = price
    return lk


def load_hcpcs_desc(path: Path) -> dict:
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    return {
        _norm(r.get("HCPC", "")): str(r.get("SHORT DESCRIPTION", "")).strip()
        for _, r in df.iterrows() if _norm(r.get("HCPC", ""))
    }


def load_raw_contract(path: Path) -> pd.DataFrame:
    """Load the raw (uncleaned) PHCC OR Contracted CSV for the contract view."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    df["_hcpcs"] = df["HCPCS"].str.strip().str.upper()
    df["_mod"]   = df["Mod"].str.strip().str.upper()
    return df


# =====================================================================
# 3.  BUILD ANALYSIS TABLE
# =====================================================================

def _delta(a: float, b: float):
    if math.isnan(a) or math.isnan(b):
        return np.nan, np.nan
    d = a - b
    p = (d / b * 100) if b != 0 else np.nan
    return d, p


def _flag(proposed: float, current: float, cms_nr: float,
          in_phcc: bool, is_phcc_only: bool,
          proposed_note: str, current_note_type: str) -> str:
    """Decision flag for one modifier slot."""
    if is_phcc_only:
        return "PHCC ONLY" if not math.isnan(current) else ""
    if math.isnan(proposed) and proposed_note:
        # Centrix has a text rate (MSRP - 30%)
        return "NON-NUMERIC PROPOSED"
    if math.isnan(proposed):
        return ""
    if not in_phcc:
        return "NEW CODE"
    if math.isnan(current):
        if current_note_type and current_note_type not in ("", "NUMERIC"):
            return "NON-NUMERIC CURRENT"
        return "NEW CODE"

    d = proposed - current
    pct = (d / current * 100) if current != 0 else 0.0
    if abs(pct) <= TOLERANCE_PCT:
        return "NO CHANGE"
    if proposed > current:
        return "RATE INCREASE"
    # proposed < current
    if not math.isnan(cms_nr):
        return "BELOW CURRENT" if proposed >= cms_nr else "BELOW CMS FLOOR"
    return "BELOW CURRENT"


def _systemic(current: float, cms_nr: float, flag: str) -> str:
    if (not math.isnan(current) and not math.isnan(cms_nr)
            and current < cms_nr):
        return f"{flag} | PHCC BELOW CMS" if flag else "PHCC BELOW CMS"
    return flag


def build_table(
    universe: set,
    centrix_lk: dict,
    centrix_meta: dict,
    phcc_lk: dict,
    cms_or: dict,
    oha_lk: dict,
    hcpcs_desc: dict,
    phcc_label: str,
) -> pd.DataFrame:
    """One row per HCPC code with NU + RR comparison columns."""

    rows = []
    for hcpcs in sorted(universe):
        in_centrix = hcpcs in centrix_lk
        in_phcc    = hcpcs in phcc_lk
        is_phcc_only = (not in_centrix) and in_phcc

        if in_centrix and in_phcc:
            source = "BOTH"
        elif in_centrix:
            source = "CENTRIX_ONLY"
        else:
            source = "PHCC_ONLY"

        # -- Centrix proposed --
        cx = centrix_lk.get(hcpcs, {})
        cx_nu      = cx.get("NU", np.nan)
        cx_rr      = cx.get("RR", np.nan)
        cx_nu_note = cx.get("NU_note", "")
        cx_rr_note = cx.get("RR_note", "")

        cm = centrix_meta.get(hcpcs, {})
        cx_cat  = cm.get("cat", "")
        cx_type = cm.get("type", "")

        # -- PHCC current --
        ph = phcc_lk.get(hcpcs, {})
        ph_nu = ph.get("NU", np.nan)
        ph_rr = ph.get("RR", np.nan)
        ph_nu_nt = ph.get("NU_note_type", "")
        ph_rr_nt = ph.get("RR_note_type", "")

        # Resolve PERCENT_OF_MEDICARE_ALLOWABLE
        for slot in ("NU", "RR"):
            if ph.get(f"{slot}_note_type") == "PERCENT_OF_MEDICARE_ALLOWABLE":
                cms_val = _cms_rate(cms_or, hcpcs, slot)
                resolved = _resolve_pct_of_medicare(
                    ph.get(f"{slot}_note_detail", ""), cms_val)
                if not math.isnan(resolved):
                    if slot == "NU":
                        ph_nu = resolved
                    else:
                        ph_rr = resolved

        # -- CMS OR --
        cms_nu = _cms_rate(cms_or, hcpcs, "NU")
        cms_rr = _cms_rate(cms_or, hcpcs, "RR")

        # -- OHA --
        oha = oha_lk.get(hcpcs, {})
        oha_nu = oha.get("NU", np.nan)
        oha_rr = oha.get("RR", np.nan)

        # -- Deltas --
        d_nu, p_nu = _delta(cx_nu, ph_nu)
        d_rr, p_rr = _delta(cx_rr, ph_rr)

        # -- Flags --
        f_nu = _flag(cx_nu, ph_nu, cms_nu,
                      in_phcc, is_phcc_only,
                      cx_nu_note, ph_nu_nt)
        f_rr = _flag(cx_rr, ph_rr, cms_rr,
                      in_phcc, is_phcc_only,
                      cx_rr_note, ph_rr_nt)
        f_nu = _systemic(ph_nu, cms_nu, f_nu)
        f_rr = _systemic(ph_rr, cms_rr, f_rr)

        rows.append({
            "HCPC":              hcpcs,
            "Description":       hcpcs_desc.get(hcpcs, ""),
            "Source":            source,
            "Centrix CAT":       cx_cat,
            "Centrix TYPE":      cx_type,
            "Centrix NU":        cx_nu,
            "Centrix RR":        cx_rr,
            "Centrix NU Note":   cx_nu_note,
            "Centrix RR Note":   cx_rr_note,
            f"PHCC {phcc_label} NU": ph_nu,
            f"PHCC {phcc_label} RR": ph_rr,
            "PHCC Note NU":      ph_nu_nt,
            "PHCC Note RR":      ph_rr_nt,
            "CMS OR NU":         cms_nu,
            "CMS OR RR":         cms_rr,
            "OHA NU":            oha_nu,
            "OHA RR":            oha_rr,
            "Delta NU":          d_nu,
            "Delta RR":          d_rr,
            "Delta% NU":         p_nu,
            "Delta% RR":         p_rr,
            "Flag NU":           f_nu,
            "Flag RR":           f_rr,
        })

    return pd.DataFrame(rows)


# =====================================================================
# 3b. BUILD CONTRACT VIEW TABLE
# =====================================================================

def build_contract_view(
    centrix_lk: dict,
    centrix_meta: dict,
    raw_contract: pd.DataFrame,
    phcc_managed: dict,
    phcc_commercial: dict,
    cms_or: dict,
    oha_lk: dict,
    hcpcs_desc: dict,
) -> pd.DataFrame:
    """Build a row-per-contract-line view showing raw contract data
    with Centrix proposed + numerical comparison + flags to the right."""

    rows = []
    for _, r in raw_contract.iterrows():
        hcpcs = r["_hcpcs"]
        if not VALID_RE.match(hcpcs):
            continue
        mod_raw = r["_mod"]

        # Raw contract columns
        desc_raw   = str(r.get("Description", "")).strip()
        unit_raw   = str(r.get("Billing Unit", "")).strip()
        mgd_rent   = str(r.get("Managed Rental Rate", "")).strip()
        mgd_purch  = str(r.get("Managed Purchase Rate", "")).strip()
        com_rent   = str(r.get("Commercial Rental Rate", "")).strip()
        com_purch  = str(r.get("Commercial Purchase Rate", "")).strip()
        comments   = str(r.get("Comments", "")).strip()

        # Determine slot for this row
        # For NU/RR split rows, mod_raw might be "NU/RR" - handle both
        mods_to_process = []
        if "/" in mod_raw:
            mods_to_process = [m.strip() for m in mod_raw.split("/")]
        else:
            mods_to_process = [mod_raw if mod_raw else "NU"]

        for mod in mods_to_process:
            slot = "RR" if mod == "RR" else "NU"

            # Centrix proposed for this code+slot
            cx = centrix_lk.get(hcpcs, {})
            cx_rate = cx.get(slot, np.nan)
            cx_note = cx.get(f"{slot}_note", "")
            in_centrix = hcpcs in centrix_lk

            # Current contract numeric (from cleaned data)
            mgd = phcc_managed.get(hcpcs, {})
            com = phcc_commercial.get(hcpcs, {})
            mgd_num = mgd.get(slot, np.nan)
            com_num = com.get(slot, np.nan)

            # CMS benchmark
            cms_nu = _cms_rate(cms_or, hcpcs, slot)

            # OHA Medicaid
            oha = oha_lk.get(hcpcs, {})
            oha_rate = oha.get(slot, np.nan)

            # Deltas: Centrix vs Managed, Centrix vs Commercial
            d_mgd, p_mgd = _delta(cx_rate, mgd_num)
            d_com, p_com = _delta(cx_rate, com_num)

            # Flag (vs Managed as primary)
            in_phcc = True  # we're iterating contract rows
            flag = _flag(cx_rate, mgd_num, cms_nu,
                         in_phcc, not in_centrix,
                         cx_note, mgd.get(f"{slot}_note_type", ""))
            flag = _systemic(mgd_num, cms_nu, flag)

            row = {
                # --- Raw contract columns (left) ---
                "HCPC":            hcpcs,
                "Mod":             mod,
                "Description":     desc_raw,
                "Billing Unit":    unit_raw,
                "Managed Rental":  mgd_rent if slot == "RR" else "",
                "Managed Purchase": mgd_purch if slot == "NU" else "",
                "Commercial Rental": com_rent if slot == "RR" else "",
                "Commercial Purchase": com_purch if slot == "NU" else "",
                "Comments":        comments,
                # --- Comparison columns (right) ---
                "Centrix Rate":      cx_rate,
                "Centrix Note":      cx_note,
                "Managed Numeric":   mgd_num,
                "Commercial Numeric": com_num,
                "CMS OR":            cms_nu,
                "OHA Medicaid":      oha_rate,
                "Delta vs Managed":  d_mgd,
                "Delta% vs Managed": p_mgd,
                "Delta vs Commercial": d_com,
                "Delta% vs Commercial": p_com,
                "Flag":              flag,
            }
            rows.append(row)

    return pd.DataFrame(rows)


# Currency/pct column indices for the Contract View tab (0-based)
# 0-HCPC, 1-Mod, 2-Desc, 3-BillingUnit, 4-MgdRent, 5-MgdPurch,
# 6-ComRent, 7-ComPurch, 8-Comments,
# 9-CentrixRate, 10-CentrixNote, 11-MgdNumeric, 12-ComNumeric,
# 13-CMS, 14-OHA, 15-DeltaMgd, 16-Delta%Mgd, 17-DeltaCom, 18-Delta%Com, 19-Flag
CV_CURRENCY_COLS = {9, 11, 12, 13, 14, 15, 17}
CV_PCT_COLS      = {16, 18}


def write_contract_tab(ws, df: pd.DataFrame):
    """Write the Contract View tab with specific formatting."""
    if df.empty:
        return
    cols = list(df.columns)

    # Header
    CONTRACT_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    COMPARE_HEADER_FILL  = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    contract_cols = {"HCPC", "Mod", "Description", "Billing Unit",
                     "Managed Rental", "Managed Purchase",
                     "Commercial Rental", "Commercial Purchase", "Comments"}

    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = (CONTRACT_HEADER_FILL if col_name in contract_cols
                     else COMPARE_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col_name in enumerate(cols, 1):
            val = row[col_name]
            if isinstance(val, float) and math.isnan(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER

            idx = ci - 1
            if idx in CV_CURRENCY_COLS and val is not None:
                cell.number_format = CURRENCY
            elif idx in CV_PCT_COLS and val is not None:
                cell.number_format = PCT_FMT

        # Flag colouring
        if "Flag" in cols:
            flag_idx = cols.index("Flag") + 1
            flag_val = str(row.get("Flag", "") or "")
            fill = _flag_fill(flag_val)
            if fill:
                ws.cell(row=ri, column=flag_idx).fill = fill

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


# =====================================================================
# 4.  XLSX FORMATTING
# =====================================================================
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
LTGRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="E2D0F0", end_color="E2D0F0", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(bold=True, size=11)
CURRENCY    = '"$"#,##0.00'
PCT_FMT     = '0.0"%"'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

FLAG_COLORS = {
    "BELOW CMS FLOOR":      RED_FILL,
    "PHCC BELOW CMS":       ORANGE_FILL,
    "BELOW CURRENT":        YELLOW_FILL,
    "RATE INCREASE":        BLUE_FILL,
    "NO CHANGE":            GREEN_FILL,
    "NEW CODE":             GRAY_FILL,
    "PHCC ONLY":            LTGRAY_FILL,
    "NON-NUMERIC PROPOSED": PURPLE_FILL,
    "NON-NUMERIC CURRENT":  GRAY_FILL,
}

# 0-based column indices for currency and pct formatting
# Columns: 0-HCPC, 1-Desc, 2-Source, 3-CAT, 4-TYPE,
#   5-Centrix NU, 6-Centrix RR, 7-CxNU Note, 8-CxRR Note,
#   9-PHCC NU, 10-PHCC RR, 11-PHCC Note NU, 12-PHCC Note RR,
#   13-CMS NU, 14-CMS RR, 15-OHA NU, 16-OHA RR,
#   17-Delta NU, 18-Delta RR, 19-Delta% NU, 20-Delta% RR,
#   21-Flag NU, 22-Flag RR
CURRENCY_COLS = {5, 6, 9, 10, 13, 14, 15, 16, 17, 18}
PCT_COLS      = {19, 20}


def _flag_fill(text: str):
    if not text:
        return None
    for kw, fill in FLAG_COLORS.items():
        if kw in text:
            return fill
    return None


def _auto_width(ws, max_w=30):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col[:80]:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(mx + 2, 8), max_w)


def write_tab(ws, df: pd.DataFrame):
    if df.empty:
        return
    cols = list(df.columns)

    # Header
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col_name in enumerate(cols, 1):
            val = row[col_name]
            if isinstance(val, float) and math.isnan(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER

            idx = ci - 1
            if idx in CURRENCY_COLS and val is not None:
                cell.number_format = CURRENCY
            elif idx in PCT_COLS and val is not None:
                cell.number_format = PCT_FMT

        # Flag colouring
        for flag_col_name in ("Flag NU", "Flag RR"):
            if flag_col_name in cols:
                flag_idx = cols.index(flag_col_name) + 1
                flag_val = str(row.get(flag_col_name, "") or "")
                fill = _flag_fill(flag_val)
                if fill:
                    ws.cell(row=ri, column=flag_idx).fill = fill

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def write_summary(ws, stats: dict):
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1,
                    value="Centrix Care OR -- Fee Schedule Analysis Summary")
    title.font = Font(bold=True, size=14)

    r = 3
    sections = [
        ("Code Universe", [
            ("Total unique HCPC codes", stats["universe"]),
            ("In Centrix", stats["centrix"]),
            ("In PHCC OR Contracted", stats["phcc"]),
            ("In Both", stats["both"]),
            ("Centrix Only", stats["centrix_only"]),
            ("PHCC Only", stats["phcc_only"]),
        ]),
        ("Centrix Rate Breakdown", [
            ("Numeric NU rates", stats["cx_nu_numeric"]),
            ("Numeric RR rates", stats["cx_rr_numeric"]),
            ("Non-numeric NU (MSRP etc.)", stats["cx_nu_text"]),
            ("Non-numeric RR (MSRP etc.)", stats["cx_rr_text"]),
        ]),
    ]

    for section_title, items in sections:
        ws.cell(row=r, column=1, value=section_title).font = BOLD_FONT
        r += 1
        for label, val in items:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        r += 1

    for tab_name, pstats in stats.get("tabs", {}).items():
        ws.cell(row=r, column=1, value=f"{tab_name} -- Flag Distribution").font = BOLD_FONT
        r += 1
        for label, val in pstats:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        r += 1

    _auto_width(ws, max_w=50)


def compute_tab_stats(df: pd.DataFrame) -> list:
    total = len(df)
    both       = len(df[df["Source"] == "BOTH"])
    cx_only    = len(df[df["Source"] == "CENTRIX_ONLY"])
    phcc_only  = len(df[df["Source"] == "PHCC_ONLY"])

    flag_nu = df["Flag NU"].fillna("").value_counts()
    flag_rr = df["Flag RR"].fillna("").value_counts()

    items = [
        ("Total codes", total),
        ("BOTH", both),
        ("CENTRIX_ONLY", cx_only),
        ("PHCC_ONLY", phcc_only),
        ("", ""),
        ("-- NU Flag Distribution --", ""),
    ]
    for flag, cnt in flag_nu.items():
        if flag:
            items.append((f"  {flag}", int(cnt)))
    items.append(("", ""))
    items.append(("-- RR Flag Distribution --", ""))
    for flag, cnt in flag_rr.items():
        if flag:
            items.append((f"  {flag}", int(cnt)))
    return items


# =====================================================================
# 5.  MAIN
# =====================================================================

def main():
    missing = [k for k, p in FILES.items() if not p.exists()]
    if missing:
        print("ERROR -- missing files:")
        for k in missing:
            print(f"  {k}: {FILES[k]}")
        print("\nEnsure clean_phcc_files.py has been run and data files exist.")
        sys.exit(1)

    OUTPUT.mkdir(exist_ok=True)

    print("Loading data ...")
    hcpcs_desc      = load_hcpcs_desc(FILES["hcpcs"])
    cms_or           = load_cms(FILES["cms_or"])
    oha_lk           = load_oha(FILES["oha"])
    centrix_lk, centrix_meta = load_centrix(FILES["centrix"])
    print(f"  Centrix: {len(centrix_lk)} unique codes")

    # Load PHCC for Managed and Commercial
    phcc_managed    = load_phcc_or_contracted(FILES["or_contracted"], "Managed")
    phcc_commercial = load_phcc_or_contracted(FILES["or_contracted"], "Commercial")
    print(f"  PHCC Managed: {len(phcc_managed)} codes")
    print(f"  PHCC Commercial: {len(phcc_commercial)} codes")
    print(f"  CMS OR: {len(cms_or)} codes")
    print(f"  OHA: {len(oha_lk)} codes")

    # Universe = Centrix + PHCC (union of both prefix loads, same codes)
    phcc_codes = set(phcc_managed.keys()) | set(phcc_commercial.keys())
    centrix_codes = set(centrix_lk.keys())
    universe = centrix_codes | phcc_codes
    overlap  = centrix_codes & phcc_codes
    print(f"\nCode universe: {len(universe)} "
          f"(Centrix={len(centrix_codes)}, PHCC={len(phcc_codes)}, "
          f"Both={len(overlap)}, "
          f"Centrix-only={len(centrix_codes - phcc_codes)}, "
          f"PHCC-only={len(phcc_codes - centrix_codes)})")

    # Centrix rate breakdown
    cx_nu_num = sum(1 for v in centrix_lk.values()
                    if not math.isnan(v.get("NU", np.nan)))
    cx_rr_num = sum(1 for v in centrix_lk.values()
                    if not math.isnan(v.get("RR", np.nan)))
    cx_nu_txt = sum(1 for v in centrix_lk.values()
                    if math.isnan(v.get("NU", np.nan)) and v.get("NU_note", ""))
    cx_rr_txt = sum(1 for v in centrix_lk.values()
                    if math.isnan(v.get("RR", np.nan)) and v.get("RR_note", ""))

    # Load raw contract for Contract View
    raw_contract = load_raw_contract(FILES["or_contracted_raw"])
    print(f"  Raw contract: {len(raw_contract)} rows")

    # Build tables
    tabs = {}
    tab_stats = {}
    for label, phcc_lk in [("Managed", phcc_managed),
                            ("Commercial", phcc_commercial)]:
        print(f"\nBuilding vs {label} ...")
        df = build_table(
            universe=universe,
            centrix_lk=centrix_lk,
            centrix_meta=centrix_meta,
            phcc_lk=phcc_lk,
            cms_or=cms_or,
            oha_lk=oha_lk,
            hcpcs_desc=hcpcs_desc,
            phcc_label=label,
        )
        tabs[f"vs {label}"] = df
        tab_stats[f"vs {label}"] = compute_tab_stats(df)
        print(f"  -> {len(df)} rows")

    # Build Contract View
    print("\nBuilding Contract View ...")
    cv_df = build_contract_view(
        centrix_lk=centrix_lk,
        centrix_meta=centrix_meta,
        raw_contract=raw_contract,
        phcc_managed=phcc_managed,
        phcc_commercial=phcc_commercial,
        cms_or=cms_or,
        oha_lk=oha_lk,
        hcpcs_desc=hcpcs_desc,
    )
    print(f"  -> {len(cv_df)} rows")

    summary = {
        "universe": len(universe),
        "centrix": len(centrix_codes),
        "phcc": len(phcc_codes),
        "both": len(overlap),
        "centrix_only": len(centrix_codes - phcc_codes),
        "phcc_only": len(phcc_codes - centrix_codes),
        "cx_nu_numeric": cx_nu_num,
        "cx_rr_numeric": cx_rr_num,
        "cx_nu_text": cx_nu_txt,
        "cx_rr_text": cx_rr_txt,
        "tabs": tab_stats,
    }

    # Write XLSX
    out_path = OUTPUT / "centrix_rate_analysis.xlsx"
    print(f"\nWriting {out_path} ...")
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary(ws_summary, summary)

    for tab_name, df in tabs.items():
        ws = wb.create_sheet(title=tab_name)
        write_tab(ws, df)

    # Contract View tab
    ws_cv = wb.create_sheet(title="Contract View")
    write_contract_tab(ws_cv, cv_df)

    wb.save(out_path)
    print(f"Done -> {out_path}")
    print(f"  Tabs: Summary, {', '.join(tabs.keys())}")


if __name__ == "__main__":
    main()
