"""Run all 3 scripts & verify contract tab numeric types."""
import subprocess, sys, os, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

os.chdir(os.path.join(os.path.dirname(__file__), ".."))

scripts = [
    "scripts/integra_rate_analysis_v2.py",
    "scripts/integra_rate_analysis_v2_formulas.py",
    "scripts/centrix_rate_analysis.py",
]

for s in scripts:
    print(f"\n{'='*60}\nRunning {s}\n{'='*60}")
    r = subprocess.run([sys.executable, s], capture_output=True, text=True,
                       encoding="utf-8", errors="replace")
    print(r.stdout[-1500:] if len(r.stdout) > 1500 else r.stdout)
    if r.returncode != 0:
        print(f"STDERR: {r.stderr[-1000:]}")
        print(f"FAILED (rc={r.returncode})")
        sys.exit(1)
    print(f"OK (rc={r.returncode})")

# Verify numeric types in contract tab cells
print(f"\n{'='*60}\nVerifying cell types in XLSX outputs\n{'='*60}")
from openpyxl import load_workbook

checks = [
    ("output/integra_rate_analysis_v2.xlsx", "CV OR Contracted", ["Managed Rental", "Managed Purchase", "Commercial Rental", "Commercial Purchase"]),
    ("output/integra_rate_analysis_v2.xlsx", "CV OR Participating", ["Rental Rate", "Purchase Rate"]),
    ("output/integra_rate_analysis_v2.xlsx", "CV WA Participating", ["Rental Rate", "Purchase Rate"]),
    ("output/integra_rate_analysis_v2_formulas.xlsx", "CV OR Contracted", ["Managed Rental", "Managed Purchase", "Commercial Rental", "Commercial Purchase"]),
    ("output/integra_rate_analysis_v2_formulas.xlsx", "CV OR Participating", ["Rental Rate", "Purchase Rate"]),
    ("output/centrix_rate_analysis.xlsx", "Contract View", ["Managed Rental", "Managed Purchase", "Commercial Rental", "Commercial Purchase"]),
]

all_ok = True
for xlsx, tab, cols_to_check in checks:
    wb = load_workbook(xlsx)
    ws = wb[tab]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    print(f"\n{xlsx} -> {tab}")
    for col_name in cols_to_check:
        ci = headers.get(col_name)
        if ci is None:
            print(f"  {col_name}: COLUMN NOT FOUND")
            continue
        nums, strs, empties = 0, 0, 0
        str_examples = []
        for r in range(2, min(ws.max_row + 1, 300)):
            v = ws.cell(r, ci).value
            if v is None or v == "":
                empties += 1
            elif isinstance(v, (int, float)):
                nums += 1
            else:
                strs += 1
                if len(str_examples) < 3:
                    str_examples.append(repr(v))
        status = "OK" if nums > 0 and strs <= nums * 0.3 else ("WARN" if strs > 0 else "OK")
        if strs > 0 and all("less" not in s.lower() and "retail" not in s.lower() and "allow" not in s.lower() for s in str_examples):
            status = "BAD - numeric strings?"
            all_ok = False
        print(f"  {col_name}: {nums} numeric, {strs} text, {empties} empty. Text examples: {str_examples}  [{status}]")
    wb.close()

print(f"\n{'='*60}")
print("ALL CHECKS PASSED" if all_ok else "SOME CHECKS FAILED")
