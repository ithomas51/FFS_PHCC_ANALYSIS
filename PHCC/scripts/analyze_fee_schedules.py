"""
PHCC / Integra Fee-Schedule Comparison Engine
==============================================
Reads cleaned PHCC CSVs (from clean_phcc_files.py), Integra proposed CSVs,
CMS DMEPOS fee schedules (Rural + Non-Rural), OHA Medicaid, and HCPCS
reference data.  Produces a 9-tab XLSX workbook with executive summary,
full comparisons, benchmark analysis, and review queue.

Run:
    python scripts/clean_phcc_files.py       # prerequisite
    python scripts/analyze_fee_schedules.py  # this script

Requires: pip install pandas openpyxl numpy
"""

from __future__ import annotations
import re, os, sys, math
from pathlib import Path
from datetime import datetime
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ───────────────────────────────────────────────────────────────────────
# 0.  PATHS
# ───────────────────────────────────────────────────────────────────────
PHCC_ROOT = Path(__file__).resolve().parent.parent          # PHCC/
CLEANED   = PHCC_ROOT / "data" / "cleaned"
INTEGRA   = PHCC_ROOT / "data" / "INTEGRA_PHP_FFS"
CMS_DIR   = PHCC_ROOT / "data" / "cms"
OUTPUT    = PHCC_ROOT / "output"
OUTPUT.mkdir(exist_ok=True)

FILES = {
    # Cleaned PHCC
    "or_contracted":  CLEANED / "PHCC_OR_CONTRACTED_CLEAN.csv",
    "or_participating": CLEANED / "PHCC_OR_PARTICIPATING_CLEAN.csv",
    "wa_participating": CLEANED / "PHCC_WA_PARTICIPATING_CLEAN.csv",
    # Integra proposed
    "integra_commercial": INTEGRA / "Integra_PHP_CARVEOUTS_COMMERCIAL.csv",
    "integra_aso":        INTEGRA / "Integra_PHP_CARVEOUTS_ASO.csv",
    "integra_medicare":   INTEGRA / "Integra_PHP_CARVEOUTS_MEDICARE.csv",
    "integra_medicaid":   INTEGRA / "INTEGRA_PHP_CARVEOUTS_MEDICAID.csv",
    # CMS benchmarks
    "cms_or": CMS_DIR / "CMS_2026_Q1_OR.csv",
    "cms_wa": CMS_DIR / "CMS_2026_Q1_WA.csv",
    # OHA Medicaid
    "oha":    CMS_DIR / "OHA_FFS_09_2025_RAW.csv",
    # HCPCS descriptions
    "hcpcs":  CMS_DIR / "2026_CMS_HCPCS.csv",
    # Audit files (for Audit Trail tab)
    "audit_hcpcs":  CLEANED / "PHCC_hcpcs_audit.csv",
    "audit_ranges": CLEANED / "PHCC_hcpcs_range_expansion_audit.csv",
}

# Payer configuration: (integra_file_key, payer_label, integra_rate_col,
#                        phcc_schedule_keys, state, benchmark_type)
PAYER_CONFIG = [
    # Commercial → OR_CONTRACTED (Commercial cols) + OR_PARTICIPATING
    ("integra_commercial", "Commercial", "Commercial",
     ["or_contracted", "or_participating"], "OR", "CMS"),
    # ASO → OR_CONTRACTED (Commercial cols) + OR_PARTICIPATING
    ("integra_aso", "ASO", "ASO/Commercial",
     ["or_contracted", "or_participating"], "OR", "CMS"),
    # Medicare → OR_CONTRACTED (Managed cols) + OR_PARTICIPATING
    ("integra_medicare", "Medicare", "Medicare",
     ["or_contracted", "or_participating"], "OR", "CMS"),
    # Medicaid → OR_CONTRACTED (Managed cols) + OR_PARTICIPATING
    ("integra_medicaid", "Medicaid", "Medicaid",
     ["or_contracted", "or_participating"], "OR", "OHA"),
    # WA — all payers vs WA_PARTICIPATING
    ("integra_commercial", "Commercial", "Commercial",
     ["wa_participating"], "WA", "CMS"),
    ("integra_aso", "ASO", "ASO/Commercial",
     ["wa_participating"], "WA", "CMS"),
    ("integra_medicare", "Medicare", "Medicare",
     ["wa_participating"], "WA", "CMS"),
    ("integra_medicaid", "Medicaid", "Medicaid",
     ["wa_participating"], "WA", None),  # No WA Medicaid benchmark
]

# ───────────────────────────────────────────────────────────────────────
# 1.  PURE FUNCTIONS
# ───────────────────────────────────────────────────────────────────────
VALID_HCPCS_RE = re.compile(r'^[A-Z][0-9]{4}$')

def normalize_hcpcs(raw: str) -> str:
    if pd.isna(raw) or str(raw).strip() == "":
        return ""
    return str(raw).strip().upper()

def validate_hcpcs(code: str) -> bool:
    return bool(VALID_HCPCS_RE.match(code))

def safe_float(val) -> float:
    if pd.isna(val):
        return np.nan
    s = str(val).strip().replace("$", "").replace(",", "")
    if s == "" or s.lower() == "nan":
        return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan

def norm_mod(raw) -> str:
    if pd.isna(raw) or str(raw).strip() == "":
        return ""
    return str(raw).strip().upper().rstrip("*")

def classify_pricing_note(val) -> tuple[float, str, str]:
    """Returns (numeric_value_or_NaN, note_type, note_detail)."""
    if pd.isna(val):
        return np.nan, "", ""
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return np.nan, "", ""
    cleaned = s.replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned), "NUMERIC", ""
    except ValueError:
        pass
    su = s.upper()
    if "NON-BILLABLE" in su or "NON BILLABLE" in su:
        return np.nan, "NON_BILLABLE", s
    m = re.match(r'RETAIL\s+LESS\s+(\d+)%', su)
    if m:
        return np.nan, "PERCENT_OF_RETAIL", f"Retail less {m.group(1)}%"
    if "QUOTE" in su:
        return np.nan, "QUOTE_REQUIRED", s
    if "MEDICARE ALLOWABLE" in su:
        return np.nan, "PERCENT_OF_MEDICARE_ALLOWABLE", s
    if "PREVAIL" in su:
        return np.nan, "PREVAILING_STATE_RATES", s
    if "COST INVOICE" in su:
        return np.nan, "COST_INVOICE", s
    if "PER 15 MIN" in su or "PER MIN" in su:
        return np.nan, "PER_TIME_UNIT", s
    return np.nan, "UNPARSED_TEXT", s


def _delta(proposed: float, current: float) -> tuple[str, float, float]:
    """Compare proposed vs current. Returns (status, amount, pct)."""
    if math.isnan(proposed) or math.isnan(current):
        return "NOT_COMPARABLE", np.nan, np.nan
    if proposed > current:
        return "HIGHER", proposed - current, ((proposed - current) / current * 100) if current != 0 else np.nan
    if proposed < current:
        return "LOWER", proposed - current, ((proposed - current) / current * 100) if current != 0 else np.nan
    return "EQUAL", 0.0, 0.0


# Cross-modifier resolution order for T3
CROSS_MOD_ORDER = {
    "NU": ["RR", ""],
    "RR": ["NU", ""],
    "AU": ["NU", ""],
    "KF": ["NU", ""],
    "":   ["NU", "RR"],
}

def _cross_mod_candidates(proposed_mod: str) -> list[str]:
    return CROSS_MOD_ORDER.get(proposed_mod, ["NU", "RR", ""])


# ───────────────────────────────────────────────────────────────────────
# 2.  LOADERS
# ───────────────────────────────────────────────────────────────────────

def load_cleaned_phcc(path: Path, schedule_label: str) -> pd.DataFrame:
    """Load a cleaned PHCC CSV. Returns DataFrame with standardized columns."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df["_schedule_label"] = schedule_label
    # Ensure key columns exist
    for col in ["hcpcs_normalised", "modifier_normalised"]:
        if col not in df.columns:
            raise ValueError(f"Missing column {col} in {path}")
    return df


def _pick_phcc_rate(row: pd.Series, schedule_label: str, payer: str, modifier: str) -> tuple[float, str, str, str]:
    """
    Given a cleaned PHCC row, pick the correct rate column based on
    schedule type, payer group, and modifier.
    Returns (numeric_rate, raw_rate, note_type, note_detail).
    """
    if schedule_label == "PHCC_OR_CONTRACTED":
        # Contracted has Managed vs Commercial split, and Rental vs Purchase
        if payer in ("Medicare", "Medicaid"):
            prefix = "Managed"
        else:
            prefix = "Commercial"

        if modifier == "RR":
            rate_col = f"{prefix} Rental Rate"
        else:  # NU, AU, KF, blank, other → purchase
            rate_col = f"{prefix} Purchase Rate"

    else:
        # OR_PARTICIPATING / WA_PARTICIPATING: single Rental/Purchase
        if modifier == "RR":
            rate_col = "Rental Rate"
        else:
            rate_col = "Purchase Rate"

    raw = row.get(f"{rate_col}_raw", "")
    num = safe_float(row.get(f"{rate_col}_numeric", ""))
    ntype = str(row.get(f"{rate_col}_note_type", ""))
    ndetail = str(row.get(f"{rate_col}_note_detail", ""))
    return num, raw, ntype, ndetail


def load_integra(path: Path, rate_col: str, payer_label: str) -> pd.DataFrame:
    """Load an Integra proposed CSV. Returns rows with normalized keys."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    # Drop unnamed filler columns
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df.columns = [c.strip() for c in df.columns]

    rows = []
    for idx, r in df.iterrows():
        hcpcs = normalize_hcpcs(r.get("HCPCS", ""))
        mod1 = norm_mod(r.get("Mod 1", ""))
        mod2 = norm_mod(r.get("Mod 2", ""))
        rate_raw = str(r.get(rate_col, "")).strip()
        rate_num = safe_float(rate_raw)
        note_num, note_type, note_detail = classify_pricing_note(rate_raw)
        if math.isnan(rate_num) and not math.isnan(note_num):
            rate_num = note_num

        rows.append({
            "hcpcs": hcpcs,
            "hcpcs_valid": validate_hcpcs(hcpcs),
            "mod1": mod1,
            "mod2": mod2,
            "proposed_rate_raw": rate_raw,
            "proposed_rate_numeric": rate_num,
            "proposed_note_type": note_type,
            "proposed_note_detail": note_detail,
            "payer": payer_label,
            "source_row": idx + 2,
        })
    return pd.DataFrame(rows)


def load_cms(path: Path, nr_col: str, r_col: str, state: str) -> dict[str, dict]:
    """
    Load CMS fee schedule. Returns dict keyed by 'HCPCS|MOD' with
    both Non-Rural and Rural rates.
    """
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    lk: dict[str, dict] = {}
    for _, r in df.iterrows():
        hcpcs = normalize_hcpcs(r.get("HCPCS", ""))
        mod = norm_mod(r.get("Mod", ""))
        if not hcpcs:
            continue
        key = f"{hcpcs}|{mod}"
        lk[key] = {
            "rate_nr": safe_float(r.get(nr_col, "")),
            "rate_r":  safe_float(r.get(r_col, "")),
            "hcpcs":   hcpcs,
            "mod":     mod,
            "state":   state,
            "desc":    str(r.get("Short Description", "")),
        }
    return lk


def load_oha(path: Path) -> dict[str, dict]:
    """Load OHA Medicaid fee schedule. Returns dict keyed by 'CODE|MOD'."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    lk: dict[str, dict] = {}
    for _, r in df.iterrows():
        code = normalize_hcpcs(r.get("Procedure Code", ""))
        mod = norm_mod(r.get("Mod1", ""))
        if not code:
            continue
        key = f"{code}|{mod}"
        rate = safe_float(r.get("Price", ""))
        lk[key] = {
            "rate": rate,
            "code": code,
            "mod":  mod,
            "desc": str(r.get("Description", "")),
        }
    return lk


def load_hcpcs_descriptions(path: Path) -> dict[str, str]:
    """Load HCPCS code → short description mapping."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    lk: dict[str, str] = {}
    for _, r in df.iterrows():
        code = normalize_hcpcs(r.get("HCPC", ""))
        desc = str(r.get("SHORT DESCRIPTION", "")).strip()
        if code:
            lk[code] = desc
    return lk


# ───────────────────────────────────────────────────────────────────────
# 3.  PHCC LOOKUP BUILDER
# ───────────────────────────────────────────────────────────────────────

def build_phcc_lookup(df: pd.DataFrame):
    """
    Build dict keyed by 'HCPCS|MOD' → list of matching rows.
    Also build a code-only index: 'HCPCS' → list of rows.
    Returns (key_lookup, code_lookup).
    """
    key_lk: dict[str, list[pd.Series]] = {}
    code_lk: dict[str, list[pd.Series]] = {}
    for _, row in df.iterrows():
        hcpcs = str(row.get("hcpcs_normalised", "")).strip()
        mod = str(row.get("modifier_normalised", "")).strip()
        key = f"{hcpcs}|{mod}"
        key_lk.setdefault(key, []).append(row)
        code_lk.setdefault(hcpcs, []).append(row)
    return key_lk, code_lk


# ───────────────────────────────────────────────────────────────────────
# 4.  MULTI-TIER MATCHING ENGINE  (T1 → T4 + NO_MATCH)
# ───────────────────────────────────────────────────────────────────────

def match_all_tiers(
    hcpcs: str, proposed_mod: str, payer: str,
    schedule_label: str,
    key_lk: dict, code_lk: dict,
) -> list[dict]:
    """
    For one proposed (hcpcs, mod) against one PHCC schedule, return
    a list of dicts — one per tier that matched:
      {tier, confidence, is_primary, is_reference, phcc_row,
       cross_mod_used, all_phcc_mods}
    """
    results: list[dict] = []
    all_phcc_rows = code_lk.get(hcpcs, [])
    all_phcc_mods = sorted({str(r.get("modifier_normalised", "")) for r in all_phcc_rows})
    all_phcc_mods_str = ",".join(all_phcc_mods)

    # T1: exact HCPCS + exact modifier
    exact_key = f"{hcpcs}|{proposed_mod}"
    t1_rows = key_lk.get(exact_key, [])
    for r in t1_rows:
        results.append({
            "tier": "T1", "confidence": "HIGH",
            "phcc_row": r, "cross_mod_used": "",
            "all_phcc_mods": all_phcc_mods_str,
        })

    # T2: proposed-mod → PHCC-blank  (only if proposed_mod is non-blank)
    if proposed_mod:
        blank_key = f"{hcpcs}|"
        t2_rows = key_lk.get(blank_key, [])
        for r in t2_rows:
            results.append({
                "tier": "T2", "confidence": "MEDIUM",
                "phcc_row": r, "cross_mod_used": "",
                "all_phcc_mods": all_phcc_mods_str,
            })

    # T3: cross-modifier
    tried_mods = {proposed_mod, ""}  # already tried in T1/T2
    for alt_mod in _cross_mod_candidates(proposed_mod):
        if alt_mod in tried_mods:
            continue
        tried_mods.add(alt_mod)
        alt_key = f"{hcpcs}|{alt_mod}"
        t3_rows = key_lk.get(alt_key, [])
        for r in t3_rows:
            results.append({
                "tier": "T3", "confidence": "MEDIUM",
                "phcc_row": r, "cross_mod_used": alt_mod,
                "all_phcc_mods": all_phcc_mods_str,
            })

    # T4: HCPCS-only (all remaining mods not already emitted)
    emitted_mods = {r["phcc_row"].get("modifier_normalised", "") for r in results}
    for r in all_phcc_rows:
        rmod = str(r.get("modifier_normalised", ""))
        if rmod not in emitted_mods:
            emitted_mods.add(rmod)
            results.append({
                "tier": "T4", "confidence": "LOW",
                "phcc_row": r, "cross_mod_used": rmod,
                "all_phcc_mods": all_phcc_mods_str,
            })

    # Determine primary: best tier found
    if results:
        best_tier = min(results, key=lambda x: x["tier"])["tier"]
        for r in results:
            r["is_primary"] = (r["tier"] == best_tier)
            r["is_reference"] = (r["tier"] != best_tier)
        # If multiple rows at best tier, keep first as primary, rest as reference
        primary_seen = False
        for r in results:
            if r["is_primary"]:
                if primary_seen:
                    r["is_primary"] = False
                    r["is_reference"] = True
                primary_seen = True

    # NO_MATCH
    if not results:
        results.append({
            "tier": "NO_MATCH", "confidence": "NONE",
            "phcc_row": None, "cross_mod_used": "",
            "all_phcc_mods": all_phcc_mods_str,
            "is_primary": True, "is_reference": False,
        })

    return results


# ───────────────────────────────────────────────────────────────────────
# 5.  BENCHMARK CASCADE  (B1 → B4)
# ───────────────────────────────────────────────────────────────────────

def _benchmark_cascade_cms(
    hcpcs: str, proposed_mod: str, cms_lk: dict
) -> dict:
    """
    Try B1→B4 for CMS. Returns dict with nr, r rates + metadata.
    """
    cascade = [
        ("B1", proposed_mod),
        ("B2", "NU"),
        ("B3", "RR"),
        ("B4", ""),
    ]
    for tier_label, try_mod in cascade:
        key = f"{hcpcs}|{try_mod}"
        rec = cms_lk.get(key)
        if rec:
            return {
                "cms_benchmark_nr": rec["rate_nr"],
                "cms_benchmark_r":  rec["rate_r"],
                "cms_benchmark_source": f"CMS {rec['state']}",
                "cms_benchmark_match_tier": tier_label,
                "cms_benchmark_mod_used": try_mod,
                "cms_benchmark_mod_mismatch": (try_mod != proposed_mod),
                "cms_desc": rec.get("desc", ""),
            }
    return {
        "cms_benchmark_nr": np.nan,
        "cms_benchmark_r":  np.nan,
        "cms_benchmark_source": "",
        "cms_benchmark_match_tier": "NOT_FOUND",
        "cms_benchmark_mod_used": "",
        "cms_benchmark_mod_mismatch": False,
        "cms_desc": "",
    }


def _benchmark_cascade_oha(
    hcpcs: str, proposed_mod: str, oha_lk: dict
) -> dict:
    """Try B1→B3 for OHA Medicaid."""
    cascade = [
        ("B1", proposed_mod),
        ("B2", "NU"),
        ("B3", ""),
    ]
    for tier_label, try_mod in cascade:
        key = f"{hcpcs}|{try_mod}"
        rec = oha_lk.get(key)
        if rec and not math.isnan(rec["rate"]):
            return {
                "oha_benchmark": rec["rate"],
                "oha_benchmark_source": "OHA",
                "oha_benchmark_match_tier": tier_label,
                "oha_benchmark_mod_used": try_mod,
                "oha_benchmark_mod_mismatch": (try_mod != proposed_mod),
            }
    return {
        "oha_benchmark": np.nan,
        "oha_benchmark_source": "",
        "oha_benchmark_match_tier": "NOT_FOUND",
        "oha_benchmark_mod_used": "",
        "oha_benchmark_mod_mismatch": False,
    }


def lookup_benchmarks(
    hcpcs: str, proposed_mod: str, state: str,
    benchmark_type: str | None,
    cms_or_lk: dict, cms_wa_lk: dict, oha_lk: dict,
) -> dict:
    """
    Master benchmark lookup. Returns a dict with all benchmark columns.
    """
    result: dict[str, Any] = {}

    # CMS benchmark (always for Medicare; also for Commercial/ASO as reference)
    cms_lk = cms_or_lk if state == "OR" else cms_wa_lk
    cms = _benchmark_cascade_cms(hcpcs, proposed_mod, cms_lk)
    result.update(cms)

    # OHA benchmark (only for Medicaid in OR)
    if benchmark_type == "OHA":
        oha = _benchmark_cascade_oha(hcpcs, proposed_mod, oha_lk)
        result.update(oha)
    elif benchmark_type is None and state == "WA":
        result["oha_benchmark"] = np.nan
        result["oha_benchmark_source"] = "WA_MEDICAID_NOT_PROVIDED"
        result["oha_benchmark_match_tier"] = "NOT_APPLICABLE"
        result["oha_benchmark_mod_used"] = ""
        result["oha_benchmark_mod_mismatch"] = False
    else:
        result["oha_benchmark"] = np.nan
        result["oha_benchmark_source"] = ""
        result["oha_benchmark_match_tier"] = "NOT_APPLICABLE"
        result["oha_benchmark_mod_used"] = ""
        result["oha_benchmark_mod_mismatch"] = False

    return result


def _evaluate_benchmark_status(proposed: float, benchmark: float) -> str:
    if math.isnan(proposed) or math.isnan(benchmark):
        return "NOT_COMPARABLE"
    if benchmark == 0:
        return "NO_RATE"
    if proposed > benchmark:
        return "ABOVE_BENCHMARK"
    if proposed < benchmark:
        return "BELOW_BENCHMARK"
    return "EQUAL_TO_BENCHMARK"


# ───────────────────────────────────────────────────────────────────────
# 6.  REVIEW TRIGGER ENGINE
# ───────────────────────────────────────────────────────────────────────

def check_review_triggers(row: dict) -> tuple[bool, str]:
    """Returns (review_required, review_reasons_concatenated)."""
    reasons = []

    if not row.get("hcpcs_valid", True):
        reasons.append("INVALID_HCPCS")
    if row.get("match_tier") == "NO_MATCH":
        reasons.append("NO_PHCC_MATCH")
    if row.get("match_tier") == "T3" and row.get("is_primary"):
        reasons.append("CROSS_MOD_PRIMARY")
    if row.get("match_tier") == "T4" and row.get("is_primary"):
        reasons.append("HCPCS_ONLY_PRIMARY")
    if row.get("current_note_type") and row["current_note_type"] not in ("", "NUMERIC"):
        reasons.append("NON_NUMERIC_CURRENT_RATE")
    if row.get("proposed_note_type") and row["proposed_note_type"] not in ("", "NUMERIC"):
        reasons.append("NON_NUMERIC_PROPOSED_RATE")
    if row.get("cms_benchmark_mod_mismatch"):
        reasons.append("CMS_MOD_MISMATCH")
    if row.get("oha_benchmark_mod_mismatch"):
        reasons.append("OHA_MOD_MISMATCH")

    # Benchmark status split: NR vs R disagree
    st_nr = row.get("benchmark_status_nr", "")
    st_r = row.get("benchmark_status_r", "")
    if st_nr and st_r and st_nr != st_r and "NOT" not in st_nr and "NOT" not in st_r and "NO_" not in st_r:
        reasons.append("RURAL_NR_STATUS_SPLIT")

    if row.get("comparison_status_current") == "MISSING_CURRENT":
        reasons.append("MISSING_CURRENT_RATE")

    # Below both benchmarks
    if row.get("benchmark_status_nr") == "BELOW_BENCHMARK":
        reasons.append("BELOW_CMS_NR_BENCHMARK")
    if row.get("benchmark_status_r") == "BELOW_BENCHMARK":
        reasons.append("BELOW_CMS_R_BENCHMARK")

    # WA Medicaid missing
    if row.get("oha_benchmark_source") == "WA_MEDICAID_NOT_PROVIDED":
        reasons.append("WA_MEDICAID_BENCHMARK_MISSING")

    return (len(reasons) > 0, "; ".join(reasons))


# ───────────────────────────────────────────────────────────────────────
# 7.  MAIN PIPELINE
# ───────────────────────────────────────────────────────────────────────

def run_analysis() -> pd.DataFrame:
    """Execute full analysis. Returns master DataFrame of all comparisons."""
    print("=" * 70)
    print("PHCC / Integra Fee Schedule Comparison Engine")
    print(f"Run date: {datetime.now():%Y-%m-%d %H:%M}")
    print("=" * 70)

    # ── Load cleaned PHCC schedules ──
    print("\n[1] Loading cleaned PHCC schedules…")
    phcc_dfs = {}
    phcc_key_lks = {}
    phcc_code_lks = {}
    for key, label in [
        ("or_contracted", "PHCC_OR_CONTRACTED"),
        ("or_participating", "PHCC_OR_PARTICIPATING"),
        ("wa_participating", "PHCC_WA_PARTICIPATING"),
    ]:
        df = load_cleaned_phcc(FILES[key], label)
        phcc_dfs[key] = df
        k_lk, c_lk = build_phcc_lookup(df)
        phcc_key_lks[key] = k_lk
        phcc_code_lks[key] = c_lk
        print(f"    {label}: {len(df)} rows, {len(c_lk)} unique HCPCS codes")

    # ── Load Integra proposed schedules ──
    print("\n[2] Loading Integra proposed schedules…")
    integra_dfs = {}
    for key, payer, rate_col in [
        ("integra_commercial", "Commercial", "Commercial"),
        ("integra_aso", "ASO", "ASO/Commercial"),
        ("integra_medicare", "Medicare", "Medicare"),
        ("integra_medicaid", "Medicaid", "Medicaid"),
    ]:
        df = load_integra(FILES[key], rate_col, payer)
        integra_dfs[key] = df
        print(f"    {payer}: {len(df)} rows")

    # ── Load CMS benchmarks (NR + R) ──
    print("\n[3] Loading CMS benchmarks (NR + Rural)…")
    cms_or_lk = load_cms(FILES["cms_or"], "OR (NR)", "OR (R)", "OR")
    cms_wa_lk = load_cms(FILES["cms_wa"], "WA (NR)", "WA (R)", "WA")
    print(f"    CMS OR: {len(cms_or_lk)} keys")
    print(f"    CMS WA: {len(cms_wa_lk)} keys")

    # ── Load OHA Medicaid ──
    print("\n[4] Loading OHA Medicaid…")
    oha_lk = load_oha(FILES["oha"])
    print(f"    OHA: {len(oha_lk)} keys")

    # ── Load HCPCS descriptions ──
    print("\n[5] Loading HCPCS descriptions…")
    hcpcs_desc = load_hcpcs_descriptions(FILES["hcpcs"])
    print(f"    Descriptions: {len(hcpcs_desc)} codes")

    # ── Process each payer config ──
    print("\n[6] Running comparisons…")
    all_rows: list[dict] = []

    for (integra_key, payer, rate_col,
         phcc_keys, state, benchmark_type) in PAYER_CONFIG:

        integra_df = integra_dfs[integra_key]
        print(f"\n    [{payer} / {state}] {len(integra_df)} proposed rows "
              f"vs {len(phcc_keys)} schedule(s)…")

        for _, prop_row in integra_df.iterrows():
            hcpcs = prop_row["hcpcs"]
            mod1 = prop_row["mod1"]
            proposed_num = prop_row["proposed_rate_numeric"]

            if not hcpcs:
                continue

            # Get benchmarks once per proposed row + state
            bench = lookup_benchmarks(
                hcpcs, mod1, state, benchmark_type,
                cms_or_lk, cms_wa_lk, oha_lk,
            )

            # Match against each PHCC schedule for this config
            for sched_key in phcc_keys:
                schedule_label = phcc_dfs[sched_key]["_schedule_label"].iloc[0]
                k_lk = phcc_key_lks[sched_key]
                c_lk = phcc_code_lks[sched_key]

                matches = match_all_tiers(
                    hcpcs, mod1, payer, schedule_label, k_lk, c_lk,
                )

                for match in matches:
                    phcc_row = match["phcc_row"]

                    # Extract PHCC current rate
                    if phcc_row is not None:
                        cur_num, cur_raw, cur_ntype, cur_ndetail = _pick_phcc_rate(
                            phcc_row, schedule_label, payer,
                            match["cross_mod_used"] if match["cross_mod_used"] else mod1,
                        )
                        phcc_mod = str(phcc_row.get("modifier_normalised", ""))
                        phcc_desc = str(phcc_row.get("orig_Description", ""))
                        phcc_billing_unit = str(phcc_row.get("orig_Billing Unit", ""))
                        phcc_hcpcs_orig = str(phcc_row.get("hcpcs_original", ""))
                        phcc_expanded = str(phcc_row.get("expanded_from_range", ""))
                    else:
                        cur_num, cur_raw, cur_ntype, cur_ndetail = np.nan, "", "", ""
                        phcc_mod = ""
                        phcc_desc = ""
                        phcc_billing_unit = ""
                        phcc_hcpcs_orig = ""
                        phcc_expanded = ""

                    # Comparison: proposed vs current
                    comp_status, comp_amt, comp_pct = _delta(proposed_num, cur_num)
                    if match["tier"] == "NO_MATCH":
                        comp_status = "MISSING_CURRENT"

                    # Benchmark evaluation (NR)
                    bench_status_nr = "NOT_APPLICABLE"
                    delta_nr = np.nan
                    pct_nr = np.nan
                    if not math.isnan(bench.get("cms_benchmark_nr", np.nan)):
                        bench_status_nr = _evaluate_benchmark_status(
                            proposed_num, bench["cms_benchmark_nr"])
                        if not math.isnan(proposed_num):
                            delta_nr = proposed_num - bench["cms_benchmark_nr"]
                            if bench["cms_benchmark_nr"] != 0:
                                pct_nr = delta_nr / bench["cms_benchmark_nr"] * 100

                    # Benchmark evaluation (R)
                    bench_status_r = "NOT_APPLICABLE"
                    delta_r = np.nan
                    pct_r = np.nan
                    cms_r = bench.get("cms_benchmark_r", np.nan)
                    if not math.isnan(cms_r):
                        if cms_r == 0:
                            bench_status_r = "NO_RURAL_RATE"
                        else:
                            bench_status_r = _evaluate_benchmark_status(proposed_num, cms_r)
                            if not math.isnan(proposed_num):
                                delta_r = proposed_num - cms_r
                                pct_r = delta_r / cms_r * 100

                    # OHA benchmark evaluation
                    oha_status = "NOT_APPLICABLE"
                    delta_oha = np.nan
                    pct_oha = np.nan
                    oha_rate = bench.get("oha_benchmark", np.nan)
                    if not math.isnan(oha_rate) and oha_rate != 0:
                        oha_status = _evaluate_benchmark_status(proposed_num, oha_rate)
                        if not math.isnan(proposed_num):
                            delta_oha = proposed_num - oha_rate
                            pct_oha = delta_oha / oha_rate * 100
                    elif bench.get("oha_benchmark_source") == "WA_MEDICAID_NOT_PROVIDED":
                        oha_status = "WA_MEDICAID_NOT_PROVIDED"
                    elif bench.get("oha_benchmark_match_tier") == "NOT_FOUND":
                        oha_status = "NOT_FOUND_IN_OHA"

                    # Determine needs_benchmark_check
                    needs_bench = (comp_status == "LOWER"
                                   and payer in ("Medicare", "Medicaid"))

                    # Review triggers
                    row_dict: dict[str, Any] = {
                        # Source identity
                        "state": state,
                        "payer_group": payer,
                        "current_schedule_type": schedule_label,
                        "integra_source_file": os.path.basename(str(FILES[integra_key])),

                        # Code / modifier
                        "hcpcs_normalised": hcpcs,
                        "hcpcs_valid": prop_row["hcpcs_valid"],
                        "modifier_proposed": mod1,
                        "modifier_2": prop_row["mod2"],
                        "modifier_current": phcc_mod,
                        "modifier_match_strategy": match["tier"],

                        # Descriptions
                        "description_hcpcs_ref": hcpcs_desc.get(hcpcs, ""),
                        "description_proposed": "",  # Integra CSVs don't have descriptions
                        "description_current": phcc_desc,
                        "billing_unit_current": phcc_billing_unit,

                        # Proposed rate
                        "proposed_rate_raw": prop_row["proposed_rate_raw"],
                        "proposed_rate_numeric": proposed_num,
                        "proposed_note_type": prop_row["proposed_note_type"],
                        "proposed_note_detail": prop_row["proposed_note_detail"],

                        # Current rate
                        "current_rate_raw": cur_raw,
                        "current_rate_numeric": cur_num,
                        "current_note_type": cur_ntype,
                        "current_note_detail": cur_ndetail,

                        # Comparison vs current
                        "comparison_status_current": comp_status,
                        "comparison_amount_current": comp_amt,
                        "comparison_pct_current": comp_pct,

                        # Match tier info
                        "match_tier": match["tier"],
                        "match_confidence": match["confidence"],
                        "is_primary_match": match.get("is_primary", False),
                        "is_reference_match": match.get("is_reference", False),
                        "cross_mod_used": match["cross_mod_used"],
                        "all_phcc_mods_available": match["all_phcc_mods"],

                        # CMS benchmark
                        "cms_benchmark_nr": bench.get("cms_benchmark_nr", np.nan),
                        "cms_benchmark_r": bench.get("cms_benchmark_r", np.nan),
                        "cms_benchmark_source": bench.get("cms_benchmark_source", ""),
                        "cms_benchmark_match_tier": bench.get("cms_benchmark_match_tier", ""),
                        "cms_benchmark_mod_used": bench.get("cms_benchmark_mod_used", ""),
                        "cms_benchmark_mod_mismatch": bench.get("cms_benchmark_mod_mismatch", False),

                        # CMS NR evaluation
                        "benchmark_status_nr": bench_status_nr,
                        "delta_vs_cms_nr": delta_nr,
                        "pct_delta_vs_cms_nr": pct_nr,

                        # CMS R evaluation
                        "benchmark_status_r": bench_status_r,
                        "delta_vs_cms_r": delta_r,
                        "pct_delta_vs_cms_r": pct_r,

                        # OHA benchmark
                        "oha_benchmark": bench.get("oha_benchmark", np.nan),
                        "oha_benchmark_source": bench.get("oha_benchmark_source", ""),
                        "oha_benchmark_match_tier": bench.get("oha_benchmark_match_tier", ""),
                        "oha_benchmark_mod_used": bench.get("oha_benchmark_mod_used", ""),
                        "oha_benchmark_mod_mismatch": bench.get("oha_benchmark_mod_mismatch", False),
                        "oha_benchmark_status": oha_status,
                        "delta_vs_oha": delta_oha,
                        "pct_delta_vs_oha": pct_oha,

                        # Flags
                        "needs_benchmark_check": needs_bench,

                        # PHCC traceability
                        "phcc_hcpcs_original": phcc_hcpcs_orig,
                        "phcc_expanded_from_range": phcc_expanded,
                        "integra_source_row": prop_row["source_row"],
                    }

                    # Review triggers
                    rev_req, rev_reasons = check_review_triggers(row_dict)
                    row_dict["review_required"] = rev_req
                    row_dict["review_reason"] = rev_reasons

                    all_rows.append(row_dict)

    master = pd.DataFrame(all_rows)
    print(f"\n[7] Total comparison rows: {len(master)}")
    if len(master) > 0:
        primary = master[master["is_primary_match"] == True]
        ref = master[master["is_reference_match"] == True]
        print(f"    Primary matches: {len(primary)}")
        print(f"    Reference matches: {len(ref)}")
        print(f"    Review required: {master['review_required'].sum()}")
    return master


# ───────────────────────────────────────────────────────────────────────
# 8.  XLSX OUTPUT WRITER  — 9 tabs
# ───────────────────────────────────────────────────────────────────────

# Formatting constants
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
CURRENCY_FMT = '"$"#,##0.00'
PCT_FMT = '0.0"%"'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws: Worksheet, ncols: int):
    """Apply header formatting to row 1."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws: Worksheet, max_width: int = 30):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:50]:  # sample first 50 rows
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _apply_currency_fmt(ws: Worksheet, col_idx: int, max_row: int):
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=col_idx).number_format = CURRENCY_FMT


def _apply_pct_fmt(ws: Worksheet, col_idx: int, max_row: int):
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=col_idx).number_format = PCT_FMT


def _write_df_to_sheet(ws: Worksheet, df: pd.DataFrame, freeze_cols: int = 0):
    """Write a DataFrame to a worksheet with headers and formatting."""
    # Headers
    for c, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=c, value=col_name)
    _style_header(ws, len(df.columns))

    # Data
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(val, float) and math.isnan(val):
                cell.value = None
            elif isinstance(val, (np.bool_, bool)):
                cell.value = bool(val)
            elif isinstance(val, (np.integer,)):
                cell.value = int(val)
            elif isinstance(val, (np.floating,)):
                cell.value = float(val) if not math.isnan(float(val)) else None
            else:
                cell.value = val

    # Freeze panes
    if freeze_cols > 0:
        ws.freeze_panes = ws.cell(row=2, column=freeze_cols + 1)
    else:
        ws.freeze_panes = ws.cell(row=2, column=1)

    # Auto-filter
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions

    _auto_width(ws)


def _find_col_idx(df: pd.DataFrame, col_name: str) -> int | None:
    """Find 1-based column index."""
    if col_name in df.columns:
        return list(df.columns).index(col_name) + 1
    return None


def _apply_conditional_fill(ws: Worksheet, col_idx: int, max_row: int,
                            value_map: dict[str, PatternFill]):
    """Apply fill based on cell value."""
    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        val = str(cell.value).strip() if cell.value else ""
        if val in value_map:
            cell.fill = value_map[val]


def write_xlsx(master: pd.DataFrame):
    """Write the 9-tab XLSX workbook."""
    out_path = OUTPUT / "fee_schedule_comparison.xlsx"
    csv_path = OUTPUT / "fee_schedule_comparison_master.csv"

    # Also write CSV
    master.to_csv(csv_path, index=False)
    print(f"\n[CSV] {csv_path.name}: {len(master)} rows")

    wb = Workbook()

    # ── Tab 1: Executive Summary ──
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    _write_executive_summary(ws_exec, master)

    # ── Tab 2: All Comparisons (primary matches only) ──
    primary = master[master["is_primary_match"] == True].copy()
    # Select key columns for readability
    all_comp_cols = [
        "state", "payer_group", "current_schedule_type",
        "hcpcs_normalised", "modifier_proposed", "modifier_current",
        "description_hcpcs_ref", "description_current",
        "billing_unit_current",
        "proposed_rate_numeric", "current_rate_numeric",
        "comparison_status_current", "comparison_amount_current",
        "comparison_pct_current",
        "match_tier", "match_confidence", "cross_mod_used",
        "all_phcc_mods_available",
        "cms_benchmark_nr", "cms_benchmark_r",
        "benchmark_status_nr", "benchmark_status_r",
        "cms_benchmark_match_tier", "cms_benchmark_mod_used",
        "oha_benchmark", "oha_benchmark_status",
        "needs_benchmark_check",
        "review_required", "review_reason",
        "proposed_rate_raw", "current_rate_raw",
        "proposed_note_type", "current_note_type",
    ]
    # Only include columns that exist
    all_comp_cols = [c for c in all_comp_cols if c in primary.columns]
    ws_all = wb.create_sheet("All Comparisons")
    _write_df_to_sheet(ws_all, primary[all_comp_cols], freeze_cols=4)
    # Format currency/pct columns
    max_r = len(primary) + 1
    for cname in ["proposed_rate_numeric", "current_rate_numeric",
                   "comparison_amount_current", "cms_benchmark_nr",
                   "cms_benchmark_r", "oha_benchmark"]:
        ci = _find_col_idx(primary[all_comp_cols], cname)
        if ci:
            _apply_currency_fmt(ws_all, ci, max_r)
    for cname in ["comparison_pct_current"]:
        ci = _find_col_idx(primary[all_comp_cols], cname)
        if ci:
            _apply_pct_fmt(ws_all, ci, max_r)
    # Conditional formatting on comparison_status_current
    ci = _find_col_idx(primary[all_comp_cols], "comparison_status_current")
    if ci:
        _apply_conditional_fill(ws_all, ci, max_r, {
            "HIGHER": GREEN_FILL, "LOWER": RED_FILL,
            "EQUAL": GREEN_FILL, "NOT_COMPARABLE": YELLOW_FILL,
            "MISSING_CURRENT": GRAY_FILL,
        })

    # ── Tab 3: Reference Matches ──
    ref = master[master["is_reference_match"] == True].copy()
    ref_cols = [
        "state", "payer_group", "current_schedule_type",
        "hcpcs_normalised", "modifier_proposed", "modifier_current",
        "match_tier", "match_confidence", "cross_mod_used",
        "proposed_rate_numeric", "current_rate_numeric",
        "comparison_status_current", "comparison_amount_current",
        "description_hcpcs_ref",
    ]
    ref_cols = [c for c in ref_cols if c in ref.columns]
    ws_ref = wb.create_sheet("Reference Matches")
    if len(ref) > 0:
        _write_df_to_sheet(ws_ref, ref[ref_cols], freeze_cols=4)
    else:
        ws_ref.cell(row=1, column=1, value="No reference matches found")

    # ── Tab 4: Lower Than Current ──
    lower = primary[primary["comparison_status_current"] == "LOWER"].copy()
    lower_cols = [
        "state", "payer_group", "current_schedule_type",
        "hcpcs_normalised", "modifier_proposed",
        "description_hcpcs_ref",
        "proposed_rate_numeric", "current_rate_numeric",
        "comparison_amount_current", "comparison_pct_current",
        "match_tier",
        "cms_benchmark_nr", "benchmark_status_nr",
        "oha_benchmark", "oha_benchmark_status",
        "review_required", "review_reason",
    ]
    lower_cols = [c for c in lower_cols if c in lower.columns]
    ws_lower = wb.create_sheet("Lower Than Current")
    if len(lower) > 0:
        _write_df_to_sheet(ws_lower, lower[lower_cols], freeze_cols=4)
        max_r = len(lower) + 1
        for cname in ["proposed_rate_numeric", "current_rate_numeric",
                       "comparison_amount_current", "cms_benchmark_nr", "oha_benchmark"]:
            ci = _find_col_idx(lower[lower_cols], cname)
            if ci:
                _apply_currency_fmt(ws_lower, ci, max_r)
    else:
        ws_lower.cell(row=1, column=1, value="No proposed rates below current")

    # ── Tab 5: Below Benchmark ──
    below_bench = primary[
        (primary["benchmark_status_nr"] == "BELOW_BENCHMARK") |
        (primary["oha_benchmark_status"] == "BELOW_BENCHMARK")
    ].copy()
    bb_cols = [
        "state", "payer_group", "current_schedule_type",
        "hcpcs_normalised", "modifier_proposed",
        "description_hcpcs_ref",
        "proposed_rate_numeric", "current_rate_numeric",
        "cms_benchmark_nr", "cms_benchmark_r",
        "benchmark_status_nr", "benchmark_status_r",
        "delta_vs_cms_nr", "pct_delta_vs_cms_nr",
        "oha_benchmark", "oha_benchmark_status",
        "delta_vs_oha", "pct_delta_vs_oha",
    ]
    bb_cols = [c for c in bb_cols if c in below_bench.columns]
    ws_bb = wb.create_sheet("Below Benchmark")
    if len(below_bench) > 0:
        _write_df_to_sheet(ws_bb, below_bench[bb_cols], freeze_cols=4)
    else:
        ws_bb.cell(row=1, column=1, value="No proposed rates below CMS/OHA benchmark")

    # ── Tab 6: Contract Lower Than Medicare ──
    ws_clm = wb.create_sheet("Contract Lower Than Medicare")
    _write_contract_lower_than_medicare(ws_clm, primary)

    # ── Tab 7: PHCC Current vs CMS NR/R ──
    ws_rc = wb.create_sheet("PHCC Current vs CMS")
    _write_phcc_vs_cms_ratecheck(ws_rc, primary)

    # ── Tab 8: Rural vs Non-Rural ──
    # Medicare primary matches with CMS data
    rural_df = primary[
        (primary["payer_group"] == "Medicare") &
        (primary["cms_benchmark_match_tier"] != "NOT_FOUND")
    ].copy()
    rural_cols = [
        "state", "hcpcs_normalised", "modifier_proposed",
        "description_hcpcs_ref",
        "proposed_rate_numeric", "current_rate_numeric",
        "cms_benchmark_nr", "cms_benchmark_r",
        "delta_vs_cms_nr", "delta_vs_cms_r",
        "pct_delta_vs_cms_nr", "pct_delta_vs_cms_r",
        "benchmark_status_nr", "benchmark_status_r",
        "cms_benchmark_match_tier", "cms_benchmark_mod_used",
    ]
    rural_cols = [c for c in rural_cols if c in rural_df.columns]
    ws_rural = wb.create_sheet("Rural vs Non-Rural")
    if len(rural_df) > 0:
        _write_df_to_sheet(ws_rural, rural_df[rural_cols], freeze_cols=3)
        max_r = len(rural_df) + 1
        # Conditional: green if above both, red if below both, yellow if split
        nr_ci = _find_col_idx(rural_df[rural_cols], "benchmark_status_nr")
        r_ci = _find_col_idx(rural_df[rural_cols], "benchmark_status_r")
        if nr_ci and r_ci:
            for row_num in range(2, max_r + 1):
                nr_val = str(ws_rural.cell(row=row_num, column=nr_ci).value or "")
                r_val = str(ws_rural.cell(row=row_num, column=r_ci).value or "")
                if "ABOVE" in nr_val and "ABOVE" in r_val:
                    fill = GREEN_FILL
                elif "BELOW" in nr_val and "BELOW" in r_val:
                    fill = RED_FILL
                elif ("ABOVE" in nr_val and "BELOW" in r_val) or \
                     ("BELOW" in nr_val and "ABOVE" in r_val):
                    fill = YELLOW_FILL
                else:
                    fill = GRAY_FILL
                ws_rural.cell(row=row_num, column=nr_ci).fill = fill
                ws_rural.cell(row=row_num, column=r_ci).fill = fill
    else:
        ws_rural.cell(row=1, column=1, value="No Medicare CMS benchmark data")

    # ── Tab 9: Review Queue ──
    review = master[master["review_required"] == True].copy()
    review_cols = [
        "state", "payer_group", "current_schedule_type",
        "hcpcs_normalised", "modifier_proposed", "modifier_current",
        "match_tier", "is_primary_match",
        "proposed_rate_numeric", "current_rate_numeric",
        "comparison_status_current",
        "review_reason",
        "description_hcpcs_ref",
    ]
    review_cols = [c for c in review_cols if c in review.columns]
    ws_review = wb.create_sheet("Review Queue")
    if len(review) > 0:
        _write_df_to_sheet(ws_review, review[review_cols], freeze_cols=4)
        max_r = len(review) + 1
        # Yellow fill on review_reason column
        ci = _find_col_idx(review[review_cols], "review_reason")
        if ci:
            for r in range(2, max_r + 1):
                ws_review.cell(row=r, column=ci).fill = YELLOW_FILL
    else:
        ws_review.cell(row=1, column=1, value="No rows flagged for review")

    # ── Tab 10: Audit Trail ──
    ws_audit = wb.create_sheet("Audit Trail")
    _write_audit_trail(ws_audit)

    # ── Tab 11: Data Sources ──
    ws_sources = wb.create_sheet("Data Sources")
    _write_data_sources(ws_sources, master)

    # Save
    wb.save(out_path)
    print(f"\n[XLSX] {out_path.name} saved with {len(wb.sheetnames)} tabs")
    print(f"    Tabs: {', '.join(wb.sheetnames)}")
    return out_path


def _write_phcc_vs_cms_ratecheck(ws: Worksheet, primary: pd.DataFrame):
    """
    Rate-check sheet: PHCC current contracted rate +/- CMS NR and Rural.
    Shows every primary match that has BOTH a current PHCC rate AND at least
    one CMS benchmark, with computed deltas for quick analysis.
    """
    # Work on rows that have a usable PHCC current rate and at least one CMS rate
    df = primary.copy()
    df["_cur"] = pd.to_numeric(df["current_rate_numeric"], errors="coerce")
    df["_nr"] = pd.to_numeric(df["cms_benchmark_nr"], errors="coerce")
    df["_r"] = pd.to_numeric(df["cms_benchmark_r"], errors="coerce")
    has_data = df["_cur"].notna() & (df["_nr"].notna() | df["_r"].notna())
    df = df[has_data].copy()

    # Compute PHCC current vs CMS deltas
    df["cur_vs_nr_amt"] = df["_cur"] - df["_nr"]
    df["cur_vs_nr_pct"] = np.where(
        df["_nr"].notna() & (df["_nr"] != 0),
        (df["_cur"] - df["_nr"]) / df["_nr"] * 100, np.nan)
    df["cur_vs_nr_status"] = np.where(
        df["_nr"].isna(), "NO_NR_RATE",
        np.where(df["_cur"] > df["_nr"], "ABOVE",
        np.where(df["_cur"] < df["_nr"], "BELOW", "EQUAL")))

    df["cur_vs_r_amt"] = df["_cur"] - df["_r"]
    df["cur_vs_r_pct"] = np.where(
        df["_r"].notna() & (df["_r"] != 0),
        (df["_cur"] - df["_r"]) / df["_r"] * 100, np.nan)
    df["cur_vs_r_status"] = np.where(
        df["_r"].isna() | (df["_r"] == 0), "NO_RURAL_RATE",
        np.where(df["_cur"] > df["_r"], "ABOVE",
        np.where(df["_cur"] < df["_r"], "BELOW", "EQUAL")))

    # ── Header ──
    ws.cell(row=1, column=1,
            value="PHCC Current Rate vs CMS Medicare (NR / Rural)").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = Font(italic=True, size=10, color="666666")

    row = 4
    ws.cell(row=row, column=1, value="PURPOSE").font = Font(bold=True, size=12, color="C00000")
    row += 1
    purpose_lines = [
        "Rate-check analysis: Are PHCC's current contracted rates above or below Medicare CMS?",
        "CMS Non-Rural (NR) = primary benchmark.  Rural (R) shown for reference.",
        "Δ Amount = PHCC Current − CMS rate.  Positive = PHCC higher, Negative = PHCC lower.",
        "Δ % = (PHCC Current − CMS) / CMS × 100.",
    ]
    for line in purpose_lines:
        ws.cell(row=row, column=1, value=line).font = Font(size=10)
        row += 1

    # ── Summary ──
    row += 1
    ws.cell(row=row, column=1, value="SUMMARY — vs CMS Non-Rural").font = Font(bold=True, size=12, color="4472C4")
    row += 1

    nr_valid = df[df["cur_vs_nr_status"] != "NO_NR_RATE"]
    nr_above = len(nr_valid[nr_valid["cur_vs_nr_status"] == "ABOVE"])
    nr_below = len(nr_valid[nr_valid["cur_vs_nr_status"] == "BELOW"])
    nr_equal = len(nr_valid[nr_valid["cur_vs_nr_status"] == "EQUAL"])
    nr_total = len(nr_valid)
    avg_nr_delta = nr_valid["cur_vs_nr_amt"].mean() if nr_total > 0 else 0
    below_avg_nr = nr_valid.loc[nr_valid["cur_vs_nr_status"] == "BELOW", "cur_vs_nr_amt"].mean()
    below_avg_nr = below_avg_nr if not math.isnan(below_avg_nr) else 0

    summary_items = [
        ("Codes with PHCC rate + CMS NR rate", nr_total),
        ("PHCC ABOVE CMS NR", f"{nr_above}  ({nr_above/nr_total*100:.1f}%)" if nr_total else "0"),
        ("PHCC BELOW CMS NR", f"{nr_below}  ({nr_below/nr_total*100:.1f}%)" if nr_total else "0"),
        ("PHCC EQUAL to CMS NR", nr_equal),
        ("Avg Δ all codes (Current − NR)", f"${avg_nr_delta:.2f}"),
        ("Avg Δ codes BELOW NR only", f"${below_avg_nr:.2f}"),
    ]
    for label, val in summary_items:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=row, column=2, value=val)
        if isinstance(val, str) and "$" in val and "-" in val:
            cell.fill = RED_FILL
        row += 1

    # ── By-State breakdown ──
    row += 1
    ws.cell(row=row, column=1, value="BY STATE — vs CMS NR").font = Font(bold=True, size=11, color="4472C4")
    row += 1
    for c, h in enumerate(["State", "Above NR", "Below NR", "Equal", "Avg Δ Below"], 1):
        ws.cell(row=row, column=c, value=h).font = BOLD_FONT
    row += 1
    for state, grp in nr_valid.groupby("state"):
        a = len(grp[grp["cur_vs_nr_status"] == "ABOVE"])
        b = len(grp[grp["cur_vs_nr_status"] == "BELOW"])
        e = len(grp[grp["cur_vs_nr_status"] == "EQUAL"])
        bavg = grp.loc[grp["cur_vs_nr_status"] == "BELOW", "cur_vs_nr_amt"].mean()
        ws.cell(row=row, column=1, value=state)
        ws.cell(row=row, column=2, value=a)
        ws.cell(row=row, column=3, value=b)
        ws.cell(row=row, column=4, value=e)
        ws.cell(row=row, column=5, value=round(bavg, 2) if not math.isnan(bavg) else 0)
        ws.cell(row=row, column=5).number_format = CURRENCY_FMT
        row += 1

    # ── Detail table ──
    row += 1
    ws.cell(row=row, column=1,
            value="DETAIL — PHCC Current Rate +/- CMS NR / Rural").font = Font(bold=True, size=12)
    row += 1

    if len(df) == 0:
        ws.cell(row=row, column=1, value="No rows with both PHCC current rate and CMS benchmark.")
        _auto_width(ws)
        return

    detail_cols = {
        "state": "State",
        "payer_group": "Payer",
        "current_schedule_type": "Schedule",
        "hcpcs_normalised": "HCPCS",
        "modifier_current": "Mod",
        "description_hcpcs_ref": "Description",
        "current_rate_numeric": "PHCC Current",
        "cms_benchmark_nr": "CMS NR",
        "cur_vs_nr_amt": "Δ vs NR",
        "cur_vs_nr_pct": "Δ% vs NR",
        "cur_vs_nr_status": "Status NR",
        "cms_benchmark_r": "CMS Rural",
        "cur_vs_r_amt": "Δ vs Rural",
        "cur_vs_r_pct": "Δ% vs Rural",
        "cur_vs_r_status": "Status Rural",
    }

    src_cols = [c for c in detail_cols if c in df.columns]
    headers = [detail_cols[c] for c in src_cols]

    # Sort: below NR first (worst delta), then by HCPCS
    df = df.sort_values(["cur_vs_nr_status", "cur_vs_nr_amt"],
                        ascending=[True, True])

    data_start = row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    for _, drow in df.iterrows():
        for c, col_name in enumerate(src_cols, 1):
            val = drow.get(col_name)
            cell = ws.cell(row=row, column=c)
            if isinstance(val, float) and math.isnan(val):
                cell.value = None
            elif isinstance(val, (np.floating,)):
                cell.value = float(val) if not math.isnan(float(val)) else None
            elif isinstance(val, (np.bool_, bool)):
                cell.value = bool(val)
            elif isinstance(val, (np.integer,)):
                cell.value = int(val)
            else:
                cell.value = val
            cell.border = THIN_BORDER
        row += 1

    max_data_row = row - 1

    # Currency / pct formatting
    currency_labels = {"PHCC Current", "CMS NR", "Δ vs NR", "CMS Rural", "Δ vs Rural"}
    pct_labels = {"Δ% vs NR", "Δ% vs Rural"}
    for c, h in enumerate(headers, 1):
        if h in currency_labels:
            _apply_currency_fmt(ws, c, max_data_row)
        elif h in pct_labels:
            _apply_pct_fmt(ws, c, max_data_row)

    # Conditional fills on Status NR / Status Rural
    nr_status_ci = headers.index("Status NR") + 1 if "Status NR" in headers else None
    r_status_ci = headers.index("Status Rural") + 1 if "Status Rural" in headers else None
    status_map = {
        "ABOVE": GREEN_FILL, "BELOW": RED_FILL,
        "EQUAL": GREEN_FILL, "NO_NR_RATE": GRAY_FILL,
        "NO_RURAL_RATE": GRAY_FILL,
    }
    if nr_status_ci:
        _apply_conditional_fill(ws, nr_status_ci, max_data_row, status_map)
    if r_status_ci:
        _apply_conditional_fill(ws, r_status_ci, max_data_row, status_map)

    # Red fill on negative Δ vs NR cells
    delta_nr_ci = headers.index("Δ vs NR") + 1 if "Δ vs NR" in headers else None
    if delta_nr_ci:
        for r in range(data_start + 1, max_data_row + 1):
            v = ws.cell(row=r, column=delta_nr_ci).value
            if v is not None and isinstance(v, (int, float)) and v < 0:
                ws.cell(row=r, column=delta_nr_ci).fill = RED_FILL

    ws.freeze_panes = ws.cell(row=data_start + 1, column=5)
    ws.auto_filter.ref = f"A{data_start}:{get_column_letter(len(headers))}{max_data_row}"
    _auto_width(ws)


def _write_contract_lower_than_medicare(ws: Worksheet, primary: pd.DataFrame):
    """
    Executive-friendly sheet: PHCC contracted codes where the proposed
    Integra rate falls below the CMS Medicare Non-Rural benchmark.
    Includes a methodology summary at the top for quick decision-making.
    """
    # Filter: Medicare payer, primary match, proposed < CMS NR benchmark
    filt = primary[
        (primary["payer_group"] == "Medicare") &
        (primary["benchmark_status_nr"] == "BELOW_BENCHMARK")
    ].copy()

    # ── Methodology Summary ──
    ws.cell(row=1, column=1, value="Contract Lower Than Medicare Benchmark").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = Font(italic=True, size=10, color="666666")

    row = 4
    ws.cell(row=row, column=1, value="HOW THIS IS CALCULATED").font = Font(bold=True, size=12, color="C00000")
    row += 1
    method_lines = [
        "1. Source: Integra PHP Medicare carveout proposed rates vs CMS 2026 Q1 DMEPOS Fee Schedule.",
        "2. CMS Benchmark: Non-Rural (NR) rate is used as the primary Medicare floor.",
        "3. Match: Each proposed HCPCS + modifier is matched to CMS using cascade B1→B4:",
        "     B1 = exact modifier match, B2 = NU (purchase), B3 = RR (rental), B4 = blank.",
        "4. Below Benchmark = Proposed rate is LESS than the CMS NR rate for that code.",
        "5. Δ Amount = Proposed − CMS NR (negative means proposed is below floor).",
        "6. Δ % = (Proposed − CMS NR) / CMS NR × 100.",
        "7. Rural rate shown for reference — rural is typically higher than NR.",
        "",
        "ACTION: Codes listed here are priced BELOW the Medicare fee schedule floor.",
        "These warrant negotiation review — the proposed rate undercuts public reimbursement.",
    ]
    for line in method_lines:
        ws.cell(row=row, column=1, value=line).font = Font(size=10)
        row += 1

    # ── Quick Summary Stats ──
    row += 1
    ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=12, color="4472C4")
    row += 1

    total_medicare_primary = len(primary[primary["payer_group"] == "Medicare"])
    below_count = len(filt)
    below_pct = (below_count / total_medicare_primary * 100) if total_medicare_primary > 0 else 0

    # Average shortfall
    filt_numeric = filt["delta_vs_cms_nr"].dropna()
    avg_shortfall = filt_numeric.mean() if len(filt_numeric) > 0 else 0
    total_shortfall = filt_numeric.sum() if len(filt_numeric) > 0 else 0

    summary_items = [
        ("Total Medicare primary comparisons", total_medicare_primary),
        ("Codes BELOW CMS NR Benchmark", below_count),
        ("% Below Benchmark", f"{below_pct:.1f}%"),
        ("Average shortfall per code (Proposed − CMS NR)", f"${avg_shortfall:.2f}"),
        ("Total shortfall across all codes", f"${total_shortfall:.2f}"),
    ]
    for label, val in summary_items:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=row, column=2, value=val)
        if isinstance(val, str) and "$" in val and "-" in val:
            cell.fill = RED_FILL
        row += 1

    # ── By State breakdown ──
    row += 1
    ws.cell(row=row, column=1, value="BY STATE").font = Font(bold=True, size=11, color="4472C4")
    row += 1
    for c, h in enumerate(["State", "Below CMS NR", "Avg Δ", "Total Δ"], 1):
        ws.cell(row=row, column=c, value=h).font = BOLD_FONT
    row += 1
    if len(filt) > 0:
        for state, grp in filt.groupby("state"):
            deltas = grp["delta_vs_cms_nr"].dropna()
            ws.cell(row=row, column=1, value=state)
            ws.cell(row=row, column=2, value=len(grp))
            ws.cell(row=row, column=3, value=round(deltas.mean(), 2) if len(deltas) > 0 else 0)
            ws.cell(row=row, column=3).number_format = CURRENCY_FMT
            ws.cell(row=row, column=4, value=round(deltas.sum(), 2) if len(deltas) > 0 else 0)
            ws.cell(row=row, column=4).number_format = CURRENCY_FMT
            row += 1

    # ── Data Table ──
    row += 1
    ws.cell(row=row, column=1, value="DETAIL — Codes Below Medicare Benchmark").font = Font(bold=True, size=12)
    row += 1

    if len(filt) == 0:
        ws.cell(row=row, column=1, value="No proposed rates below CMS Medicare NR benchmark.")
        _auto_width(ws)
        return

    # Slim columns for exec readability
    detail_cols = {
        "state": "State",
        "hcpcs_normalised": "HCPCS",
        "modifier_proposed": "Mod",
        "description_hcpcs_ref": "Description",
        "proposed_rate_numeric": "Proposed Rate",
        "cms_benchmark_nr": "CMS NR Rate",
        "delta_vs_cms_nr": "Δ Amount",
        "pct_delta_vs_cms_nr": "Δ %",
        "cms_benchmark_r": "CMS Rural Rate",
        "benchmark_status_r": "vs Rural",
        "current_rate_numeric": "PHCC Current",
        "cms_benchmark_match_tier": "CMS Match",
    }

    headers = list(detail_cols.values())
    src_cols = list(detail_cols.keys())
    src_cols = [c for c in src_cols if c in filt.columns]
    headers = [detail_cols[c] for c in src_cols]

    # Sort by shortfall (worst first)
    filt = filt.sort_values("delta_vs_cms_nr", ascending=True)

    # Write headers
    data_start_row = row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    # Write data
    for _, drow in filt.iterrows():
        for c, col_name in enumerate(src_cols, 1):
            val = drow.get(col_name)
            cell = ws.cell(row=row, column=c)
            if isinstance(val, float) and math.isnan(val):
                cell.value = None
            elif isinstance(val, (np.floating,)):
                cell.value = float(val) if not math.isnan(float(val)) else None
            elif isinstance(val, (np.bool_, bool)):
                cell.value = bool(val)
            elif isinstance(val, (np.integer,)):
                cell.value = int(val)
            else:
                cell.value = val
            cell.border = THIN_BORDER
        row += 1

    max_data_row = row - 1

    # Currency formatting on rate/delta columns
    currency_labels = {"Proposed Rate", "CMS NR Rate", "Δ Amount", "CMS Rural Rate", "PHCC Current"}
    pct_labels = {"Δ %"}
    for c, h in enumerate(headers, 1):
        if h in currency_labels:
            _apply_currency_fmt(ws, c, max_data_row)
        elif h in pct_labels:
            _apply_pct_fmt(ws, c, max_data_row)

    # Red fill on Δ Amount column (all negative)
    delta_ci = headers.index("Δ Amount") + 1 if "Δ Amount" in headers else None
    if delta_ci:
        for r in range(data_start_row + 1, max_data_row + 1):
            ws.cell(row=r, column=delta_ci).fill = RED_FILL

    # Conditional on vs Rural
    rural_ci = headers.index("vs Rural") + 1 if "vs Rural" in headers else None
    if rural_ci:
        _apply_conditional_fill(ws, rural_ci, max_data_row, {
            "BELOW_BENCHMARK": RED_FILL,
            "ABOVE_BENCHMARK": GREEN_FILL,
            "NO_RURAL_RATE": GRAY_FILL,
        })

    # Freeze at data table headers
    ws.freeze_panes = ws.cell(row=data_start_row + 1, column=3)

    # Auto-filter on data table
    ws.auto_filter.ref = f"A{data_start_row}:{get_column_letter(len(headers))}{max_data_row}"

    _auto_width(ws)


def _write_executive_summary(ws: Worksheet, master: pd.DataFrame):
    """Write the Executive Summary tab."""
    ws.cell(row=1, column=1, value="PHCC / Integra Fee Schedule Comparison — Executive Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}")
    ws.cell(row=2, column=1).font = Font(italic=True, size=10)

    primary = master[master["is_primary_match"] == True]

    # Overall stats
    row = 4
    stats = [
        ("Total comparison rows (all tiers)", len(master)),
        ("Primary match rows", len(primary)),
        ("Reference match rows", len(master[master["is_reference_match"] == True])),
        ("Review required rows", int(master["review_required"].sum())),
    ]
    for label, val in stats:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=val)
        row += 1

    # Breakdown by payer × state
    row += 1
    ws.cell(row=row, column=1, value="Primary Match Breakdown by Payer × State")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    headers = ["Payer", "State", "Schedule", "Total", "Higher", "Lower",
               "Equal", "Not Comparable", "Missing Current",
               "% Higher", "% Lower", "Avg Δ (Lower)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row += 1

    if len(primary) > 0:
        groups = primary.groupby(["payer_group", "state", "current_schedule_type"])
        for (payer, state, sched), grp in sorted(groups):
            total = len(grp)
            higher = len(grp[grp["comparison_status_current"] == "HIGHER"])
            lower = len(grp[grp["comparison_status_current"] == "LOWER"])
            equal = len(grp[grp["comparison_status_current"] == "EQUAL"])
            not_comp = len(grp[grp["comparison_status_current"] == "NOT_COMPARABLE"])
            missing = len(grp[grp["comparison_status_current"] == "MISSING_CURRENT"])
            pct_higher = higher / total * 100 if total > 0 else 0
            pct_lower = lower / total * 100 if total > 0 else 0
            lower_grp = grp[grp["comparison_status_current"] == "LOWER"]
            avg_delta = lower_grp["comparison_amount_current"].mean() if len(lower_grp) > 0 else 0

            vals = [payer, state, sched, total, higher, lower, equal,
                    not_comp, missing, f"{pct_higher:.1f}%",
                    f"{pct_lower:.1f}%", f"${avg_delta:.2f}" if not math.isnan(avg_delta) else "N/A"]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = THIN_BORDER
            # Color the % Lower cell
            cell = ws.cell(row=row, column=11)
            if pct_lower > 20:
                cell.fill = RED_FILL
            elif pct_lower > 10:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = GREEN_FILL
            row += 1

    # Match tier distribution
    row += 2
    ws.cell(row=row, column=1, value="Match Tier Distribution (Primary)").font = Font(bold=True, size=12)
    row += 1
    if len(primary) > 0:
        tier_counts = primary["match_tier"].value_counts().sort_index()
        for c, h in enumerate(["Tier", "Count", "% of Primary"], 1):
            ws.cell(row=row, column=c, value=h).font = BOLD_FONT
        row += 1
        for tier, cnt in tier_counts.items():
            ws.cell(row=row, column=1, value=tier)
            ws.cell(row=row, column=2, value=cnt)
            ws.cell(row=row, column=3, value=f"{cnt/len(primary)*100:.1f}%")
            row += 1

    _auto_width(ws)


def _write_audit_trail(ws: Worksheet):
    """Write audit data from cleaned pipeline."""
    ws.cell(row=1, column=1, value="PHCC Cleanup Audit Trail")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    row = 3
    # HCPCS audit
    audit_path = FILES["audit_hcpcs"]
    if audit_path.exists():
        ws.cell(row=row, column=1, value="HCPCS Validation Issues").font = BOLD_FONT
        row += 1
        audit_df = pd.read_csv(audit_path, dtype=str, keep_default_na=False)
        if len(audit_df) > 0:
            for c, col in enumerate(audit_df.columns, 1):
                ws.cell(row=row, column=c, value=col).font = BOLD_FONT
            row += 1
            for _, arow in audit_df.iterrows():
                for c, val in enumerate(arow, 1):
                    ws.cell(row=row, column=c, value=val)
                row += 1
        else:
            ws.cell(row=row, column=1, value="No HCPCS issues found")
            row += 1
    else:
        ws.cell(row=row, column=1, value="HCPCS audit file not found — run clean_phcc_files.py first")
        row += 1

    row += 1
    # Range expansion audit
    range_path = FILES["audit_ranges"]
    if range_path.exists():
        ws.cell(row=row, column=1, value="HCPCS Range Expansions").font = BOLD_FONT
        row += 1
        range_df = pd.read_csv(range_path, dtype=str, keep_default_na=False)
        if len(range_df) > 0:
            for c, col in enumerate(range_df.columns, 1):
                ws.cell(row=row, column=c, value=col).font = BOLD_FONT
            row += 1
            for _, rrow in range_df.iterrows():
                for c, val in enumerate(rrow, 1):
                    ws.cell(row=row, column=c, value=val)
                row += 1
        else:
            ws.cell(row=row, column=1, value="No ranges expanded")
            row += 1

    _auto_width(ws)


def _write_data_sources(ws: Worksheet, master: pd.DataFrame):
    """Write Data Sources tab with file inventory."""
    ws.cell(row=1, column=1, value="Data Source Inventory")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Analysis date: {datetime.now():%Y-%m-%d %H:%M}")

    row = 4
    headers = ["File", "Category", "Path", "Exists", "Size (KB)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h).font = BOLD_FONT
    row += 1

    categories = {
        "or_contracted": "PHCC Cleaned",
        "or_participating": "PHCC Cleaned",
        "wa_participating": "PHCC Cleaned",
        "integra_commercial": "Integra Proposed",
        "integra_aso": "Integra Proposed",
        "integra_medicare": "Integra Proposed",
        "integra_medicaid": "Integra Proposed",
        "cms_or": "CMS Benchmark",
        "cms_wa": "CMS Benchmark",
        "oha": "OHA Benchmark",
        "hcpcs": "HCPCS Reference",
    }

    for key, cat in categories.items():
        path = FILES[key]
        exists = path.exists()
        size = path.stat().st_size / 1024 if exists else 0
        ws.cell(row=row, column=1, value=path.name)
        ws.cell(row=row, column=2, value=cat)
        ws.cell(row=row, column=3, value=str(path.relative_to(PHCC_ROOT)))
        ws.cell(row=row, column=4, value="YES" if exists else "NO")
        ws.cell(row=row, column=5, value=f"{size:.1f}")
        if not exists:
            ws.cell(row=row, column=4).fill = RED_FILL
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Output Summary").font = BOLD_FONT
    row += 1
    ws.cell(row=row, column=1, value="Total comparison rows")
    ws.cell(row=row, column=2, value=len(master))
    row += 1
    if len(master) > 0:
        ws.cell(row=row, column=1, value="Primary matches")
        ws.cell(row=row, column=2, value=int(master["is_primary_match"].sum()))
        row += 1
        ws.cell(row=row, column=1, value="Reference matches")
        ws.cell(row=row, column=2, value=int(master["is_reference_match"].sum()))
        row += 1
        ws.cell(row=row, column=1, value="Unique HCPCS codes")
        ws.cell(row=row, column=2, value=master["hcpcs_normalised"].nunique())

    _auto_width(ws)


# ───────────────────────────────────────────────────────────────────────
# 9.  ENTRY POINT
# ───────────────────────────────────────────────────────────────────────

def main():
    # Verify cleaned data exists
    for key in ("or_contracted", "or_participating", "wa_participating"):
        if not FILES[key].exists():
            print(f"ERROR: Cleaned file not found: {FILES[key]}")
            print("Run  python scripts/clean_phcc_files.py  first.")
            sys.exit(1)

    master = run_analysis()
    if len(master) == 0:
        print("\nWARNING: No comparison rows produced!")
        sys.exit(1)

    xlsx_path = write_xlsx(master)
    print(f"\n{'=' * 70}")
    print(f"DONE — Output in: {OUTPUT}")
    print(f"  XLSX: {xlsx_path.name}")
    print(f"  CSV:  fee_schedule_comparison_master.csv")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
