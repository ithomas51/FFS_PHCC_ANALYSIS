#!/usr/bin/env python3
"""Validate centrix_cms_analysis.xlsx output -- expected vs actual sheets/columns."""

import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT   = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "output" / "centrix_cms_analysis.xlsx"

EXPECTED_SHEETS = {
    "Summary":        None,  # free-form, skip column check
    "Centrix Fees":   ["HCPC", "Description", "CAT", "TYPE",
                       "NU Rate", "RR Rate", "NU Note", "RR Note"],
    "CMS NR Fees":    ["HCPC", "Description", "NR NU", "NR RR",
                       "Rural NU", "Rural RR"],
    "OHA Fees":       ["HCPC", "Description", "OHA NU", "OHA RR"],
    "CMS Comparison": ["HCPC", "Description", "Centrix NU", "CMS NR NU",
                       "Delta $", "Delta %", "Flag NU",
                       "Centrix RR", "CMS NR RR",
                       "Delta $ RR", "Delta % RR", "Flag RR"],
    "OHA Comparison": ["HCPC", "Description", "Centrix NU", "OHA NU",
                       "Delta $", "Delta %", "Flag NU",
                       "Centrix RR", "OHA RR",
                       "Delta $ RR", "Delta % RR", "Flag RR"],
    "Delta Flags":    ["HCPC", "Description", "Benchmark",
                       "CMS Flag NU", "CMS Flag RR",
                       "OHA Flag NU", "OHA Flag RR"],
}

ok = True

def fail(msg):
    global ok
    ok = False
    print(f"  FAIL: {msg}")

def info(msg):
    print(f"  OK:   {msg}")


print("=" * 60)
print(f"Validating: {OUTPUT.name}")
print("=" * 60)

if not OUTPUT.exists():
    print(f"\nERROR: File not found -> {OUTPUT}")
    print("Run centrix_cms_analysis.py first.")
    sys.exit(1)

wb = load_workbook(OUTPUT, read_only=True, data_only=True)

# 1. Sheet names
actual_sheets = wb.sheetnames
print(f"\nSheets: {actual_sheets}")
for name in EXPECTED_SHEETS:
    if name in actual_sheets:
        info(f"Sheet '{name}' present")
    else:
        fail(f"Sheet '{name}' MISSING")

extra = set(actual_sheets) - set(EXPECTED_SHEETS)
if extra:
    fail(f"Unexpected sheets: {extra}")

# 2. Column headers per sheet
for name, expected_cols in EXPECTED_SHEETS.items():
    if expected_cols is None or name not in actual_sheets:
        continue
    ws = wb[name]
    actual_cols = [str(c.value).strip() if c.value else "" for c in ws[1]]
    print(f"\n-- {name} --")
    print(f"   Expected cols: {expected_cols}")
    print(f"   Actual cols:   {actual_cols}")
    for ec in expected_cols:
        if ec in actual_cols:
            info(f"Col '{ec}'")
        else:
            fail(f"Col '{ec}' MISSING")

# 3. Row counts (all data sheets should have same row count)
print("\n-- Row Counts --")
row_counts = {}
for name in EXPECTED_SHEETS:
    if name == "Summary" or name not in actual_sheets:
        continue
    ws = wb[name]
    rows = ws.max_row - 1  # subtract header
    row_counts[name] = rows
    print(f"  {name}: {rows} data rows")

unique = set(row_counts.values())
if len(unique) == 1:
    info(f"All sheets have {unique.pop()} rows (consistent)")
else:
    fail(f"Inconsistent row counts: {row_counts}")

# 4. Spot-check: first data row of CMS Comparison has HCPC
if "CMS Comparison" in actual_sheets:
    ws = wb["CMS Comparison"]
    hcpc = ws.cell(row=2, column=1).value
    if hcpc and len(str(hcpc)) == 5:
        info(f"CMS Comparison row 1 HCPC = {hcpc}")
    else:
        fail(f"CMS Comparison row 1 HCPC unexpected: {hcpc}")

# 5. Flag columns should only have known values
VALID_FLAGS = {"BELOW", "ABOVE", "AT BENCHMARK", "NO BENCHMARK", "NON-NUMERIC", "", None}
for sheet_name in ("CMS Comparison", "OHA Comparison", "Delta Flags"):
    if sheet_name not in actual_sheets:
        continue
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    flag_cols = [i for i, h in enumerate(headers) if "Flag" in h]
    for ci in flag_cols:
        vals = set()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=ci + 1).value
            vals.add(str(v).strip() if v else "")
        bad = vals - VALID_FLAGS - {"None"}
        if bad:
            fail(f"{sheet_name} col '{headers[ci]}' has unexpected flags: {bad}")
        else:
            info(f"{sheet_name} col '{headers[ci]}' flags valid: {vals - {'', 'None'}}")

wb.close()

print("\n" + "=" * 60)
if ok:
    print("ALL CHECKS PASSED")
else:
    print("SOME CHECKS FAILED -- review above")
print("=" * 60)
