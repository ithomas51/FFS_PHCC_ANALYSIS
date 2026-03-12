"""
PHCC vs Integra PHP Fee Schedule Comparison Analysis  (v2)
============================================================
Implements the full PROMPT.md specification:
  - HCPCS normalization & validation
  - HCPCS range expansion
  - Pricing note classification
  - Multi-state, multi-payer matching with fallback
  - Modifier-aware matching (NU=purchase, RR=rental)
  - Benchmark comparison (CMS Medicare, OHA Medicaid)
  - Audit / review queue outputs
  - Executive summary with Excel formatting

Run:  python analyze_fee_schedules.py
Requires: pip install pandas openpyxl
"""

import pandas as pd
import numpy as np
import re
import logging
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════════════════
# PATHS
# ════════════════════════════════════════════════════════════════════════
BASE = Path(__file__).resolve().parent / "data"
CONTRACT = BASE / "Contract"
CMS = BASE / "cms"
INTEGRA = BASE / "INTEGRA_PHP_FFS"
OUT_DIR = Path(__file__).resolve().parent / "output"
OUT_DIR.mkdir(exist_ok=True)

FILES = {
    "integra_commercial": INTEGRA / "Integra_PHP_CARVEOUTS_COMMERCIAL.csv",
    "integra_medicare":   INTEGRA / "Integra_PHP_CARVEOUTS_MEDICARE.csv",
    "integra_medicaid":   INTEGRA / "INTEGRA_PHP_CARVEOUTS_MEDICAID.csv",
    "integra_aso":        INTEGRA / "Integra_PHP_CARVEOUTS_ASO.csv",
    "phcc_or":            CONTRACT / "PHCC_OR_CONTRACTED.csv",
    "phcc_wa":            CONTRACT / "PHCC_WA_PARTICIPATING.csv",
    "cms_or":             CMS / "CMS_2026_Q1_OR.csv",
    "cms_wa":             CMS / "CMS_2026_Q1_WA.csv",
    "oha":                CMS / "OHA_FFS_09_2025_RAW.csv",
    "hcpcs_ref":          CMS / "2026_CMS_HCPCS.csv",
}

VALID_HCPCS = re.compile(r"^[A-Z][0-9]{4}$")

# ════════════════════════════════════════════════════════════════════════
# PURE FUNCTIONS: normalization & classification
# ════════════════════════════════════════════════════════════════════════

def normalize_hcpcs(raw: str) -> str:
    """Upper-case, strip, remove embedded line breaks, fix OCR."""
    if pd.isna(raw):
        return ""
    s = str(raw).strip().upper()
    s = re.sub(r"[\r\n]+", "", s)
    # Fix common OCR: letter-O → digit-0 in K-codes (KO→K0)
    s = re.sub(r"^KO", "K0", s)
    return s


def validate_hcpcs(code: str) -> tuple:
    """Return (is_valid: bool, issue: str)."""
    if not code:
        return False, "EMPTY"
    if VALID_HCPCS.match(code):
        return True, ""
    if re.match(r"^[A-Z]\d{4}\s*-\s*[A-Z]?\d{4}$", code):
        return False, "RANGE"
    if "?" in code:
        return False, "OCR_ARTIFACT"
    if re.search(r"[^A-Z0-9]", code):
        return False, "ILLEGAL_CHARS"
    if len(code) != 5:
        return False, f"BAD_LENGTH_{len(code)}"
    return False, "UNKNOWN_FORMAT"


def parse_hcpcs_range(raw: str):
    """
    Try to parse 'A6530 - A6541' or 'A6530-A6541' into a list of codes.
    Returns (list_of_codes, range_start, range_end) or (None, raw, '') if unparsable.
    """
    s = normalize_hcpcs(raw)
    m = re.match(r"^([A-Z])(\d{4})\s*-\s*([A-Z]?)(\d{4})$", s)
    if not m:
        return None, s, ""
    prefix1, num_start, prefix2, num_end = m.groups()
    prefix2 = prefix2 if prefix2 else prefix1
    if prefix1 != prefix2:
        return None, s, ""
    start, end = int(num_start), int(num_end)
    if start > end or (end - start) > 500:
        return None, s, ""
    codes = [f"{prefix1}{i:04d}" for i in range(start, end + 1)]
    return codes, f"{prefix1}{num_start}", f"{prefix2}{num_end}"


def safe_float(val) -> float:
    """Parse a dollar amount; return NaN if non-numeric."""
    if pd.isna(val):
        return np.nan
    s = str(val).strip().replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return np.nan


def classify_pricing_note(raw: str) -> tuple:
    """
    Classify non-numeric rate text.
    Returns (note_type, note_detail).
    """
    if pd.isna(raw):
        return "", ""
    s = str(raw).strip()
    if not s:
        return "", ""
    num = safe_float(s)
    if not np.isnan(num):
        return "", ""
    sl = s.lower()
    if "non-billable" in sl or "non billable" in sl:
        return "NON_BILLABLE", s
    if "quote" in sl:
        return "QUOTE_REQUIRED", s
    if re.search(r"retail\s+less\s+(\d+)%", sl):
        return "PERCENT_OF_RETAIL", s
    if re.search(r"medicare\s+allowable\s+less\s+(\d+)%", sl):
        return "PERCENT_OF_MEDICARE_ALLOWABLE", s
    if "prevail" in sl:
        return "PREVAILING_STATE_RATES", s
    if "per 15 min" in sl or "per 30 min" in sl or "per hour" in sl:
        return "PER_TIME_UNIT", s
    if "cost invoice" in sl:
        return "COST_INVOICE", s
    return "UNPARSED_TEXT", s


def norm_mod(val) -> str:
    """Uppercase, stripped modifier."""
    if pd.isna(val):
        return ""
    return str(val).strip().upper()


# ════════════════════════════════════════════════════════════════════════
# LOADERS
# ════════════════════════════════════════════════════════════════════════

def _clean_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df.columns = [c.strip() for c in df.columns]
    return df


def load_integra(path, rate_col, payer_group, source_file):
    """Load one Integra CSV → list of normalized dicts."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df = _clean_cols(df)
    log.info(f"  Loaded {len(df)} rows from {source_file}")
    rows = []
    for idx, r in df.iterrows():
        hcpcs_raw = str(r.get("HCPCS", "")).strip()
        hcpcs_norm = normalize_hcpcs(hcpcs_raw)
        is_valid, issue = validate_hcpcs(hcpcs_norm)

        rate_raw = str(r.get(rate_col, "")).strip()
        rate_num = safe_float(rate_raw)
        note_type, note_detail = classify_pricing_note(rate_raw)

        mod1 = norm_mod(r.get("Mod 1", ""))
        mod2 = norm_mod(r.get("Mod 2", ""))

        rows.append({
            "source_file": source_file,
            "source_row": idx + 2,
            "payer_group": payer_group,
            "hcpcs_original": hcpcs_raw,
            "hcpcs_normalized": hcpcs_norm,
            "hcpcs_is_valid": is_valid,
            "hcpcs_validation_issue": issue,
            "expanded_from_range": "",
            "range_start": "",
            "range_end": "",
            "modifier_1": mod1,
            "modifier_2": mod2,
            "proposed_rate_raw": rate_raw,
            "proposed_rate_numeric": rate_num,
            "proposed_rate_note_type": note_type,
            "proposed_rate_note_detail": note_detail,
        })
    return rows


def load_phcc_or(path):
    """Load PHCC Oregon contracted fee schedule with range expansion."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df = _clean_cols(df)
    log.info(f"  Loaded {len(df)} rows from PHCC_OR_CONTRACTED.csv")
    records = []
    for idx, r in df.iterrows():
        hcpcs_raw = str(r.get("HCPCS", "")).strip()
        hcpcs_norm = normalize_hcpcs(hcpcs_raw)
        mod_raw = str(r.get("Mod", "")).strip()

        mods = [m.strip().upper() for m in mod_raw.split("/") if m.strip()]
        if not mods:
            mods = [""]

        desc = str(r.get("Description", "")).strip()
        billing_unit = str(r.get("Billing Unit", "")).strip()

        managed_rental = str(r.get("Managed Rental Rate", "")).strip()
        managed_purchase = str(r.get("Managed Purchase Rate", "")).strip()
        comm_rental = str(r.get("Commercial Rental Rate", "")).strip()
        comm_purchase = str(r.get("Commercial Purchase Rate", "")).strip()

        is_range = "-" in hcpcs_norm and not VALID_HCPCS.match(hcpcs_norm)
        expanded_codes = [hcpcs_norm]
        range_start, range_end = "", ""
        if is_range:
            codes, rs, re_ = parse_hcpcs_range(hcpcs_raw)
            if codes:
                expanded_codes = codes
                range_start, range_end = rs, re_

        for code in expanded_codes:
            for mod in mods:
                rec = {
                    "hcpcs_original": hcpcs_raw,
                    "hcpcs_normalized": code,
                    "modifier": mod,
                    "modifier_raw": mod_raw,
                    "description": desc,
                    "billing_unit": billing_unit,
                    "managed_rental_raw": managed_rental,
                    "managed_purchase_raw": managed_purchase,
                    "commercial_rental_raw": comm_rental,
                    "commercial_purchase_raw": comm_purchase,
                    "managed_rental_num": safe_float(managed_rental),
                    "managed_purchase_num": safe_float(managed_purchase),
                    "commercial_rental_num": safe_float(comm_rental),
                    "commercial_purchase_num": safe_float(comm_purchase),
                    "expanded_from_range": hcpcs_raw if len(expanded_codes) > 1 else "",
                    "range_start": range_start,
                    "range_end": range_end,
                    "schedule": "PHCC_OR_CONTRACTED",
                    "state": "OR",
                    "source_row": idx + 2,
                }
                records.append(rec)
    return records


def load_phcc_wa(path):
    """Load PHCC Washington participating fee schedule with range expansion."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df = _clean_cols(df)
    log.info(f"  Loaded {len(df)} rows from PHCC_WA_PARTICIPATING.csv")
    records = []
    for idx, r in df.iterrows():
        hcpcs_raw = str(r.get("HCPCS", "")).strip()
        hcpcs_norm = normalize_hcpcs(hcpcs_raw)
        mod_raw = str(r.get("Modifier", "")).strip()

        mods = [m.strip().upper() for m in mod_raw.split("/") if m.strip()]
        if not mods:
            mods = [""]

        desc = str(r.get("Description", "")).strip()
        billing_unit = str(r.get("Billing Unit", "")).strip()
        rental_raw = str(r.get("Rental Rate", "")).strip()
        purchase_raw = str(r.get("Purchase Rate", "")).strip()

        is_range = "-" in hcpcs_norm and not VALID_HCPCS.match(hcpcs_norm)
        expanded_codes = [hcpcs_norm]
        range_start, range_end = "", ""
        if is_range:
            codes, rs, re_ = parse_hcpcs_range(hcpcs_raw)
            if codes:
                expanded_codes = codes
                range_start, range_end = rs, re_

        for code in expanded_codes:
            for mod in mods:
                rec = {
                    "hcpcs_original": hcpcs_raw,
                    "hcpcs_normalized": code,
                    "modifier": mod,
                    "modifier_raw": mod_raw,
                    "description": desc,
                    "billing_unit": billing_unit,
                    "rental_raw": rental_raw,
                    "purchase_raw": purchase_raw,
                    "rental_num": safe_float(rental_raw),
                    "purchase_num": safe_float(purchase_raw),
                    "expanded_from_range": hcpcs_raw if len(expanded_codes) > 1 else "",
                    "range_start": range_start,
                    "range_end": range_end,
                    "schedule": "PHCC_WA_PARTICIPATING",
                    "state": "WA",
                    "source_row": idx + 2,
                }
                records.append(rec)
    return records


def load_cms(path, nr_col, state):
    """Load CMS DMEPOS fee schedule."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df = _clean_cols(df)
    log.info(f"  Loaded {len(df)} rows from CMS_2026_Q1_{state}.csv")
    records = {}
    for _, r in df.iterrows():
        code = normalize_hcpcs(r.get("HCPCS", ""))
        mod = norm_mod(r.get("Mod", ""))
        rate = safe_float(r.get(nr_col, ""))
        desc = str(r.get("Short Description", "")).strip()
        key = f"{code}|{mod}"
        if key not in records:
            records[key] = {"rate": rate, "description": desc, "hcpcs": code, "mod": mod}
    return records


def load_oha(path):
    """Load OHA Medicaid fee schedule."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    df = _clean_cols(df)
    log.info(f"  Loaded {len(df)} rows from OHA_FFS_09_2025_RAW.csv")
    records = {}
    for _, r in df.iterrows():
        code = normalize_hcpcs(r.get("Procedure Code", ""))
        mod = norm_mod(r.get("Mod1", ""))
        rate = safe_float(r.get("Price", ""))
        desc = str(r.get("Description", "")).strip()
        key = f"{code}|{mod}"
        if key not in records:
            records[key] = {"rate": rate, "description": desc}
    return records


def load_hcpcs_descriptions(path) -> dict:
    """Load HCPCS reference for description enrichment."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False,
                     on_bad_lines="skip", encoding_errors="replace")
    df.columns = [c.strip() for c in df.columns]
    desc_map = {}
    for _, r in df.iterrows():
        code = str(r.get("HCPC", "")).strip().upper()
        desc = str(r.get("SHORT DESCRIPTION", "")).strip()
        if code and code not in desc_map:
            desc_map[code] = desc
    log.info(f"  Loaded {len(desc_map)} HCPCS descriptions")
    return desc_map


# ════════════════════════════════════════════════════════════════════════
# MATCHING ENGINE
# ════════════════════════════════════════════════════════════════════════

def build_phcc_lookup(records: list) -> dict:
    """Build multi-level lookup: key=hcpcs_normalized → list of records."""
    lk = {}
    for rec in records:
        code = rec["hcpcs_normalized"]
        if code not in lk:
            lk[code] = []
        lk[code].append(rec)
    return lk


def pick_rate_for_modifier(mod: str, rec: dict, schedule: str) -> tuple:
    """
    Given a modifier (NU/RR/blank) and a PHCC record, pick the best rate.
    Returns (rate_num, rate_raw, note_type, note_detail, rate_context).
    """
    if schedule == "PHCC_OR_CONTRACTED":
        if mod == "RR":
            candidates = [
                ("managed_rental_num", "managed_rental_raw", "Managed Rental"),
                ("commercial_rental_num", "commercial_rental_raw", "Commercial Rental"),
            ]
        elif mod == "NU":
            candidates = [
                ("managed_purchase_num", "managed_purchase_raw", "Managed Purchase"),
                ("commercial_purchase_num", "commercial_purchase_raw", "Commercial Purchase"),
            ]
        else:
            candidates = [
                ("managed_purchase_num", "managed_purchase_raw", "Managed Purchase"),
                ("commercial_purchase_num", "commercial_purchase_raw", "Commercial Purchase"),
                ("managed_rental_num", "managed_rental_raw", "Managed Rental"),
                ("commercial_rental_num", "commercial_rental_raw", "Commercial Rental"),
            ]
        for num_key, raw_key, ctx in candidates:
            val = rec.get(num_key, np.nan)
            if not np.isnan(val) and val > 0:
                return val, rec.get(raw_key, ""), "", "", ctx
        for num_key, raw_key, ctx in candidates:
            raw = rec.get(raw_key, "")
            if raw:
                nt, nd = classify_pricing_note(raw)
                return np.nan, raw, nt, nd, ctx
        return np.nan, "", "", "", ""

    elif schedule == "PHCC_WA_PARTICIPATING":
        if mod == "RR":
            candidates = [("rental_num", "rental_raw", "Rental"),
                          ("purchase_num", "purchase_raw", "Purchase")]
        elif mod == "NU":
            candidates = [("purchase_num", "purchase_raw", "Purchase"),
                          ("rental_num", "rental_raw", "Rental")]
        else:
            candidates = [("purchase_num", "purchase_raw", "Purchase"),
                          ("rental_num", "rental_raw", "Rental")]
        for num_key, raw_key, ctx in candidates:
            val = rec.get(num_key, np.nan)
            if not np.isnan(val) and val > 0:
                return val, rec.get(raw_key, ""), "", "", ctx
        for num_key, raw_key, ctx in candidates:
            raw = rec.get(raw_key, "")
            if raw:
                nt, nd = classify_pricing_note(raw)
                return np.nan, raw, nt, nd, ctx
        return np.nan, "", "", "", ""
    return np.nan, "", "", "", ""


def match_proposed_to_current(proposed: dict, phcc_lk: dict, schedule: str) -> dict:
    """
    Match one proposed row to PHCC current schedule.
    3-tier strategy: exact mod → proposed-mod-current-blank → HCPCS-only fallback.
    """
    code = proposed["hcpcs_normalized"]
    mod1 = proposed["modifier_1"]

    result = {
        "current_schedule_type": schedule,
        "current_rate_raw": "",
        "current_rate_numeric": np.nan,
        "current_rate_note_type": "",
        "current_rate_note_detail": "",
        "description_current": "",
        "billing_unit_current": "",
        "modifier_current": "",
        "modifier_match_strategy": "",
        "match_confidence": "",
        "match_method": "",
        "research_notes": "",
    }

    candidates = phcc_lk.get(code, [])
    if not candidates:
        result["match_method"] = "NO_MATCH"
        return result

    # Strategy 1: Exact modifier match
    for cand in candidates:
        cand_mod = cand.get("modifier", "")
        if cand_mod == mod1 or (mod1 == "" and cand_mod == ""):
            rate_num, rate_raw, nt, nd, ctx = pick_rate_for_modifier(
                cand_mod if cand_mod else mod1, cand, schedule
            )
            result.update({
                "current_rate_raw": rate_raw,
                "current_rate_numeric": rate_num,
                "current_rate_note_type": nt,
                "current_rate_note_detail": nd,
                "description_current": cand.get("description", ""),
                "billing_unit_current": cand.get("billing_unit", ""),
                "modifier_current": cand.get("modifier_raw", cand.get("modifier", "")),
                "modifier_match_strategy": f"EXACT_MOD_{ctx}",
                "match_confidence": "HIGH",
                "match_method": "EXACT_HCPCS_MOD",
            })
            return result

    # Strategy 2: Integra has modifier but PHCC entry is blank
    blank_candidates = [c for c in candidates if c.get("modifier", "") == ""]
    if blank_candidates and mod1:
        cand = blank_candidates[0]
        rate_num, rate_raw, nt, nd, ctx = pick_rate_for_modifier(mod1, cand, schedule)
        result.update({
            "current_rate_raw": rate_raw,
            "current_rate_numeric": rate_num,
            "current_rate_note_type": nt,
            "current_rate_note_detail": nd,
            "description_current": cand.get("description", ""),
            "billing_unit_current": cand.get("billing_unit", ""),
            "modifier_current": cand.get("modifier_raw", ""),
            "modifier_match_strategy": f"PROPOSED_MOD_CURRENT_BLANK_{ctx}",
            "match_confidence": "MEDIUM",
            "match_method": "HCPCS_ONLY_FALLBACK",
        })
        return result

    # Strategy 3: HCPCS-only fallback (use first candidate)
    cand = candidates[0]
    cand_mod = cand.get("modifier", "")
    rate_num, rate_raw, nt, nd, ctx = pick_rate_for_modifier(
        cand_mod if cand_mod else mod1, cand, schedule
    )
    other_mods = ", ".join(set(c.get("modifier", "") for c in candidates))
    result.update({
        "current_rate_raw": rate_raw,
        "current_rate_numeric": rate_num,
        "current_rate_note_type": nt,
        "current_rate_note_detail": nd,
        "description_current": cand.get("description", ""),
        "billing_unit_current": cand.get("billing_unit", ""),
        "modifier_current": cand.get("modifier_raw", cand.get("modifier", "")),
        "modifier_match_strategy": f"HCPCS_FALLBACK_{ctx}",
        "match_confidence": "LOW",
        "match_method": "HCPCS_ONLY_FALLBACK",
        "research_notes": f"Multiple candidates with mods: [{other_mods}]",
    })
    return result


# ════════════════════════════════════════════════════════════════════════
# COMPARISON LOGIC
# ════════════════════════════════════════════════════════════════════════

def compare_rates(proposed_num, current_num):
    """Compare proposed vs current. Returns (status, diff, pct)."""
    if np.isnan(proposed_num) or np.isnan(current_num):
        return "NOT_COMPARABLE", np.nan, np.nan
    if current_num == 0 and proposed_num == 0:
        return "EQUAL", 0.0, 0.0
    if current_num == 0:
        return "NOT_COMPARABLE", np.nan, np.nan
    diff = proposed_num - current_num
    pct = (diff / current_num) * 100
    if abs(diff) < 0.005:
        return "EQUAL", 0.0, 0.0
    return ("HIGHER" if diff > 0 else "LOWER"), round(diff, 4), round(pct, 2)


def compare_to_benchmark(proposed_num, bench_num, bench_label):
    """Compare proposed vs benchmark floor."""
    if np.isnan(proposed_num) or np.isnan(bench_num):
        if np.isnan(bench_num):
            return "MISSING_BENCHMARK", np.nan, np.nan
        return "NOT_COMPARABLE", np.nan, np.nan
    if bench_num == 0:
        return "NOT_COMPARABLE", np.nan, np.nan
    diff = proposed_num - bench_num
    pct = (diff / bench_num) * 100
    if abs(diff) < 0.005:
        return "EQUAL_TO_BENCHMARK", 0.0, 0.0
    return ("ABOVE_BENCHMARK" if diff > 0 else "BELOW_BENCHMARK"), round(diff, 4), round(pct, 2)


def lookup_benchmark(code, mod, cms_lk, oha_lk, payer_group, state):
    """Return (bench_rate, bench_raw, bench_source, bench_note_type, bench_note_detail)."""
    if payer_group == "Medicare":
        for key in [f"{code}|{mod}", f"{code}|"]:
            if key in cms_lk:
                rec = cms_lk[key]
                return rec["rate"], str(rec["rate"]), f"CMS_2026_Q1_{state}", "", ""
        return np.nan, "", f"CMS_2026_Q1_{state}", "", "NOT_FOUND_IN_CMS"
    elif payer_group == "Medicaid":
        if state == "WA":
            return np.nan, "", "MISSING", "", "WA_MEDICAID_NOT_PROVIDED"
        for key in [f"{code}|{mod}", f"{code}|NU", f"{code}|"]:
            if key in oha_lk:
                rec = oha_lk[key]
                return rec["rate"], str(rec["rate"]), "OHA_FFS_09_2025", "", ""
        return np.nan, "", "OHA_FFS_09_2025", "", "NOT_FOUND_IN_OHA"
    return np.nan, "", "", "", "NOT_APPLICABLE"


# ════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ════════════════════════════════════════════════════════════════════════

def _process_one_proposed(prop, phcc_or_lk, phcc_wa_lk,
                          cms_or_lk, cms_wa_lk, oha_lk, desc_map,
                          comparison_rows, review_rows):
    """Match one proposed row against both OR and WA schedules and emit comparison rows."""
    code = prop["hcpcs_normalized"]
    mod1 = prop["modifier_1"]
    payer = prop["payer_group"]

    state_configs = [
        ("OR", phcc_or_lk, cms_or_lk, oha_lk),
        ("WA", phcc_wa_lk, cms_wa_lk, oha_lk),
    ]

    for state, phcc_lk, cms_lk, oha_ref in state_configs:
        schedule = "PHCC_OR_CONTRACTED" if state == "OR" else "PHCC_WA_PARTICIPATING"
        match = match_proposed_to_current(prop, phcc_lk, schedule)

        proposed_num = prop["proposed_rate_numeric"]
        current_num = match["current_rate_numeric"]

        if prop["proposed_rate_note_type"] and match["current_rate_note_type"]:
            comp_status = "NOT_COMPARABLE"
            comp_diff, comp_pct = np.nan, np.nan
        elif match["match_method"] == "NO_MATCH":
            comp_status = "MISSING_CURRENT"
            comp_diff, comp_pct = np.nan, np.nan
        else:
            comp_status, comp_diff, comp_pct = compare_rates(proposed_num, current_num)

        needs_bench = (comp_status == "LOWER" and payer in ("Medicare", "Medicaid"))
        bench_rate, bench_raw, bench_src, bench_nt, bench_nd = (np.nan, "", "", "", "")
        bench_status, bench_diff, bench_pct = "NOT_APPLICABLE", np.nan, np.nan

        if needs_bench:
            bench_rate, bench_raw, bench_src, bench_nt, bench_nd = lookup_benchmark(
                code, mod1, cms_lk, oha_ref, payer, state
            )
            if bench_nd and "NOT_FOUND" in bench_nd:
                bench_status = "MISSING_BENCHMARK"
            elif bench_nd and "NOT_PROVIDED" in bench_nd:
                bench_status = "MISSING_BENCHMARK"
            else:
                bench_status, bench_diff, bench_pct = compare_to_benchmark(
                    proposed_num, bench_rate, bench_src
                )

        review_required = False
        review_reason = []
        if not prop["hcpcs_is_valid"]:
            review_required = True
            review_reason.append(f"INVALID_HCPCS:{prop['hcpcs_validation_issue']}")
        if match["match_method"] == "NO_MATCH":
            review_required = True
            review_reason.append("MISSING_CURRENT")
        if match["match_confidence"] == "LOW":
            review_required = True
            review_reason.append("LOW_CONFIDENCE_MATCH")
        if prop["proposed_rate_note_type"]:
            review_required = True
            review_reason.append(f"NON_NUMERIC_PROPOSED:{prop['proposed_rate_note_type']}")
        if match["current_rate_note_type"]:
            review_required = True
            review_reason.append(f"NON_NUMERIC_CURRENT:{match['current_rate_note_type']}")
        if bench_status == "MISSING_BENCHMARK" and needs_bench:
            review_required = True
            review_reason.append("MISSING_BENCHMARK")

        desc_proposed = desc_map.get(code, "")

        row = {
            "source_file": prop["source_file"],
            "source_row": prop["source_row"],
            "state": state,
            "payer_group": payer,
            "current_schedule_type": schedule,
            "hcpcs_original": prop["hcpcs_original"],
            "hcpcs_normalized": code,
            "hcpcs_is_valid": prop["hcpcs_is_valid"],
            "hcpcs_validation_issue": prop["hcpcs_validation_issue"],
            "expanded_from_range": prop.get("expanded_from_range", ""),
            "range_start": prop.get("range_start", ""),
            "range_end": prop.get("range_end", ""),
            "modifier_1": mod1,
            "modifier_2": prop["modifier_2"],
            "modifier_current": match["modifier_current"],
            "modifier_match_strategy": match["modifier_match_strategy"],
            "description_proposed": desc_proposed,
            "description_current": match["description_current"],
            "billing_unit_current": match["billing_unit_current"],
            "proposed_rate_raw": prop["proposed_rate_raw"],
            "proposed_rate_numeric": proposed_num,
            "proposed_rate_note_type": prop["proposed_rate_note_type"],
            "proposed_rate_note_detail": prop["proposed_rate_note_detail"],
            "current_rate_raw": match["current_rate_raw"],
            "current_rate_numeric": current_num,
            "current_rate_note_type": match["current_rate_note_type"],
            "current_rate_note_detail": match["current_rate_note_detail"],
            "comparison_status_current": comp_status,
            "comparison_amount_current": comp_diff,
            "comparison_pct_current": comp_pct,
            "needs_benchmark_check": needs_bench,
            "benchmark_source": bench_src,
            "benchmark_rate_raw": bench_raw,
            "benchmark_rate_numeric": bench_rate,
            "benchmark_rate_note_type": bench_nt,
            "benchmark_rate_note_detail": bench_nd,
            "comparison_status_benchmark": bench_status,
            "comparison_amount_benchmark": bench_diff,
            "comparison_pct_benchmark": bench_pct,
            "review_required": review_required,
            "review_reason": "; ".join(review_reason),
            "research_notes": match["research_notes"],
            "match_confidence": match["match_confidence"],
            "match_method": match["match_method"],
        }
        comparison_rows.append(row)

        if review_required:
            review_rows.append({
                "source_file": prop["source_file"],
                "source_row": prop["source_row"],
                "payer_group": payer,
                "state": state,
                "hcpcs_original": prop["hcpcs_original"],
                "hcpcs_normalized": code,
                "review_reason": "; ".join(review_reason),
                "review_required": True,
                "proposed_rate_raw": prop["proposed_rate_raw"],
                "current_rate_raw": match["current_rate_raw"],
                "match_method": match["match_method"],
                "match_confidence": match["match_confidence"],
            })


def run_analysis():
    log.info("=" * 70)
    log.info("LOADING DATA FILES")
    log.info("=" * 70)

    # ── Load Integra proposed schedules ──────────────────────────────
    integra_rows = []
    for key, rate_col, payer in [
        ("integra_commercial", "Commercial", "Commercial"),
        ("integra_medicare",   "Medicare",   "Medicare"),
        ("integra_medicaid",   "Medicaid",   "Medicaid"),
        ("integra_aso",        "ASO/Commercial", "ASO"),
    ]:
        integra_rows.extend(load_integra(FILES[key], rate_col, payer, key))
    log.info(f"  Total proposed rows: {len(integra_rows)}")

    # ── Load PHCC current schedules ──────────────────────────────────
    phcc_or_records = load_phcc_or(FILES["phcc_or"])
    phcc_wa_records = load_phcc_wa(FILES["phcc_wa"])
    log.info(f"  PHCC_OR expanded records: {len(phcc_or_records)}")
    log.info(f"  PHCC_WA expanded records: {len(phcc_wa_records)}")

    phcc_or_lk = build_phcc_lookup(phcc_or_records)
    phcc_wa_lk = build_phcc_lookup(phcc_wa_records)
    log.info(f"  PHCC_OR unique HCPCS: {len(phcc_or_lk)}")
    log.info(f"  PHCC_WA unique HCPCS: {len(phcc_wa_lk)}")

    # ── Load benchmarks ──────────────────────────────────────────────
    cms_or_lk = load_cms(FILES["cms_or"], "OR (NR)", "OR")
    cms_wa_lk = load_cms(FILES["cms_wa"], "WA (NR)", "WA")
    oha_lk = load_oha(FILES["oha"])
    desc_map = load_hcpcs_descriptions(FILES["hcpcs_ref"])

    # ── Process each proposed row ────────────────────────────────────
    log.info("=" * 70)
    log.info("MATCHING & COMPARING")
    log.info("=" * 70)

    comparison_rows = []
    audit_rows = []
    range_expansion_rows = []
    review_rows = []

    for prop in integra_rows:
        code = prop["hcpcs_normalized"]
        is_valid = prop["hcpcs_is_valid"]
        issue = prop["hcpcs_validation_issue"]

        if not is_valid:
            audit_rows.append({
                "source_file": prop["source_file"],
                "row_number": prop["source_row"],
                "payer_group": prop["payer_group"],
                "hcpcs_original": prop["hcpcs_original"],
                "hcpcs_normalized": code,
                "issue_type": issue,
                "issue_detail": f"Failed validation: {issue}",
                "contains_illegal_chars": bool(re.search(r"[^A-Z0-9]", code)),
                "suggested_manual_review": True,
            })
            if issue == "RANGE":
                codes, rs, re_ = parse_hcpcs_range(prop["hcpcs_original"])
                if codes:
                    for exp_code in codes:
                        range_expansion_rows.append({
                            "source_file": prop["source_file"],
                            "source_row": prop["source_row"],
                            "hcpcs_original": prop["hcpcs_original"],
                            "hcpcs_expanded": exp_code,
                            "range_start": rs,
                            "range_end": re_,
                        })
                        exp_prop = prop.copy()
                        exp_prop["hcpcs_normalized"] = exp_code
                        exp_prop["hcpcs_is_valid"] = True
                        exp_prop["expanded_from_range"] = prop["hcpcs_original"]
                        exp_prop["range_start"] = rs
                        exp_prop["range_end"] = re_
                        _process_one_proposed(
                            exp_prop, phcc_or_lk, phcc_wa_lk,
                            cms_or_lk, cms_wa_lk, oha_lk, desc_map,
                            comparison_rows, review_rows
                        )
                    continue
                else:
                    review_rows.append({
                        "source_file": prop["source_file"],
                        "source_row": prop["source_row"],
                        "payer_group": prop["payer_group"],
                        "hcpcs_original": prop["hcpcs_original"],
                        "hcpcs_normalized": code,
                        "review_reason": f"Unparseable HCPCS range: {issue}",
                        "review_required": True,
                    })
                    continue

            review_rows.append({
                "source_file": prop["source_file"],
                "source_row": prop["source_row"],
                "payer_group": prop["payer_group"],
                "hcpcs_original": prop["hcpcs_original"],
                "hcpcs_normalized": code,
                "review_reason": f"Invalid HCPCS: {issue}",
                "review_required": True,
            })
            _process_one_proposed(
                prop, phcc_or_lk, phcc_wa_lk,
                cms_or_lk, cms_wa_lk, oha_lk, desc_map,
                comparison_rows, review_rows
            )
            continue

        _process_one_proposed(
            prop, phcc_or_lk, phcc_wa_lk,
            cms_or_lk, cms_wa_lk, oha_lk, desc_map,
            comparison_rows, review_rows
        )

    # ── Build DataFrames ─────────────────────────────────────────────
    master = pd.DataFrame(comparison_rows)
    audit_df = pd.DataFrame(audit_rows) if audit_rows else pd.DataFrame()
    range_df = pd.DataFrame(range_expansion_rows) if range_expansion_rows else pd.DataFrame()
    review_df = pd.DataFrame(review_rows) if review_rows else pd.DataFrame()

    # ── Summary ──────────────────────────────────────────────────────
    summary_rows = []
    if len(master) > 0:
        for (payer, state), grp in master.groupby(["payer_group", "state"]):
            vc = grp["comparison_status_current"].value_counts()
            bc = grp["comparison_status_benchmark"].value_counts() if "comparison_status_benchmark" in grp.columns else pd.Series(dtype=int)
            summary_rows.append({
                "payer_group": payer,
                "state": state,
                "total_rows": len(grp),
                "matched": int((grp["match_method"] != "NO_MATCH").sum()),
                "unmatched": int((grp["match_method"] == "NO_MATCH").sum()),
                "HIGHER": int(vc.get("HIGHER", 0)),
                "LOWER": int(vc.get("LOWER", 0)),
                "EQUAL": int(vc.get("EQUAL", 0)),
                "NOT_COMPARABLE": int(vc.get("NOT_COMPARABLE", 0)),
                "MISSING_CURRENT": int(vc.get("MISSING_CURRENT", 0)),
                "lower_above_benchmark": int(bc.get("ABOVE_BENCHMARK", 0)),
                "lower_below_benchmark": int(bc.get("BELOW_BENCHMARK", 0)),
                "non_numeric_proposed": int(grp["proposed_rate_note_type"].astype(bool).sum()),
            })
    summary_df = pd.DataFrame(summary_rows)

    # ── Log summary ──────────────────────────────────────────────────
    log.info("=" * 70)
    log.info("RESULTS SUMMARY")
    log.info("=" * 70)
    log.info(f"  Total comparison rows: {len(master)}")
    log.info(f"  HCPCS audit issues:    {len(audit_df)}")
    log.info(f"  Range expansions:      {len(range_df)}")
    log.info(f"  Review queue items:    {len(review_df)}")
    if len(master) > 0:
        vc = master["comparison_status_current"].value_counts()
        for status in ["HIGHER", "LOWER", "EQUAL", "NOT_COMPARABLE", "MISSING_CURRENT"]:
            log.info(f"    {status}: {vc.get(status, 0)}")

    # ── Write outputs ────────────────────────────────────────────────
    _write_outputs(master, audit_df, range_df, review_df, summary_df)
    log.info(f"\nAll outputs written to: {OUT_DIR}")
    return master, audit_df, range_df, review_df, summary_df


# ════════════════════════════════════════════════════════════════════════
# OUTPUT WRITER
# ════════════════════════════════════════════════════════════════════════

def _write_outputs(master, audit_df, range_df, review_df, summary_df):
    """Write CSV + formatted Excel."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # CSV outputs
    master.to_csv(OUT_DIR / "fee_schedule_comparison_master.csv", index=False)
    if len(audit_df): audit_df.to_csv(OUT_DIR / "hcpcs_audit.csv", index=False)
    if len(range_df): range_df.to_csv(OUT_DIR / "hcpcs_range_expansion_audit.csv", index=False)
    if len(review_df): review_df.to_csv(OUT_DIR / "fee_schedule_review_queue.csv", index=False)
    summary_df.to_csv(OUT_DIR / "comparison_summary.csv", index=False)
    log.info("  CSV files written.")

    # Excel workbook
    xlsx_path = OUT_DIR / "fee_schedule_comparison_master.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        _write_executive_summary(writer, master, summary_df)
        summary_df.to_excel(writer, sheet_name="Summary Counts", index=False)
        master.to_excel(writer, sheet_name="Full Comparison", index=False)

        if len(master) > 0:
            lower = master[master["comparison_status_current"] == "LOWER"]
            if len(lower): lower.to_excel(writer, sheet_name="Proposed LOWER", index=False)

            higher = master[master["comparison_status_current"] == "HIGHER"]
            if len(higher): higher.to_excel(writer, sheet_name="Proposed HIGHER", index=False)

            below_bench = master[master["comparison_status_benchmark"] == "BELOW_BENCHMARK"]
            if len(below_bench): below_bench.to_excel(writer, sheet_name="Below Benchmark", index=False)

        if len(audit_df): audit_df.to_excel(writer, sheet_name="HCPCS Audit", index=False)
        if len(review_df): review_df.to_excel(writer, sheet_name="Review Queue", index=False)

        # ── Apply formatting ─────────────────────────────────────────
        GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        GRAY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ORANGE = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
        THIN_BORDER = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        verdict_keywords = {
            "HIGHER": GREEN, "ABOVE_BENCHMARK": GREEN,
            "LOWER": RED, "BELOW_BENCHMARK": RED,
            "EQUAL": YELLOW, "EQUAL_TO_BENCHMARK": YELLOW,
            "NOT_COMPARABLE": GRAY, "MISSING_CURRENT": ORANGE,
            "MISSING_BENCHMARK": ORANGE, "NOT_APPLICABLE": GRAY,
        }

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
                ws.column_dimensions[col_letter].width = min(max_len + 3, 35)

            for cell in ws[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = THIN_BORDER

            ws.freeze_panes = "A2"

            if sheet_name in ("Executive Summary", "Summary Counts"):
                continue

            verdict_col_indices = []
            for idx, cell in enumerate(ws[1], 1):
                val = str(cell.value or "")
                if "comparison_status" in val or val in ("Verdict",):
                    verdict_col_indices.append(idx)

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx in verdict_col_indices:
                    cell = row[col_idx - 1]
                    v = str(cell.value or "").upper()
                    for keyword, fill in verdict_keywords.items():
                        if keyword in v:
                            cell.fill = fill
                            break

            for idx, cell in enumerate(ws[1], 1):
                val = str(cell.value or "")
                if "pct" in val.lower() or "%" in val:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                            min_col=idx, max_col=idx):
                        for c in row:
                            if isinstance(c.value, (int, float)) and not (isinstance(c.value, float) and np.isnan(c.value)):
                                c.number_format = '0.0"%"'

    log.info(f"  Excel written to: {xlsx_path}")


def _write_executive_summary(writer, master, summary_df):
    """Concise executive summary sheet for quick decision-making."""
    if len(master) == 0:
        pd.DataFrame({"Note": ["No comparison data"]}).to_excel(
            writer, sheet_name="Executive Summary", index=False)
        return

    exec_rows = []
    for payer in ["Commercial", "ASO", "Medicare", "Medicaid"]:
        payer_data = master[master["payer_group"] == payer]
        if len(payer_data) == 0:
            continue

        for state in ["OR", "WA"]:
            state_data = payer_data[payer_data["state"] == state]
            if len(state_data) == 0:
                continue

            total = len(state_data)
            matched = (state_data["match_method"] != "NO_MATCH").sum()
            vc = state_data["comparison_status_current"].value_counts()
            higher_n = vc.get("HIGHER", 0)
            lower_n = vc.get("LOWER", 0)
            equal_n = vc.get("EQUAL", 0)
            not_comp = vc.get("NOT_COMPARABLE", 0)
            missing = vc.get("MISSING_CURRENT", 0)

            comparable = state_data[state_data["comparison_status_current"].isin(["HIGHER", "LOWER", "EQUAL"])]
            avg_pct = comparable["comparison_pct_current"].mean() if len(comparable) > 0 else np.nan

            lower_items = state_data[state_data["comparison_status_current"] == "LOWER"]
            total_lower_impact = lower_items["comparison_amount_current"].sum() if len(lower_items) > 0 else 0

            bc = state_data["comparison_status_benchmark"].value_counts()
            below_bench = bc.get("BELOW_BENCHMARK", 0)
            above_bench = bc.get("ABOVE_BENCHMARK", 0)

            exec_rows.append({
                "Payer": payer,
                "State": state,
                "Total Codes": total,
                "Matched to PHCC": int(matched),
                "HIGHER than PHCC": int(higher_n),
                "LOWER than PHCC": int(lower_n),
                "EQUAL to PHCC": int(equal_n),
                "Not Comparable": int(not_comp),
                "Missing PHCC Rate": int(missing),
                "Avg % Diff (comparable)": round(avg_pct, 1) if not np.isnan(avg_pct) else "N/A",
                "Total $ Impact (LOWER)": round(total_lower_impact, 2),
                "Below Public Benchmark": int(below_bench),
                "Above Public Benchmark": int(above_bench),
            })

    exec_df = pd.DataFrame(exec_rows)
    exec_df.to_excel(writer, sheet_name="Executive Summary", index=False)


if __name__ == "__main__":
    run_analysis()
